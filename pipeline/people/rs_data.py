"""rs_data.py — 태린이 '주도주찾기' 시트에서 종목별 RS(모멘텀) 로드.

유동성체크 엑셀의 '주도주찾기' 시트 = 종목별 1개월/3개월 수익률(주도주 스크린).
funnel의 '주도(RS)' 축에 쓴다. 자체 완결(taerini ingest 파이프라인 미변경).

한계: 시총 상위 600 블록(1개월+3개월 둘 다)만 파싱. ETF 소라티노는 별도 파일(미도착).
"""
import glob
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).parent.parent.parent
_GLOB = str(_ROOT / "raw" / "매일 엑셀넣을것" / "유동성*.xlsm")
_SHEET = "주도주찾기"


def _find_file():
    files = sorted(glob.glob(_GLOB))
    return files[-1] if files else None


_CACHE = {}  # path -> (mtime, result). 30MB xlsm 반복 파싱 방지.


def load_rs(path=None) -> dict:
    """{종목명: {'m1': 1개월수익률, 'm3': 3개월수익률, 'sector': 업종}}.
    파일/시트 없으면 {} (호출측이 graceful 처리). mtime 기반 캐시."""
    path = path or _find_file()
    if not path or not Path(path).exists():
        return {}
    mtime = Path(path).stat().st_mtime
    cached = _CACHE.get(path)
    if cached and cached[0] == mtime:
        return cached[1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if _SHEET not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb[_SHEET]
    rows = list(ws.iter_rows(min_row=1, max_row=700, values_only=True))
    wb.close()

    # 헤더 행 찾기: 'Name'과 '3개월'이 함께 있는 행
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c) if c is not None else "" for c in row[:6]]
        if any(c == "Name" for c in cells) and any("3개월" in c for c in cells):
            header_idx = i
            break
    if header_idx is None:
        return {}
    header = [str(c) if c is not None else "" for c in rows[header_idx][:6]]
    col = {}
    for j, h in enumerate(header):
        if h == "Name": col["name"] = j
        elif "1개월" in h: col["m1"] = j
        elif "3개월" in h: col["m3"] = j
        elif "업종" in h: col["sector"] = j
    if "name" not in col or "m3" not in col:
        return {}

    out = {}
    for row in rows[header_idx + 1:]:
        name = row[col["name"]] if col["name"] < len(row) else None
        if not name or not str(name).strip():
            break  # 블록 끝(빈 행)
        name = str(name).strip()

        def _num(key):
            j = col.get(key)
            if j is None or j >= len(row):
                return None
            v = row[j]
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        out[name] = {
            "m1": _num("m1"),
            "m3": _num("m3"),
            "sector": (str(row[col["sector"]]).strip()
                       if col.get("sector") is not None and col["sector"] < len(row)
                       and row[col["sector"]] else None),
        }
    _CACHE[path] = (mtime, out)
    return out


if __name__ == "__main__":
    import sys, io
    if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    rs = load_rs()
    print(f"주도주찾기 RS 로드: {len(rs)}종목")
    for name, v in list(rs.items())[:8]:
        print(f"  {name:12s} 1M {v['m1']}  3M {v['m3']}  {v['sector']}")

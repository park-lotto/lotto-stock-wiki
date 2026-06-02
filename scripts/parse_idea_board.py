"""
태린이아빠 투자아이디어 엑셀 파서 + 위키 저장 + HTML 이벤트 보드
"""
import openpyxl
import sys
import io
from pathlib import Path
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "raw/매일 엑셀넣을것/투자아이디어정리0602.xlsx"
WIKI_OUT = ROOT / "wiki/외부인사이트/태린이아빠_투자아이디어.md"
HTML_OUT = ROOT / "out/idea_board.html"

TODAY = date(2026, 6, 2)

# 분기 → 우선순위 정렬용
QUARTER_ORDER = {
    "2025 4Q": 0, "2026 1Q": 1, "2026 2Q": 2,
    "2026 3Q": 3, "2026 4Q": 4,
    "단기": 5, "중기": 6, "장기": 7,
}

# 현재 분기 = 2026 2Q
CURRENT_Q = "2026 2Q"

SKIP_SHEETS = {"정기변경", "Sheet1"}

def parse_excel():
    wb = openpyxl.load_workbook(str(XLSX), read_only=True, data_only=True)

    # 정기변경 시트 별도 파싱
    msci = parse_msci(wb["정기변경"])

    # 종목 아이디어 파싱
    sectors = {}
    for sh in wb.sheetnames:
        if sh in SKIP_SHEETS:
            continue
        rows = list(wb[sh].iter_rows(values_only=True))
        parsed = parse_sheet(sh, rows)
        if parsed:
            sectors[sh] = parsed

    return sectors, msci


def parse_msci(ws):
    events = []
    for row in ws.iter_rows(values_only=True):
        if not row[1]:
            continue
        v = str(row[1]).strip()
        if v in ("분기", "MSCI", "구분", "심사기간", " ", "수급이 계속 발생할 수 있으면 리밸런싱이후도 좋다", "3달전 매수 1달전 매도", "코스피 200"):
            continue
        if row[2] and "일" in str(row[2]):
            events.append({
                "name": f"MSCI {v}",
                "announce": str(row[2]) if row[2] else "",
                "rebal": str(row[3]) if row[3] else "",
                "reflect": str(row[4]) if row[4] else "",
            })
    return events


def parse_sheet(sector, rows):
    """한 시트에서 {종목: [{분기, 핵심모멘텀, 상세}]} 파싱"""
    stocks = {}
    current_stock = None

    for row in rows:
        # 유효 셀 수 계산
        vals = [str(v).strip() if v is not None else "" for v in row]
        non_empty = [v for v in vals if v]

        if not non_empty:
            continue

        col1 = vals[1] if len(vals) > 1 else ""
        col2 = vals[2] if len(vals) > 2 else ""
        col3 = vals[3] if len(vals) > 3 else ""

        # 헤더행 스킵
        if col1 in ("분기", "핵심 모멘텀", "심사기간"):
            continue
        if col1 in ("단기 카탈리스트", "단기 카탈리스트 요약", "중장기 인베스트먼트", "중장기 인베스트먼트 요약"):
            continue

        # 종목명 행: col1에 값 있고 col2·col3 비어있으면
        if col1 and not col2 and not col3:
            current_stock = col1
            if current_stock not in stocks:
                stocks[current_stock] = []
            continue

        # 데이터 행
        if current_stock and col1 and col2:
            q = normalize_quarter(col1)
            stocks[current_stock].append({
                "quarter": q,
                "momentum": col2,
                "detail": col3,
            })

    return stocks if stocks else None


def normalize_quarter(q):
    q = str(q).strip()
    mappings = {
        "25년 4Q": "2025 4Q", "26년 1Q": "2026 1Q",
        "26년 2Q": "2026 2Q", "26년 3Q": "2026 3Q", "26년 4Q": "2026 4Q",
        "2025 4Q": "2025 4Q", "2026 1Q": "2026 1Q",
        "2026 2Q": "2026 2Q", "2026 3Q": "2026 3Q", "2026 4Q": "2026 4Q",
    }
    if q in mappings:
        return mappings[q]
    for k in ["단기", "중기", "장기"]:
        if q.startswith(k):
            return k
    return q


# ─── WIKI MD 생성 ────────────────────────────────────────────────────────────

def build_wiki(sectors, msci):
    lines = [
        f"# 태린이아빠 투자아이디어 (업데이트: {TODAY})",
        "",
        "> 출처: `raw/매일 엑셀넣을것/투자아이디어정리0602.xlsx`  ",
        "> 분류: 외부인사이트 / 섹터별 분기 모멘텀 정리  ",
        "",
        "---",
        "",
        "## 정기 이벤트 캘린더",
        "",
        "### MSCI 정기변경",
        "| 구분 | 발표일 | 리밸런싱일 | 지수반영일 |",
        "|------|--------|----------|---------|",
    ]
    for e in msci:
        lines.append(f"| {e['name']} | {e['announce']} | {e['rebal']} | {e['reflect']} |")

    lines += ["", "---", "", "## 섹터별 아이디어", ""]

    for sector, stocks in sectors.items():
        lines.append(f"### {sector}")
        lines.append("")
        for stock, items in stocks.items():
            lines.append(f"#### {stock}")
            lines.append("")
            lines.append("| 분기 | 핵심 모멘텀 | 상세 |")
            lines.append("|------|-----------|------|")
            for item in sorted(items, key=lambda x: QUARTER_ORDER.get(x["quarter"], 99)):
                q = item["quarter"]
                badge = " 🔴" if q == CURRENT_Q else (" ✅" if q in ("2025 4Q", "2026 1Q") else "")
                lines.append(f"| **{q}**{badge} | {item['momentum']} | {item['detail']} |")
            lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


# ─── HTML 이벤트 보드 ─────────────────────────────────────────────────────────

def build_html(sectors, msci):
    # 이번 달(2Q) 이벤트 집계
    now_q_events = []
    next_q_events = []

    for sector, stocks in sectors.items():
        for stock, items in stocks.items():
            for item in items:
                q = item["quarter"]
                row = {"sector": sector, "stock": stock, "quarter": q,
                       "momentum": item["momentum"], "detail": item["detail"]}
                if q == "2026 2Q":
                    now_q_events.append(row)
                elif q == "2026 3Q":
                    next_q_events.append(row)

    # 섹터별 종목 수
    sector_counts = {s: len(v) for s, v in sectors.items()}

    now_rows = build_event_rows(now_q_events)
    next_rows = build_event_rows(next_q_events)
    msci_rows = "".join([
        f'<tr><td>{e["name"]}</td><td>{e["announce"]}</td><td>{e["rebal"]}</td><td>{e["reflect"]}</td></tr>'
        for e in msci
    ])

    # 전체 종목 × 분기 매트릭스
    all_stocks_html = build_all_stocks_matrix(sectors)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>투자아이디어 보드 — {TODAY}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Pretendard', 'Apple SD Gothic Neo', sans-serif; background: #0f1117; color: #e2e8f0; }}
  .header {{ background: linear-gradient(135deg, #1a1f35 0%, #0f1117 100%); padding: 32px 40px; border-bottom: 1px solid #2d3748; }}
  .header h1 {{ font-size: 28px; font-weight: 700; color: #f0f0f0; }}
  .header .sub {{ font-size: 14px; color: #718096; margin-top: 6px; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 32px 40px; }}
  .section {{ margin-bottom: 48px; }}
  .section-title {{ font-size: 20px; font-weight: 700; color: #f0f0f0; margin-bottom: 16px; display: flex; align-items: center; gap: 10px; }}
  .badge-now {{ background: #e53e3e; color: white; font-size: 11px; padding: 3px 8px; border-radius: 12px; }}
  .badge-next {{ background: #3182ce; color: white; font-size: 11px; padding: 3px 8px; border-radius: 12px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #1a202c; color: #a0aec0; font-weight: 600; text-align: left; padding: 10px 14px; border-bottom: 1px solid #2d3748; }}
  td {{ padding: 10px 14px; border-bottom: 1px solid #1a202c; vertical-align: top; }}
  tr:hover td {{ background: #1a202c; }}
  .sector-tag {{ display: inline-block; background: #2d3748; color: #90cdf4; font-size: 11px; padding: 2px 8px; border-radius: 4px; margin-bottom: 4px; }}
  .stock-name {{ font-weight: 700; color: #f6e05e; font-size: 14px; }}
  .momentum {{ color: #68d391; }}
  .detail {{ color: #a0aec0; font-size: 12px; margin-top: 4px; }}
  .grid-4 {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; }}
  .stat-card {{ background: #1a1f2e; border: 1px solid #2d3748; border-radius: 12px; padding: 20px; text-align: center; }}
  .stat-num {{ font-size: 32px; font-weight: 800; color: #f6e05e; }}
  .stat-label {{ font-size: 12px; color: #718096; margin-top: 4px; }}
  .msci-highlight {{ background: #1a2744; border: 1px solid #3182ce; border-radius: 8px; padding: 12px 16px; margin-bottom: 8px; display: flex; gap: 24px; align-items: center; }}
  .msci-label {{ font-size: 12px; color: #90cdf4; }}
  .msci-date {{ font-size: 14px; font-weight: 600; }}
  .matrix-wrap {{ overflow-x: auto; }}
  .matrix-wrap table td {{ min-width: 140px; }}
  .q-now {{ background: rgba(229, 62, 62, 0.1); border-left: 3px solid #e53e3e; }}
  .q-next {{ background: rgba(49, 130, 206, 0.08); }}
  .q-done {{ opacity: 0.5; }}
  .stock-row-header {{ background: #1a202c; }}
  .no-data {{ color: #4a5568; font-size: 12px; }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 투자아이디어 이벤트 보드</h1>
  <div class="sub">태린이아빠 분기별 모멘텀 정리 · 업데이트: {TODAY} · 현재 분기: 2026 2Q</div>
</div>
<div class="container">

  <!-- 요약 통계 -->
  <div class="section">
    <div class="grid-4">
      <div class="stat-card"><div class="stat-num">{sum(len(v) for v in sectors.values())}</div><div class="stat-label">총 종목 수</div></div>
      <div class="stat-card"><div class="stat-num">{len(sectors)}</div><div class="stat-label">섹터 수</div></div>
      <div class="stat-card"><div class="stat-num" style="color:#e53e3e">{len(now_q_events)}</div><div class="stat-label">2Q 이벤트 (현재)</div></div>
      <div class="stat-card"><div class="stat-num" style="color:#3182ce">{len(next_q_events)}</div><div class="stat-label">3Q 이벤트 (다음)</div></div>
    </div>
  </div>

  <!-- MSCI 정기변경 -->
  <div class="section">
    <div class="section-title">🗓️ MSCI 정기변경 일정</div>
    <table>
      <tr><th>구분</th><th>발표일</th><th>리밸런싱일</th><th>지수반영일</th></tr>
      {msci_rows}
    </table>
  </div>

  <!-- 이번 달 (2026 2Q) 이벤트 -->
  <div class="section">
    <div class="section-title">🔴 이번 분기 이벤트 (2026 2Q) <span class="badge-now">현재 진행 중</span></div>
    <table>
      <tr><th>섹터</th><th>종목</th><th>핵심 모멘텀</th><th>상세</th></tr>
      {now_rows}
    </table>
  </div>

  <!-- 다음 분기 (2026 3Q) -->
  <div class="section">
    <div class="section-title">🔵 다음 분기 이벤트 (2026 3Q) <span class="badge-next">준비 구간</span></div>
    <table>
      <tr><th>섹터</th><th>종목</th><th>핵심 모멘텀</th><th>상세</th></tr>
      {next_rows}
    </table>
  </div>

  <!-- 전체 종목 × 분기 매트릭스 -->
  <div class="section">
    <div class="section-title">📋 전체 종목 분기별 모멘텀 매트릭스</div>
    <div class="matrix-wrap">
      {all_stocks_html}
    </div>
  </div>

</div>
</body>
</html>"""
    return html


def build_event_rows(events):
    if not events:
        return '<tr><td colspan="4" class="no-data" style="text-align:center;padding:20px">데이터 없음</td></tr>'
    rows = []
    for e in sorted(events, key=lambda x: x["sector"]):
        rows.append(
            f'<tr>'
            f'<td><span class="sector-tag">{e["sector"]}</span></td>'
            f'<td><span class="stock-name">{e["stock"]}</span></td>'
            f'<td><span class="momentum">{e["momentum"]}</span></td>'
            f'<td><span class="detail">{e["detail"]}</span></td>'
            f'</tr>'
        )
    return "".join(rows)


def build_all_stocks_matrix(sectors):
    quarters = ["2025 4Q", "2026 1Q", "2026 2Q", "2026 3Q", "2026 4Q", "단기", "중기", "장기"]
    q_labels = {"2025 4Q": "25 4Q", "2026 1Q": "26 1Q", "2026 2Q": "26 2Q 🔴",
                "2026 3Q": "26 3Q", "2026 4Q": "26 4Q", "단기": "단기", "중기": "중기", "장기": "장기"}

    html_parts = []
    for sector, stocks in sectors.items():
        html_parts.append(f'<div style="margin-bottom:32px"><div style="font-size:16px;font-weight:700;color:#f6e05e;margin-bottom:12px">{sector}</div>')
        html_parts.append('<table><tr><th style="min-width:120px">종목</th>')
        for q in quarters:
            css = 'style="background:#1a2744"' if q == CURRENT_Q else ""
            html_parts.append(f'<th {css}>{q_labels[q]}</th>')
        html_parts.append('</tr>')

        for stock, items in stocks.items():
            q_map = {}
            for item in items:
                q = item["quarter"]
                if q not in q_map:
                    q_map[q] = []
                q_map[q].append(item["momentum"])

            html_parts.append(f'<tr><td style="font-weight:700;color:#f6e05e;white-space:nowrap">{stock}</td>')
            for q in quarters:
                if q == CURRENT_Q:
                    td_style = 'class="q-now"'
                elif q in ("2025 4Q", "2026 1Q"):
                    td_style = 'class="q-done"'
                else:
                    td_style = ''

                if q in q_map:
                    bullets = "<br>".join([f"· {m}" for m in q_map[q]])
                    html_parts.append(f'<td {td_style}><span style="font-size:12px;color:#68d391">{bullets}</span></td>')
                else:
                    html_parts.append(f'<td {td_style}><span class="no-data">—</span></td>')
            html_parts.append('</tr>')

        html_parts.append('</table></div>')

    return "".join(html_parts)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("파싱 중...")
    sectors, msci = parse_excel()

    total_stocks = sum(len(v) for v in sectors.values())
    total_sectors = len(sectors)
    print(f"  섹터 {total_sectors}개, 종목 {total_stocks}개 파싱 완료")

    # 위키 저장
    WIKI_OUT.parent.mkdir(parents=True, exist_ok=True)
    wiki_content = build_wiki(sectors, msci)
    WIKI_OUT.write_text(wiki_content, encoding="utf-8")
    print(f"  위키 저장: {WIKI_OUT}")

    # HTML 저장
    HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
    html_content = build_html(sectors, msci)
    HTML_OUT.write_text(html_content, encoding="utf-8")
    print(f"  HTML 저장: {HTML_OUT}")

    print("완료!")

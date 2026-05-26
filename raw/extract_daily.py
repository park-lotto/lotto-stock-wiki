# -*- coding: utf-8 -*-
"""
매일 Excel 파일 → openpyxl 추출 → Gemini 분석 → 변화 감지 → MD 저장

처리 대상 (6종):
  추정이익변경.xlsm   → Rating_Up/Down, TP_Up/Down  → 탑픽 #4·#3
  컨센서프쇼크.xlsx   → 서프라이즈/쇼크/컨센상향    → 탑픽 #3
  일정수주잔고.xlsx   → 이벤트·투경·수주잔고         → 탑픽 #7
  투자아이디어.xlsx   → 섹터별 분기 모멘텀           → 탑픽 #7
  액티브ETF.xlsx      → ETF 편입 종목 변화            → 섹터 흐름
  페어트레이딩.xlsx   → 스프레드 모니터               → 페어 알림

PS1이 처리하는 3종 (별도):
  유동성.xlsm         → extract_유동성.ps1 (완성)
  종목별오실레이터.xlsm → extract_oscillator.ps1
  업종별오실레이터.xlsm → 미작성

사용법:
  python raw/extract_daily.py             # 전체 처리
  python raw/extract_daily.py 추정이익    # 특정 유형만
  python raw/extract_daily.py --no-skip   # done 파일 무시하고 재처리
"""

import os
import sys
import pathlib
import time
from datetime import datetime, timedelta
from dotenv import load_dotenv
import openpyxl

from google import genai

sys.stdout.reconfigure(encoding="utf-8")

ROOT = pathlib.Path(__file__).parent.parent
load_dotenv(ROOT / ".env")

API_KEY = os.environ.get("GEMINI_API_KEY")
if not API_KEY:
    print("ERROR: .env에 GEMINI_API_KEY 없음")
    sys.exit(1)

client = genai.Client(api_key=API_KEY)
MODEL = "models/gemini-2.5-flash"

EXCEL_DIR = ROOT / "raw" / "매일 엑셀넣을것"
SUMMARY_DIR = ROOT / "raw" / "매일요약"
SUMMARY_DIR.mkdir(exist_ok=True)

TODAY = datetime.now().strftime("%Y%m%d")
TODAY_DISPLAY = f"{TODAY[:4]}-{TODAY[4:6]}-{TODAY[6:]}"
YESTERDAY = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
FORCE = "--no-skip" in sys.argv

# ── Excel 읽기 유틸 ───────────────────────────────────

def read_sheet_as_text(wb, sheet_name: str, max_rows: int = 100) -> str:
    """시트 데이터를 마크다운 테이블 형태 텍스트로 반환."""
    if sheet_name not in wb.sheetnames:
        return f"[{sheet_name} 시트 없음]"
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            rows.append(f"... ({ws.max_row - max_rows}행 더 있음)")
            break
        # None 셀은 빈 문자열로
        cells = [str(c) if c is not None else "" for c in row]
        if any(c for c in cells):  # 완전히 빈 행 제외
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def read_sheet_recent(wb, sheet_name: str, date_col: int = 0, days: int = 14, max_rows: int = 50) -> str:
    """날짜 컬럼 기준 최근 N일 행만 추출."""
    if sheet_name not in wb.sheetnames:
        return f"[{sheet_name} 시트 없음]"
    ws = wb[sheet_name]
    cutoff = datetime.now() - timedelta(days=days)
    rows = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = " | ".join([str(c) if c else "" for c in row])
            rows.append(header)
            continue
        if len(rows) > max_rows:
            break
        # 날짜 필터 시도
        try:
            cell_val = row[date_col]
            if hasattr(cell_val, 'date'):
                if cell_val < cutoff:
                    continue
            elif isinstance(cell_val, str):
                pass  # 문자열 날짜는 그냥 포함
        except Exception:
            pass
        cells = [str(c) if c is not None else "" for c in row]
        if any(c for c in cells):
            rows.append(" | ".join(cells))
    return "\n".join(rows)


def open_workbook(path: pathlib.Path):
    """openpyxl로 워크북 열기 (data_only=True로 계산된 값 읽기)."""
    try:
        return openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"    [경고] openpyxl 열기 실패: {e}")
        return None


# ── Gemini 텍스트 분석 ────────────────────────────────

def ask_gemini(prompt: str) -> str:
    response = client.models.generate_content(
        model=MODEL,
        contents=[prompt],
    )
    return response.text


# ── 변화 감지 ──────────────────────────────────────────

def load_yesterday_md(ftype: str) -> str:
    p = SUMMARY_DIR / f"{YESTERDAY}_{ftype}.md"
    return p.read_text(encoding="utf-8") if p.exists() else ""


def detect_changes(ftype: str, today_text: str) -> str:
    yesterday = load_yesterday_md(ftype)
    if not yesterday:
        return today_text  # 첫 실행 = 전체가 신규

    prompt = f"""아래 '어제 데이터'와 '오늘 데이터'를 비교해서
오늘 데이터에서 **새로 추가되거나 수치가 변화된 항목만** 추출하세요.
변화 없는 항목은 출력하지 마세요.
변화가 전혀 없으면 첫 줄에 "변화없음" 이라고만 쓰세요.

=== 어제 ({YESTERDAY}) ===
{yesterday[:4000]}

=== 오늘 ({TODAY_DISPLAY}) ===
{today_text[:4000]}

=== 변화 요약 ({TODAY_DISPLAY}) ==="""
    return ask_gemini(prompt)


# ── 파일 찾기 ──────────────────────────────────────────

def find_file(*patterns: str) -> pathlib.Path | None:
    for pat in patterns:
        matches = list(EXCEL_DIR.glob(pat))
        if matches:
            return sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return None


# ── 각 파일 처리 함수 ────────────────────────────────

def process_추정이익() -> str:
    path = find_file("추정이익*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    sections = []
    for sheet in ["Rating_Up", "Rating_Down", "TP_Up", "TP_Down", "New_KOSPI", "New_KOSDAQ"]:
        text = read_sheet_as_text(wb, sheet, max_rows=30)
        sections.append(f"[{sheet}]\n{text}")
    wb.close()

    raw_data = "\n\n".join(sections)
    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 증권사 추정이익 변경 데이터입니다.

{raw_data}

위 데이터에서 아래 형식으로 핵심만 요약하세요:

[Rating_Up] — 등급 상향 (상위 15개)
종목명 | 증권사 | 이전등급→신규등급 | 날짜

[Rating_Down] — 등급 하향 (상위 15개)
종목명 | 증권사 | 이전등급→신규등급 | 날짜

[TP_Up] — 목표주가 상향 (상위 20개, 변화율 큰 순)
종목명 | 증권사 | 이전TP→신규TP | 변화율%

[TP_Down] — 목표주가 하향 (상위 20개)
종목명 | 증권사 | 이전TP→신규TP | 변화율%

[신규커버리지]
종목명 | 증권사 | 등급 | TP
"""
    return ask_gemini(prompt)


def process_컨센() -> str:
    path = find_file("컨센*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    sections = []
    for sheet in ["서프라이즈", "쇼크", "컨센상향", "컨센하향"]:
        text = read_sheet_as_text(wb, sheet, max_rows=30)
        sections.append(f"[{sheet}]\n{text}")
    # 1개월 컨센 변화 상위/하위
    text5 = read_sheet_as_text(wb, "1개월 컨센 1개월일 변화", max_rows=30)
    sections.append(f"[1개월컨센변화]\n{text5}")
    wb.close()

    raw_data = "\n\n".join(sections)
    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 컨센서스 서프라이즈/쇼크 데이터입니다.

{raw_data}

위 데이터에서 아래 형식으로 요약하세요:

[서프라이즈] — 실적 크게 상회 (상위 20개)
종목명 | 서프라이즈율% | 실적 | 컨센 | 분기

[쇼크] — 실적 크게 하회 (상위 20개)
종목명 | 쇼크율% | 실적 | 컨센 | 분기

[컨센상향] — 추정치 상향 (상위 20개)
종목명 | 변화율% | 이전→현재

[컨센하향] — 추정치 하향 (상위 20개)
종목명 | 변화율%

[1개월컨센변화 상위] — 상위 10개
종목명 | 변화율% | 섹터
"""
    return ask_gemini(prompt)


def process_일정수주() -> str:
    path = find_file("일정 및 수주잔고 투경*", "일정 및 수주잔고*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    # 이벤트 일정
    event_text = read_sheet_as_text(wb, "25년말~26년", max_rows=80)
    # 투경관리
    투경_text = read_sheet_as_text(wb, "투경관리", max_rows=50)
    # 수주잔고
    수주_text = read_sheet_as_text(wb, "수주잔고", max_rows=50)
    wb.close()

    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 이벤트 일정·투경관리·수주잔고 데이터입니다.

[이벤트 캘린더]
{event_text[:2000]}

[투경관리]
{투경_text[:1000]}

[수주잔고]
{수주_text[:2000]}

위 데이터에서 아래 형식으로 요약하세요:

[이벤트일정] D-60 이내만:
날짜 | D-? | 관련종목/업종 | 내용

[투경관리] 현재 상태:
경고예고: 종목명 | 공시일 | 판단마지막날
경고중: 종목명 | 공시일 | 해제판단일

[수주잔고] QoQ 변화 큰 순 상위 10개:
종목명 | 업종 | 직전분기 | 최신분기 | QoQ%
"""
    return ask_gemini(prompt)


def process_투자아이디어() -> str:
    path = find_file("투자아이디어*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    # 주요 시트 추출 (해외주식 제외, 과거 분기 제외)
    skip_keywords = ["해외주식", "해외", "Sheet"]
    sections = []
    for name in wb.sheetnames:
        if any(kw in name for kw in skip_keywords):
            continue
        text = read_sheet_as_text(wb, name, max_rows=20)
        if text.strip():
            sections.append(f"[{name}]\n{text}")
    wb.close()

    if not sections:
        return "추출된 시트 없음"

    raw_data = "\n\n".join(sections[:25])  # 최대 25개 시트
    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 섹터별 투자아이디어 분기 모멘텀 데이터입니다.

{raw_data[:5000]}

2026년 2분기(2Q) 이후 분기의 모멘텀만 추출하세요.
과거 분기(1Q 이전)는 제외하세요.

[섹터/종목그룹명]
종목명 | 분기 | 핵심모멘텀요약

(모멘텀이 없는 섹터는 생략)
"""
    return ask_gemini(prompt)


def process_액티브ETF() -> str:
    path = find_file("액티브ETF*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    # 주요 섹터 요약 시트들만 (과거 제외, 유니버스 제외)
    target_sheets = [s for s in wb.sheetnames
                     if "과거" not in s and "유니버스" not in s and len(s) < 20]
    sections = []
    for name in target_sheets[:10]:  # 상위 10개 시트
        text = read_sheet_as_text(wb, name, max_rows=15)
        if text.strip():
            sections.append(f"[{name}]\n{text}")
    wb.close()

    if not sections:
        return "추출된 시트 없음"

    raw_data = "\n\n".join(sections)
    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 액티브ETF 편입 종목 데이터입니다.

{raw_data[:5000]}

아래 형식으로 요약하세요:

[ETF명]
상위편입: 종목명(비중%) 나열

[공통편입 상위] — 여러 ETF에서 동시 편입된 종목:
종목명 | ETF 수 | 편입 ETF 목록
"""
    return ask_gemini(prompt)


def process_페어트레이딩() -> str:
    path = find_file("페어트레이딩*")
    if not path:
        return ""
    wb = open_workbook(path)
    if not wb:
        return ""

    페어_text = read_sheet_as_text(wb, "페어", max_rows=50)
    체크_text = read_sheet_as_text(wb, "페어트레이딩 체크", max_rows=60)
    wb.close()

    prompt = f"""오늘 날짜: {TODAY_DISPLAY}
아래는 페어트레이딩 현황 데이터입니다.

[페어 목록]
{페어_text[:1500]}

[페어트레이딩 체크 — 스프레드 현황]
{체크_text[:3000]}

위 데이터에서 아래 형식으로 요약하세요:

[스프레드 주목 페어] — 스프레드가 크게 벌어진 TOP 10:
A종목 vs B종목 | 현재스프레드 | 평균 | 방향(A매수/B매도 등)

[스프레드 축소 페어] — 수렴 중인 TOP 5:
A종목 vs B종목 | 방향

[오늘의 추천 페어] 1~2개 + 이유:
"""
    return ask_gemini(prompt)


# ── 처리 라우팅 ──────────────────────────────────────

PROCESSORS = {
    "추정이익": process_추정이익,
    "컨센":     process_컨센,
    "일정수주": process_일정수주,
    "투자아이디어": process_투자아이디어,
    "액티브ETF": process_액티브ETF,
    "페어트레이딩": process_페어트레이딩,
}

# ── 메인 실행 ──────────────────────────────────────────

def run(ftype: str) -> bool:
    today_path = SUMMARY_DIR / f"{TODAY}_{ftype}.md"
    if today_path.exists() and not FORCE:
        print(f"  [{ftype}] 이미 처리됨 (--no-skip으로 재처리)")
        return True

    print(f"\n[{ftype}] 처리 중...")
    try:
        result = PROCESSORS[ftype]()
        if not result:
            print(f"  → 파일 없음 또는 추출 실패")
            return False

        # 오늘 전체 저장
        today_path.write_text(
            f"# {ftype} — {TODAY_DISPLAY}\n\n{result}\n",
            encoding="utf-8"
        )

        # 변화 감지
        changes = detect_changes(ftype, result)
        change_path = SUMMARY_DIR / f"{TODAY}_{ftype}_변화.md"
        change_path.write_text(
            f"# {ftype} 변화 — {TODAY_DISPLAY}\n\n{changes}\n",
            encoding="utf-8"
        )

        is_unchanged = changes.strip().startswith("변화없음")
        status = "변화없음" if is_unchanged else f"변화 감지 → {change_path.name}"
        print(f"  → {status}")
        time.sleep(1)
        return True

    except Exception as e:
        print(f"  [오류] {e}")
        return False


def main():
    # 인수에서 유형 필터 추출 (--no-skip 제외)
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    targets = [t for t in PROCESSORS if not args or any(a in t for a in args)]

    print(f"\n매일 Excel 추출 — {TODAY_DISPLAY}")
    print(f"저장 위치: raw/매일요약/\n")

    results = {t: run(t) for t in targets}

    print(f"\n{'='*40}")
    print(f"처리: {sum(results.values())}/{len(results)}개")

    changed = []
    for ftype in targets:
        cp = SUMMARY_DIR / f"{TODAY}_{ftype}_변화.md"
        if cp.exists():
            txt = cp.read_text(encoding="utf-8")
            if not txt.strip().startswith("변화없음") and "변화없음" not in txt[:30]:
                changed.append(ftype)

    if changed:
        print(f"\n변화 감지: {', '.join(changed)}")
        print(f"\n다음 명령으로 wiki 반영:")
        for ftype in changed:
            print(f"  /ingest raw/매일요약/{TODAY}_{ftype}_변화.md")
    else:
        print("\n오늘 변화 없음 — wiki 업데이트 불필요")


if __name__ == "__main__":
    main()

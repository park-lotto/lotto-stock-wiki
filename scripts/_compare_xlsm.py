"""3개 수급오실레이터 xlsm 비교"""
import openpyxl, os, sys
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

files = [
    ('0521', os.path.join(BASE, 'raw', '외국인기관수급오실레이터0521.xlsm')),
    ('0528', os.path.join(BASE, 'raw', '외국인기관수급오실레이터 0528.xlsm')),
    ('0601', os.path.join(BASE, 'raw', '매일 엑셀넣을것', '외국인기관수급오실레이터0601.xlsm')),
]

KEY_SHEETS = ['수급오실레이터', '외국인매수데이터', '기관매수데이터', '시가총액', '시기외']

for label, path in files:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    print(f'=== {label} ({os.path.getsize(path)//1024}KB) ===')
    print(f'  시트목록: {wb.sheetnames}')
    for sname in KEY_SHEETS:
        if sname in wb.sheetnames:
            ws = wb[sname]
            print(f'  [{sname}] {ws.max_row}행 x {ws.max_column}열')
    # 종목 수: 외국인매수데이터 9행 컬럼 수
    if '외국인매수데이터' in wb.sheetnames:
        ws = wb['외국인매수데이터']
        cols = 0
        for cell in ws[9]:
            if cell.value and isinstance(cell.value, str):
                cols += 1
        print(f'  종목수 (9행 문자열): {cols}개')
    # 데이터 날짜 범위: 외국인매수데이터 1열 15~91행
    if '외국인매수데이터' in wb.sheetnames:
        ws = wb['외국인매수데이터']
        dates = []
        for r in range(15, 92):
            v = ws.cell(r, 1).value
            if v: dates.append(str(v)[:10])
        if dates:
            print(f'  날짜범위: {dates[0]} ~ {dates[-1]} ({len(dates)}거래일)')
    # 수급오실레이터 L10(p25), L11(p10) 기준값
    if '수급오실레이터' in wb.sheetnames:
        ws = wb['수급오실레이터']
        p10 = ws.cell(11, 12).value
        p25 = ws.cell(10, 12).value
        print(f'  기준값 p10={p10}  p25={p25}')
    wb.close()
    print()

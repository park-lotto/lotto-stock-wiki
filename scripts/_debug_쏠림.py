import sys; sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path
from datetime import datetime

DIR = Path(__file__).parent.parent / 'raw' / '매일 엑셀넣을것'
wb = openpyxl.load_workbook(str(DIR / '특정업종쏠림지수국내0602.xlsx'), read_only=True, data_only=True)

# 업종쏠림지수 시트 확인
print('=== 업종쏠림지수 시트 ===')
ws = wb['업종쏠림지수']
rows = list(ws.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True))
for i, r in enumerate(rows):
    vals = [(j,v) for j,v in enumerate(r) if v is not None]
    if vals: print(f'  행{i+1}: {vals}')

print()
print('=== 최종 시트 — 헤더+최근3행 ===')
ws2 = wb['최종']
all_rows = list(ws2.iter_rows(min_row=1, max_row=9999, max_col=12, values_only=True))
# 헤더 찾기
for i, r in enumerate(all_rows[:20]):
    vals = [(j,v) for j,v in enumerate(r) if v is not None]
    if vals: print(f'  행{i+1}: {vals}')

# 최신 5개 데이터
valid = [r for r in all_rows if r[0] and isinstance(r[0], datetime)]
print(f'\n최신 3개:')
for r in valid[-3:]:
    print(f'  {[(j,v) for j,v in enumerate(r) if v is not None]}')

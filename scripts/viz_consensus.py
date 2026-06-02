"""
viz_consensus.py — 종목 TP 변화 시각화 HTML 생성

사용법:
    python scripts/viz_consensus.py SK하이닉스
    python scripts/viz_consensus.py LG이노텍 삼성전자
    python scripts/viz_consensus.py --list          ← 오늘 변경된 전체 종목 목록
"""

import sys, argparse, json, re, subprocess
from pathlib import Path
from datetime import datetime, date

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("openpyxl 없음. pip install openpyxl")
    sys.exit(1)

ROOT      = Path(__file__).parent.parent
EXCEL_DIR = ROOT / "raw" / "매일 엑셀넣을것"
OUT_DIR   = ROOT / "out"
OUT_DIR.mkdir(exist_ok=True)
TODAY     = datetime.today().strftime("%Y-%m-%d")

# ── 증권사별 색상 팔레트
BROKER_COLORS = [
    "#60a5fa","#34d399","#fbbf24","#f472b6","#a78bfa",
    "#38bdf8","#fb923c","#4ade80","#e879f9","#94a3b8",
    "#f9a8d4","#6ee7b7","#fde68a","#c4b5fd","#7dd3fc",
]

def find_excel(pattern):
    for f in EXCEL_DIR.iterdir():
        if f.name.startswith("~$"): continue
        if pattern in f.name and f.suffix in (".xlsx", ".xlsm"):
            return f
    return None

def load_all(wb) -> list[dict]:
    """추정이익변경 파일 전체 읽기"""
    records = []
    sheet_types = {
        "Rating_Up":   "rating_up",
        "Rating_Down": "rating_down",
        "TP_Up":       "tp_up",
        "TP_Down":     "tp_down",
        "New_KOSPI":   "new",
        "New_KOSDAQ":  "new",
    }
    for sheet_name, stype in sheet_types.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=300, max_col=16, values_only=True):
            if any(v is not None for v in row):
                rows.append(list(row))

        header_idx = None
        for i, r in enumerate(rows):
            vals = [str(v) for v in r if v is not None]
            if any(k in vals for k in ["일자", "코드", "From"]):
                header_idx = i
                break
        if header_idx is None: continue

        for r in rows[header_idx + 2:]:
            date_v = r[1] if len(r) > 1 else None
            code_v = r[2] if len(r) > 2 else None
            name_v = r[3] if len(r) > 3 else None
            if not isinstance(date_v, (datetime, date)): continue
            if not isinstance(code_v, str) or not code_v.startswith("A"): continue
            if not name_v: continue
            tp_old = int(r[13]) if len(r) > 13 and isinstance(r[13], (int, float)) else None
            tp_new = int(r[14]) if len(r) > 14 and isinstance(r[14], (int, float)) else None
            if tp_new is None: continue

            op_old = str(r[11]).strip() if len(r) > 11 and r[11] else "—"
            op_new = str(r[12]).strip() if len(r) > 12 and r[12] else "—"

            records.append({
                "date":      date_v.strftime("%Y-%m-%d") if isinstance(date_v, (datetime, date)) else str(date_v),
                "code":      code_v,
                "name":      str(name_v).strip(),
                "brokerage": str(r[5]).strip() if len(r) > 5 and r[5] else "—",
                "op_old":    op_old,
                "op_new":    op_new,
                "tp_old":    tp_old,
                "tp_new":    tp_new,
                "stype":     stype,
            })
    return records

def filter_stock(records: list, name: str) -> list:
    name = name.strip()
    return [r for r in records if name in r["name"] or r["name"] in name]

def build_html(stock_name: str, items: list) -> str:
    if not items:
        return f"<h2>{stock_name} 데이터 없음</h2>"

    from datetime import timedelta
    RECENT_DAYS = 45  # 최근 45일 이내만 의미있는 데이터로 간주
    cutoff = (datetime.today() - timedelta(days=RECENT_DAYS)).strftime("%Y-%m-%d")

    # 최신 TP per brokerage (날짜순 정렬 → 가장 최근 것 덮어쓰기)
    latest_all = {}
    for it in sorted(items, key=lambda x: x["date"]):
        latest_all[it["brokerage"]] = it

    # 최근 45일 이내 = 의미있는 데이터
    latest_recent = {b: it for b, it in latest_all.items() if it["date"] >= cutoff}
    latest_old    = {b: it for b, it in latest_all.items() if it["date"] < cutoff}

    # 최신순 정렬 (날짜 최신 → 오래된 순)
    recent_list = sorted(latest_recent.values(), key=lambda x: x["date"], reverse=True)
    old_list    = sorted(latest_old.values(),    key=lambda x: x["tp_new"], reverse=True)

    # 컨센서스는 최근 것만
    tp_vals   = [it["tp_new"] for it in recent_list if it["tp_new"]]
    consensus = int(sum(tp_vals) / len(tp_vals)) if tp_vals else 0
    max_tp    = max(tp_vals) if tp_vals else 0
    min_tp    = min(tp_vals) if tp_vals else 0
    top_tp    = max_tp  # 신고가 기준

    brokers   = list(dict.fromkeys(it["brokerage"] for it in items))
    color_map = {b: BROKER_COLORS[i % len(BROKER_COLORS)] for i, b in enumerate(brokers)}

    # 전체(recent만) 표시용
    latest_list = recent_list

    # 바차트 (최신 TP, 높은 순으로 재정렬)
    bar_sorted = sorted(recent_list, key=lambda x: x["tp_new"], reverse=True)
    bar_labels = [it["brokerage"] for it in bar_sorted]
    bar_tp     = [it["tp_new"] for it in bar_sorted]
    bar_pct    = []
    bar_colors = []
    for it in bar_sorted:
        pct = (it["tp_new"] - it["tp_old"]) / it["tp_old"] * 100 if it["tp_old"] and it["tp_new"] else 0
        bar_pct.append(round(pct, 1))
        if it["stype"] in ("rating_up", "tp_up"):
            bar_colors.append("rgba(34,197,94,0.85)")
        elif it["stype"] in ("rating_down", "tp_down"):
            bar_colors.append("rgba(239,68,68,0.85)")
        else:
            bar_colors.append("rgba(99,102,241,0.85)")

    # 타임라인 (최근 브로커만)
    recent_brokers = [it["brokerage"] for it in recent_list]
    timeline_datasets = []
    for broker in recent_brokers:
        pts  = sorted([it for it in items if it["brokerage"] == broker], key=lambda x: x["date"])
        data = [{"x": p["date"], "y": p["tp_new"]} for p in pts]
        timeline_datasets.append({
            "label": broker, "data": data,
            "borderColor": color_map[broker], "backgroundColor": color_map[broker],
            "pointRadius": 5, "pointHoverRadius": 8,
            "showLine": True, "tension": 0.3, "borderWidth": 2,
        })

    # 레이팅 변경 (최근 45일)
    rating_changes = sorted(
        [it for it in items if it["op_old"] != it["op_new"] and it["op_old"] != "—" and it["date"] >= cutoff],
        key=lambda x: x["date"], reverse=True
    )

    def days_ago(d):
        delta = (datetime.today() - datetime.strptime(d, "%Y-%m-%d")).days
        return "오늘" if delta == 0 else f"{delta}일 전"

    def rbadge(op_old, op_new, stype):
        c = "#10b981" if "up" in stype else "#ef4444"
        icon = "📈" if "up" in stype else "📉"
        return f'<span class="rbadge" style="background:{c}">{icon} {op_old} → {op_new}</span>'

    # TP 카드 — 날짜 최신순, 신고가 뱃지
    tp_cards = ""
    for it in recent_list:
        pct   = round((it["tp_new"] - it["tp_old"]) / it["tp_old"] * 100 if it["tp_old"] and it["tp_new"] else 0, 1)
        cls   = "up" if pct > 0 else ("down" if pct < 0 else "flat")
        pcls  = "up-c" if pct > 0 else ("dn-c" if pct < 0 else "fl-c")
        arrow = "▲" if pct > 0 else ("▼" if pct < 0 else "—")
        old_s = f"{it['tp_old']//10000}만" if it['tp_old'] else "신규"
        is_top  = it["tp_new"] == top_tp
        is_high = it["tp_new"] >= top_tp * 0.93 and not is_top
        badge_html = '<div class="top-badge">🏆 신고가</div>' if is_top else ('<div class="top-badge hi">🔥 상위</div>' if is_high else '<div class="top-badge empty"></div>')
        da = days_ago(it["date"])
        tp_cards += (
            f'<div class="card {cls}">'
            + badge_html
            + f'<div class="card-broker">{it["brokerage"]}</div>'
            + f'<div class="card-tp">{it["tp_new"]//10000}<span class="card-unit">만원</span></div>'
            + f'<div class="card-pct {pcls}">{arrow} {abs(pct):.0f}%<span class="card-from"> ({old_s}→)</span></div>'
            + f'<div class="card-date">{da}</div>'
            + f'</div>'
        )

    # 제외된 오래된 데이터
    old_note = ""
    if old_list:
        names_str = ", ".join(it["brokerage"] for it in old_list[:4]) + ("…" if len(old_list) > 4 else "")
        old_note  = f'<div class="old-note">📌 {RECENT_DAYS}일 이상 지난 자료 제외: {names_str} ({len(old_list)}개사)</div>'

    # 범위 바 % 위치
    range_pct_con = round((consensus - min_tp) / (max_tp - min_tp) * 100) if max_tp != min_tp else 50

    # 레이팅 섹션
    rating_rows = "".join(
        f"<tr><td class='da'>{days_ago(it['date'])}</td><td><b>{it['brokerage']}</b></td>"
        f"<td>{rbadge(it['op_old'],it['op_new'],it['stype'])}</td>"
        f"<td>{it['tp_old']:,}원→<b>{it['tp_new']:,}원</b></td></tr>"
        for it in rating_changes[:5]
    )
    rating_section = (
        f'<div class="panel-title">📢 투자의견 변경 <span class="ptag">(매수/매도 의견)</span></div>'
        f'<div class="rating-wrap"><table class="rtable"><tbody>{rating_rows}</tbody></table></div>'
    ) if rating_rows else f'<div style="color:var(--text3);font-size:12px;padding:8px 0">최근 {RECENT_DAYS}일 투자의견 변경 없음</div>'

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>{stock_name} 목표주가</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3.0.0/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
*{{box-sizing:border-box;margin:0;padding:0}}
html,body{{height:100%;overflow:hidden;font-family:'Inter','Malgun Gothic',sans-serif}}

/* ══ CSS 변수 — 다크 테마 (기본) ══ */
:root{{
  --bg:        #060b18;
  --bg2:       rgba(255,255,255,.03);
  --border:    rgba(255,255,255,.07);
  --border2:   rgba(255,255,255,.13);
  --text:      #e2e8f0;
  --text2:     #94a3b8;
  --text3:     #475569;
  --text4:     #374151;
  --grid:      rgba(255,255,255,.05);
  --card-bg:   rgba(255,255,255,.04);
  --card-hov:  rgba(255,255,255,.07);
  --glow1:     rgba(59,130,246,.07);
  --glow2:     rgba(139,92,246,.06);
  --tick:      #475569;
  --tt-bg:     rgba(15,23,42,.95);
}}

/* ══ CSS 변수 — 라이트 테마 ══ */
html.light{{
  --bg:        #f0f4ff;
  --bg2:       rgba(255,255,255,.8);
  --border:    rgba(0,0,0,.08);
  --border2:   rgba(0,0,0,.16);
  --text:      #0f172a;
  --text2:     #475569;
  --text3:     #64748b;
  --text4:     #94a3b8;
  --grid:      rgba(0,0,0,.06);
  --card-bg:   rgba(255,255,255,.9);
  --card-hov:  #ffffff;
  --glow1:     rgba(59,130,246,.12);
  --glow2:     rgba(139,92,246,.10);
  --tick:      #64748b;
  --tt-bg:     rgba(255,255,255,.98);
}}

body{{background:var(--bg);color:var(--text)}}

/* ── 배경 그라디언트 ── */
body::before{{
  content:'';position:fixed;inset:0;
  background:
    radial-gradient(ellipse 80% 50% at 20% 20%, var(--glow1) 0%, transparent 60%),
    radial-gradient(ellipse 60% 40% at 80% 80%, var(--glow2) 0%, transparent 60%);
  pointer-events:none;
}}

/* ── 레이아웃: 헤더(auto) + 범위바(auto) + 카드(auto) + 차트(1fr) ── */
.layout{{
  display:grid;
  grid-template-rows:62px auto auto 1fr;
  height:100vh;
  padding:10px 14px 10px;
  gap:7px;
}}

/* ── 헤더 ── */
.header{{display:flex;align-items:center;gap:12px}}
.stock-name{{font-size:26px;font-weight:800;color:var(--text);letter-spacing:-.5px}}
.stock-sub{{font-size:11px;color:var(--text3);line-height:1.6;margin-left:2px}}
.con-box{{
  background:linear-gradient(135deg,rgba(96,165,250,.15),rgba(167,139,250,.15));
  border:1px solid rgba(96,165,250,.3);
  border-radius:10px;padding:6px 16px;margin-left:4px;
}}
.con-label{{font-size:9px;color:#60a5fa;font-weight:600;letter-spacing:.8px;text-transform:uppercase}}
.con-value{{font-size:22px;font-weight:800;color:#f8fafc}}
.stat-box{{display:flex;gap:16px;margin-left:6px}}
.stat-item{{text-align:center}}
.stat-label{{font-size:9px;color:var(--text3);margin-bottom:2px}}
.stat-val{{font-size:14px;font-weight:700}}
.meta{{margin-left:auto;font-size:10px;color:var(--text3);line-height:1.7;text-align:right}}
#themeBtn{{
  background:var(--card-bg);border:1px solid var(--border);
  border-radius:8px;padding:5px 9px;cursor:pointer;
  font-size:15px;line-height:1;margin-left:6px;
  transition:all .2s;
}}
#themeBtn:hover{{background:var(--card-hov);border-color:var(--border2)}}

/* ── 범위 바 ── */
.range-wrap{{
  background:var(--bg2);border:1px solid var(--border);
  border-radius:10px;padding:10px 16px;
}}
.range-label{{font-size:10px;color:var(--text3);margin-bottom:6px;font-weight:600}}
.range-track{{position:relative;height:8px;background:rgba(255,255,255,.08);border-radius:99px;margin:0 4px}}
.range-fill{{
  position:absolute;left:0;top:0;height:100%;
  background:linear-gradient(90deg,#3b82f6,#a78bfa);
  border-radius:99px;width:{range_pct_con}%;
}}
.range-marker{{
  position:absolute;top:50%;transform:translateY(-50%) translateX(-50%);
  left:{range_pct_con}%;
  width:14px;height:14px;background:#fff;border-radius:50%;
  border:2px solid #a78bfa;box-shadow:0 0 8px rgba(167,139,250,.5);
}}
.range-ends{{display:flex;justify-content:space-between;font-size:10px;margin-top:4px;color:var(--text3)}}
.range-ends .lo{{color:#f87171;font-weight:600}}
.range-ends .hi{{color:#22d3ee;font-weight:600}}
.range-con-label{{
  position:absolute;top:-18px;
  left:{range_pct_con}%;transform:translateX(-50%);
  font-size:9px;color:#a78bfa;font-weight:700;white-space:nowrap;
}}
.old-note{{font-size:10px;color:var(--text3);margin-top:5px}}

/* ── TP 카드 ── */
.cards-row{{display:flex;gap:7px;overflow-x:auto;scrollbar-width:none;align-items:stretch}}
.cards-row::-webkit-scrollbar{{display:none}}
.card{{
  flex-shrink:0;position:relative;
  background:var(--card-bg);border:1px solid var(--border);
  border-radius:10px;padding:10px 13px 8px;min-width:114px;
  transition:all .2s;cursor:default;
}}
.card:hover{{background:var(--card-hov);border-color:var(--border2);transform:translateY(-1px)}}
.card.up  {{border-top:3px solid #22d3ee}}
.card.down{{border-top:3px solid #f87171}}
.card.flat{{border-top:3px solid #818cf8}}
.top-badge{{
  font-size:9px;font-weight:700;margin-bottom:3px;height:14px;
  color:#fbbf24;letter-spacing:.3px;
}}
.top-badge.hi{{color:#34d399}}
.top-badge.empty{{visibility:hidden}}
.card-broker{{font-size:10px;color:var(--text3);margin-bottom:3px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:100px;font-weight:500}}
.card-tp{{font-size:22px;font-weight:800;color:var(--text);line-height:1}}
.card-unit{{font-size:11px;font-weight:400;color:var(--text3)}}
.card-pct{{font-size:11px;font-weight:600;margin-top:3px}}
.card-from{{font-size:9px;color:var(--text4);font-weight:400}}
.card-date{{font-size:9px;color:var(--text4);margin-top:3px}}
.up-c{{color:#22d3ee}} .dn-c{{color:#f87171}} .fl-c{{color:#818cf8}}

/* ── 메인 차트 영역 ── */
.main{{display:grid;grid-template-columns:42% 1fr;gap:8px;min-height:0}}
.left-col{{display:grid;grid-template-rows:1fr 150px;gap:8px;min-height:0}}
.right-col{{min-height:0}}
.panel{{
  background:var(--bg2);border:1px solid var(--border);
  border-radius:12px;padding:11px 14px;
  display:flex;flex-direction:column;min-height:0;overflow:hidden;
}}
.panel-title{{font-size:10px;font-weight:700;color:var(--text3);margin-bottom:7px;flex-shrink:0;letter-spacing:.3px}}
.ptag{{font-weight:400;font-size:9px;opacity:.7}}
.chart-wrap{{flex:1;min-height:0;position:relative}}
canvas{{width:100%!important;height:100%!important}}
.rating-wrap{{overflow-y:auto;flex:1;scrollbar-width:none}}
.rtable{{width:100%;border-collapse:collapse;font-size:11px}}
.rtable td{{padding:5px 4px;border-bottom:1px solid var(--border);color:var(--text2);white-space:nowrap;vertical-align:middle}}
.rtable .da{{color:var(--text3);font-size:10px;width:46px}}
.rtable td:nth-child(2){{color:var(--text);font-weight:600}}
.rtable td:nth-child(4){{color:var(--text2);font-size:10px}}
.rtable tr:last-child td{{border:none}}
.rbadge{{display:inline-block;padding:2px 7px;border-radius:4px;font-size:10px;font-weight:700;color:#fff}}
</style>
</head>
<body>
<div class="layout">

  <!-- 헤더 -->
  <div class="header">
    <div>
      <div class="stock-name">{stock_name}</div>
    </div>
    <div class="con-box">
      <div class="con-label">전문가 평균 목표주가</div>
      <div class="con-value">{consensus:,}원</div>
    </div>
    <div class="stat-box">
      <div class="stat-item">
        <div class="stat-label">신고가</div>
        <div class="stat-val" style="color:#22d3ee">{max_tp:,}원</div>
      </div>
      <div class="stat-item">
        <div class="stat-label">최저</div>
        <div class="stat-val" style="color:#f87171">{min_tp:,}원</div>
      </div>
      <div class="stat-item">
        <div class="stat-label">분석 증권사</div>
        <div class="stat-val" style="color:#a78bfa">{len(recent_list)}개사</div>
      </div>
    </div>
    <div class="meta">{TODAY}<br><span>최근 {RECENT_DAYS}일 기준</span></div>
    <button id="themeBtn" onclick="toggleTheme()" title="밝기 전환"><span id="themeIcon">☀️</span></button>
  </div>

  <!-- 목표주가 범위 바 -->
  <div class="range-wrap">
    <div class="range-label">📊 목표주가 범위 — 전문가들이 예상하는 주가 구간</div>
    <div class="range-track">
      <div class="range-fill"></div>
      <div class="range-marker">
        <span class="range-con-label">평균 {consensus//10000}만</span>
      </div>
    </div>
    <div class="range-ends">
      <span class="lo">최저 {min_tp:,}원</span>
      <span class="hi">신고가 {max_tp:,}원</span>
    </div>
    {old_note}
  </div>

  <!-- TP 카드 (최신순, 신고가 뱃지) -->
  <div class="cards-row">{tp_cards}</div>

  <!-- 차트 + 투자의견 -->
  <div class="main">
    <div class="left-col">
      <div class="panel">
        <div class="panel-title">🎯 목표주가 비교 (높은 순)</div>
        <div class="chart-wrap"><canvas id="barChart"></canvas></div>
      </div>
      <div class="panel">{rating_section}</div>
    </div>
    <div class="panel right-col">
      <div class="panel-title">📈 목표주가 변화 흐름 <span class="ptag">(날짜별)</span></div>
      <div class="chart-wrap"><canvas id="lineChart"></canvas></div>
    </div>
  </div>
</div>

<script>
const BAR_PCTS = {json.dumps(bar_pct)};
const BAR_LABELS = {json.dumps(bar_labels, ensure_ascii=False)};

let isLight = false;
function toggleTheme() {{
  isLight = !isLight;
  document.documentElement.classList.toggle('light', isLight);
  document.getElementById('themeIcon').textContent = isLight ? '🌙' : '☀️';
  const g = getComputedStyle(document.documentElement);
  const GRID = g.getPropertyValue('--grid').trim();
  const TICK = g.getPropertyValue('--tick').trim();
  const TT   = g.getPropertyValue('--tt-bg').trim();
  [barChart, lineChart].forEach(c => {{
    if (!c) return;
    c.options.scales.x.grid.color = GRID;
    c.options.scales.y.grid.color = GRID;
    c.options.scales.x.ticks.color = TICK;
    c.options.scales.y.ticks.color = TICK;
    c.options.plugins.tooltip.backgroundColor = TT;
    c.options.plugins.tooltip.titleColor = isLight ? '#0f172a' : '#e2e8f0';
    c.options.plugins.tooltip.bodyColor  = isLight ? '#475569' : '#94a3b8';
    if (c.options.plugins.legend)
      c.options.plugins.legend.labels.color = isLight ? '#64748b' : '#64748b';
    c.update();
  }});
}}

function gv(v) {{ return getComputedStyle(document.documentElement).getPropertyValue(v).trim(); }}
let GRID_C = gv('--grid');
let TICK_C = gv('--tick');

const barChart = new Chart(document.getElementById('barChart'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(bar_labels, ensure_ascii=False)},
    datasets:[{{
      data:{json.dumps(bar_tp)},
      backgroundColor:{json.dumps(bar_colors)},
      borderRadius:5, borderWidth:0,
    }}]
  }},
  options:{{
    indexAxis:'y', responsive:true, maintainAspectRatio:false,
    plugins:{{
      legend:{{display:false}},
      tooltip:{{
        backgroundColor:'rgba(15,23,42,.95)',
        borderColor:'rgba(255,255,255,.1)',borderWidth:1,
        titleColor:'#e2e8f0',bodyColor:'#94a3b8',
        callbacks:{{label:ctx=>{{
          const p=BAR_PCTS[ctx.dataIndex],s=p>0?'+':'';
          return ` ${{ctx.parsed.x.toLocaleString()}}원  (${{s}}${{p}}%)`;
        }}}}
      }}
    }},
    scales:{{
      x:{{
        beginAtZero:false,
        ticks:{{color:TICK_C,font:{{size:10}},callback:v=>(v/10000).toFixed(0)+'만'}},
        grid:{{color:GRID_C}}
      }},
      y:{{ticks:{{color:'#94a3b8',font:{{size:10.5}}}},grid:{{display:false}}}}
    }}
  }}
}});

const lineChart = new Chart(document.getElementById('lineChart'), {{
  type:'scatter',
  data:{{datasets:{json.dumps(timeline_datasets, ensure_ascii=False)}}},
  options:{{
    responsive:true, maintainAspectRatio:false,
    plugins:{{
      legend:{{
        position:'right',
        labels:{{boxWidth:8,boxHeight:8,borderRadius:4,useBorderRadius:true,
          font:{{size:10.5}},color:'#64748b',padding:8}}
      }},
      tooltip:{{
        backgroundColor:'rgba(15,23,42,.95)',
        borderColor:'rgba(255,255,255,.1)',borderWidth:1,
        titleColor:'#e2e8f0',bodyColor:'#94a3b8',
        callbacks:{{label:ctx=>` ${{ctx.dataset.label}}: ${{ctx.parsed.y.toLocaleString()}}원`}}
      }}
    }},
    scales:{{
      x:{{
        type:'time',
        time:{{unit:'day',tooltipFormat:'yyyy-MM-dd',displayFormats:{{day:'MM/dd'}}}},
        grid:{{color:GRID_C}},
        ticks:{{color:TICK_C,font:{{size:10}}}}
      }},
      y:{{
        grid:{{color:GRID_C}},
        ticks:{{color:TICK_C,font:{{size:10}},callback:v=>(v/10000).toFixed(0)+'만'}}
      }}
    }}
  }}
}});
</script>
</body>
</html>"""

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("stocks", nargs="*", help="종목명 (예: SK하이닉스 LG이노텍)")
    parser.add_argument("--list", action="store_true", help="오늘 변경 전체 종목 목록")
    args = parser.parse_args()

    path = find_excel("추정이익 변경")
    if not path:
        print("❌ 추정이익 변경 파일 없음 — raw/매일 엑셀넣을것/ 확인")
        sys.exit(1)

    print(f"📂 {path.name} 읽는 중...")
    wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=False, data_only=True)
    records = load_all(wb)
    wb.close()
    print(f"   전체 {len(records)}건 로드")

    if args.list:
        names = sorted(set(r["name"] for r in records))
        print(f"\n📋 오늘 변경된 종목 ({len(names)}개):")
        for n in names:
            cnt = len([r for r in records if r["name"] == n])
            print(f"  {n} ({cnt}건)")
        return

    if not args.stocks:
        parser.print_help()
        return

    opened = []
    for stock in args.stocks:
        items = filter_stock(records, stock)
        if not items:
            print(f"⚠️  {stock} — 데이터 없음 (--list 로 목록 확인)")
            continue

        html = build_html(stock, items)
        out_path = OUT_DIR / f"consensus_{stock}_{TODAY}.html"
        out_path.write_text(html, encoding="utf-8")
        print(f"✅ {stock} — {len(items)}건 → {out_path.name}")
        opened.append(str(out_path))

    for p in opened:
        subprocess.Popen(["start", p], shell=True)

if __name__ == "__main__":
    main()

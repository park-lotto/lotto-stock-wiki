"""섹터 히트맵 데이터 빌드.

raw/내 관심종목.xlsx → 전체 계층 파싱:
  L일봉H 이후 첫 텍스트행 = 대섹터
  이후 텍스트행 = 서브섹터(소분류)
  종목코드 행 = 종목
"""
import re
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
WATCH_FILE = ROOT / "raw" / "내 관심종목.xlsx"
SECTOR_CUSTOM_PATH = ROOT / "pipeline" / "sector_custom.json"

# ── NXT(넥스트트레이드) 애프터아워 전용 히트맵 — 정규장 마감가(15:30) 대비 등락률 ──
_CLOSE_SNAPSHOT_PATH = ROOT / "output" / "regular_close_prices.json"


def _load_close_snapshot() -> dict:
    """오늘자 정규장 마감가 스냅샷 로드. 날짜 안 맞으면(장 시작 전 등) 빈 dict."""
    try:
        import json as _json
        with open(str(_CLOSE_SNAPSHOT_PATH), encoding="utf-8") as f:
            d = _json.load(f)
        if d.get("date") == datetime.now().date().isoformat():
            return d.get("prices", {})
    except Exception:
        pass
    return {}


def _maybe_save_close_snapshot(prices: dict) -> None:
    """정규장 마감 직후(15:30~15:40) 히트맵 빌드 시 딱 한 번만 오늘자 종가 스냅샷 저장.
    이 시각 안에는 여러 번 불려도 이미 저장돼 있으면 건너뛴다."""
    now = datetime.now()
    if now.weekday() >= 5:
        return
    hm = now.hour * 60 + now.minute
    if not (15 * 60 + 30 <= hm < 15 * 60 + 40):
        return
    today = now.date().isoformat()
    try:
        import json as _json
        with open(str(_CLOSE_SNAPSHOT_PATH), encoding="utf-8") as f:
            existing = _json.load(f)
        if existing.get("date") == today:
            return
    except Exception:
        pass
    snap = {c: p.get("price", 0) for c, p in prices.items() if (p or {}).get("price")}
    if not snap:
        return
    try:
        _CLOSE_SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(str(_CLOSE_SNAPSHOT_PATH), "w", encoding="utf-8") as f:
            _json.dump({"date": today, "prices": snap}, f, ensure_ascii=False)
    except Exception:
        pass


def _load_sector_custom() -> dict:
    if SECTOR_CUSTOM_PATH.exists():
        try:
            import json as _json
            with open(str(SECTOR_CUSTOM_PATH), encoding="utf-8") as _f:
                return _json.load(_f)
        except Exception:
            pass
    return {"custom_tiles": [], "extra_stocks": {}, "hidden_sectors": []}

# 전체 탭에 세부섹터를 '다 꺼낼' 테마 부모들(사용자 요청). 메타(ETF*)는 제외 유지.
# 각 부모의 서브섹터(예: 반도체 테마별→HBM·CXL·액침냉각·온디바이스AI)를 독립 타일로 표시.
SURFACE_THEME_PARENTS = [
    "반도체 테마별", "블랙테마 원자재", "코인/STO", "메타버스/6G",
    "코로나 전염병", "가구 인테리어", "정책", "대왕고래",
]

# 전체 탭에서 대섹터를 서브섹터 타일로 '분할'(사용자 요청). 뭉친 타일 대신 서브별 타일.
# key=표시(대섹터)명, value=raw 부모명 리스트(MERGE 전 원본, 예: 로봇=로봇-1+로봇-2).
SPLIT_SECTORS = {
    "에너지 신재생":     ["에너지 신재생"],       # 수소·SOFC·태양광·풍력
    "에너지 전선/변압기": ["에너지 전선/변압기"],   # 변압기·전선
    "AI 보안/양자":      ["AI 보안/양자"],         # 보안·양자컴퓨터
    "우주방산":          ["우주방산"],             # 방산·우주
    "로봇":              ["로봇-1", "로봇-2"],      # 삼성향·현대향… + 의료용·방산용·엑츄에이터…
}

# ── 제외 섹터 (전체 탭용) ─────────────────────────────────
EXCLUDE = {
    "ETF 시장전체", "ETF시장 체크",
    "블랙테마 원자재", "블랙테마 전쟁",  # SECTOR_EXTRACT로 일부 추출
    "코로나 전염병", "코인/STO",         # STO만 SECTOR_EXTRACT
    "메타버스/6G",                       # 광통신 6G만 SECTOR_EXTRACT
    "반도체 테마별",                      # 서브탭에서만 일부 포함
    "가구 인테리어",
    "트럼프<재건/경협>",
    "정책",    # [수정2] 삭제
    "대왕고래", # [수정8] 알레스카LNG 삭제
}

RENAME = {
    "소비재": "의류 소비재",
    # 대왕고래→알레스카LNG 제거 (EXCLUDE로 이동)
}

MERGE = {
    "바이오 섹터별":   "바이오",
    "바이오 섹터별-2": "바이오",
    "이차전지-1":     "이차전지",
    "이차전지-2":     "이차전지",
    "로봇-1":        "로봇",   # [수정1] 통합
    "로봇-2":        "로봇",
}

TRUMP_SUB_MAP = {"남북경협": "남북경협", "재건": "전쟁재건"}

# ── 제외된 섹터에서 특정 서브섹터 추출 → 독립 메인섹터로 ────
# [수정3,4,5,6]
SECTOR_EXTRACT = [
    {"name": "광통신", "from": "메타버스/6G",   "sub": "광통신 6G"},
    {"name": "STO",   "from": "코인/STO",      "sub": "STO"},
    {"name": "해운",   "from": "블랙테마 전쟁",  "sub": "해운"},
    {"name": "전쟁",   "from": "블랙테마 전쟁",  "sub": "전쟁"},
]

# ── 서브탭: 특정 (탭ID, 대섹터) 내 제외 서브섹터 ─────────────
TAB_SUB_EXCLUDE: dict = {
    ("semi", "반도체 디스플레이"): {"통신"},         # [서브3]
    ("semi", "반도체 테마별"):   {"HBM", "AGI 뉴럴링크"},  # CXL·액침냉각·온디바이스만
}

# ── 서브탭: 특정 (탭ID, 대섹터)의 모든 서브섹터를 하나의 타일로 통합 ─
TAB_TILE_MERGE: dict = {
    ("semi",   "반도체 데이터센터"): "데이터센터",  # [서브4]
    ("beauty", "뷰티 화장품"):     "화장품",        # [서브7]
    ("beauty", "뷰티 미용"):       "미용",          # [서브7]
    ("theme",  "식품"):           "식품",           # [서브8]
}

# ── 서브탭: 서브섹터명 변경 (대섹터명, 서브명) → 표시명 ─────
SUB_RENAME: dict = {
    ("에너지 원전", "대장"):   "원전 대장주",    # [서브5]
    ("에너지 원전", "중소형"): "원전 중소형주",  # [서브6]
}

# ── 하드코딩 타일 (특정 탭에 고정 삽입) ──────────────────────
HARDCODED_TILES: dict = {
    "semi": [
        {   # [서브2] MLCC/기판
            "name": "MLCC/기판", "parent": "부품",
            "stocks": [
                {"name": "삼성전기",   "code": "009150"},
                {"name": "LG이노텍",   "code": "011070"},
                {"name": "삼화콘덴서", "code": "001820"},
            ],
        },
    ],
}

# ── 탭 그룹 정의 ─────────────────────────────────────────
TAB_GROUPS = [
    {"id": "all",     "label": "전체"},
    {"id": "semi",    "label": "반도체",   "sectors": [
        "반도체 설계 및 파운드리", "반도체 전공정",
        "반도체 후공정", "반도체 디스플레이", "반도체 데이터센터",
        "반도체 테마별",  # [서브1] CXL·액침냉각·온디바이스AI
    ]},
    {"id": "ship",    "label": "조선/LNG", "sectors": ["조선", "LNG"]},
    {"id": "def",     "label": "방산/우주", "sectors": ["우주방산"]},
    {"id": "robot",   "label": "로봇",     "sectors": ["로봇-1", "로봇-2"]},
    {"id": "bio",     "label": "바이오",   "sectors": [
        "바이오 섹터별", "바이오 섹터별-2", "바이오 비만당뇨치매",
    ]},
    {"id": "battery", "label": "이차전지", "sectors": ["이차전지-1", "이차전지-2"]},
    {"id": "energy",  "label": "에너지",   "sectors": [
        "에너지 전선/변압기", "에너지 원전", "에너지 신재생",
    ]},
    {"id": "beauty",  "label": "뷰티",     "sectors": ["뷰티 화장품", "뷰티 미용"]},
    {"id": "ai",      "label": "AI",       "sectors": [
        "AI 소프트웨어", "AI 의료", "AI 보안/양자",
    ]},
    {"id": "auto",    "label": "자동차",   "sectors": ["자동차", "자동차 자율주행"]},
    {"id": "theme",   "label": "테마",     "sectors": [
        "트럼프<재건/경협>",
        "식품", "여행/항공", "웹툰/게임", "엔터", "소비재",
        "육계/수산", "원자재", "시멘트/페인트",
        # 대왕고래·정책 제거 [서브8]
    ]},
]

TAB_BY_ID = {t["id"]: t for t in TAB_GROUPS}

PARENT_LABEL = {
    "반도체 설계 및 파운드리": "설계/파운드리",
    "반도체 전공정": "전공정",
    "반도체 후공정": "후공정",
    "반도체 디스플레이": "디스플레이",
    "반도체 데이터센터": "데이터센터",
    "반도체 테마별": "테마",
    "에너지 전선/변압기": "전선/변압기",
    "에너지 원전": "원전",
    "에너지 신재생": "신재생",
    "뷰티 화장품": "화장품",
    "뷰티 미용": "미용",
    "AI 소프트웨어": "소프트웨어",
    "AI 의료": "AI의료",
    "AI 보안/양자": "보안/양자",
    "자동차 자율주행": "자율주행",
    "바이오 섹터별": "바이오①",
    "바이오 섹터별-2": "바이오②",
    "바이오 비만당뇨치매": "비만/당뇨",
    "이차전지-1": "이차전지①",
    "이차전지-2": "이차전지②",
    "로봇-1": "로봇①",
    "로봇-2": "로봇②",
    "소비재": "의류소비재",
}


# ── 파서 ─────────────────────────────────────────────────
def _parse_raw_full() -> dict:
    """xlsx → {대섹터명: [{name: 서브명, stocks: [...]}]}"""
    import openpyxl
    wb = openpyxl.load_workbook(str(WATCH_FILE), read_only=True, data_only=True)
    ws = wb["Sheet1"]

    result: dict = {}
    cur_sector = None
    cur_subsec = None
    next_is_major = False

    for row in ws.iter_rows(values_only=True):
        c1, c2, c5 = row[0], row[1], row[4]

        if c1 == "L일봉H":
            next_is_major = True
            cur_subsec = None
            continue

        code_str = str(c5 or "").strip()
        is_stock = code_str.isdigit() and 1 <= len(code_str) <= 6

        if is_stock:
            name = str(c2 or "").strip()
            code = code_str.zfill(6)
            if cur_sector and name:
                if cur_subsec is None:
                    subs = result.setdefault(cur_sector, [])
                    if not subs:
                        subs.append({"name": "", "stocks": []})
                    subs[0]["stocks"].append({"name": name, "code": code})
                else:
                    cur_subsec["stocks"].append({"name": name, "code": code})
        elif c1:
            label = str(c1).strip()
            if not label or label.startswith("0 0"):
                continue
            if next_is_major:
                cur_sector = label
                next_is_major = False
                cur_subsec = None
                if cur_sector not in result:
                    result[cur_sector] = []
            else:
                if cur_sector is not None:
                    cur_subsec = {"name": label, "stocks": []}
                    result[cur_sector].append(cur_subsec)

    wb.close()
    return result


def _dedup_stocks(stocks: list) -> list:
    seen, out = set(), []
    for s in stocks:
        if s["code"] not in seen:
            seen.add(s["code"])
            out.append(s)
    return out


def _shorten_sub(name: str, parent: str) -> str:
    for prefix in [parent + " ", parent]:
        if name.startswith(prefix):
            return name[len(prefix):].strip(" /·-")
    return name


# ── 전체 탭 파싱 ─────────────────────────────────────────
def parse_watchlist(top_n: int = 3) -> list:
    raw = _parse_raw_full()
    merged: dict = {}

    for sec_name, subsecs in raw.items():
        if sec_name.startswith("트럼프"):
            for sub in subsecs:
                key = None
                for k, v in TRUMP_SUB_MAP.items():
                    if k in sub["name"]:
                        key = v
                        break
                if key:
                    merged.setdefault(key, [])
                    merged[key].extend(sub["stocks"])
            continue

        if sec_name in EXCLUDE:
            continue

        target = MERGE.get(sec_name, sec_name)
        target = RENAME.get(target, target)
        merged.setdefault(target, [])
        for sub in subsecs:
            merged[target].extend(sub["stocks"])

    # 제외된 섹터에서 서브섹터 추출
    for ext in SECTOR_EXTRACT:
        raw_subs = raw.get(ext["from"], [])
        stocks = []
        for sub in raw_subs:
            if sub["name"] == ext["sub"]:
                stocks.extend(sub["stocks"])
        if stocks:
            merged.setdefault(ext["name"], [])
            merged[ext["name"]].extend(stocks)

    result = []
    for sec_name, stocks in merged.items():
        deduped = _dedup_stocks(stocks)
        result.append({"sector": sec_name, "stocks": deduped[:top_n]})
    return result


def parse_etf_bar() -> list:
    """ETF시장 체크 섹션 파싱 (코드/이름/OHLC)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(WATCH_FILE), read_only=True, data_only=True)
    ws = wb["Sheet1"]
    in_etf = False
    etfs = []
    for row in ws.iter_rows(values_only=True):
        c1, c2, c4, c5 = row[0], row[1], row[3], row[4]
        if str(c1 or "").strip() == "ETF시장 체크":
            in_etf = True
            continue
        if c1 == "L일봉H" and in_etf:
            break
        if in_etf:
            code_str = str(c5 or "").strip()
            if code_str.isdigit() and 1 <= len(code_str) <= 6:
                try:
                    ohlc = [float(v) for v in str(c1).split()]
                except Exception:
                    ohlc = []
                etfs.append({
                    "name": str(c2 or "").strip(),
                    "code": code_str.zfill(6),
                    "ohlc": ohlc,
                    "change_rate": 0.0,  # KIS API로 실시간 조회
                    "price": 0,
                })
    wb.close()
    return etfs


# ── 탭별 세부 히트맵 ──────────────────────────────────────
def build_heatmap_tab(tab_id: str) -> dict:
    import sys
    sys.path.insert(0, str(ROOT / "scripts"))
    import kis_api

    tab = TAB_BY_ID.get(tab_id)
    if not tab:
        return {"error": f"알 수 없는 탭: {tab_id}", "tiles": []}

    raw = _parse_raw_full()
    tiles_raw = []

    for sec_name in (tab.get("sectors") or []):
        subsecs = raw.get(sec_name, [])

        # 트럼프 섹션 → 남북경협/전쟁재건 분리
        if sec_name.startswith("트럼프"):
            for sub in subsecs:
                display = None
                for k, v in TRUMP_SUB_MAP.items():
                    if k in sub["name"]:
                        display = v
                        break
                if display and sub["stocks"]:
                    tiles_raw.append({
                        "name": display, "parent": "",
                        "stocks": _dedup_stocks(sub["stocks"]),
                    })
            continue

        parent_label = PARENT_LABEL.get(sec_name, sec_name)

        # TAB_TILE_MERGE: 서브섹터 전체를 하나의 타일로
        merge_name = TAB_TILE_MERGE.get((tab_id, sec_name))
        if merge_name is not None:
            all_stocks = _dedup_stocks([s for sub in subsecs for s in sub["stocks"]])
            if all_stocks:
                tiles_raw.append({
                    "name": merge_name, "parent": parent_label,
                    "stocks": all_stocks,
                })
            continue

        # 일반: 서브섹터별 타일
        sub_exclude = TAB_SUB_EXCLUDE.get((tab_id, sec_name), set())
        has_named = any(s["name"] for s in subsecs)

        if has_named:
            for sub in subsecs:
                if not sub["stocks"]:
                    continue
                if sub["name"] in sub_exclude:
                    continue
                # 이름 변경 → 단축 → 원본 순으로
                renamed = SUB_RENAME.get((sec_name, sub["name"]))
                if renamed:
                    tile_name = renamed
                elif sub["name"]:
                    tile_name = _shorten_sub(sub["name"], sec_name)
                else:
                    tile_name = parent_label
                tiles_raw.append({
                    "name": tile_name, "parent": parent_label,
                    "stocks": _dedup_stocks(sub["stocks"]),
                })
        else:
            all_stocks = _dedup_stocks([s for sub in subsecs for s in sub["stocks"]])
            if all_stocks:
                tiles_raw.append({
                    "name": parent_label, "parent": "",
                    "stocks": all_stocks,
                })

    # 하드코딩 타일
    for ht in HARDCODED_TILES.get(tab_id, []):
        tiles_raw.append({
            "name": ht["name"], "parent": ht.get("parent", ""),
            "stocks": list(ht["stocks"]),
        })

    # ── 커스텀 오버레이 (sector_custom.json) ─────────────────
    _custom = _load_sector_custom()
    _hidden_set = set(_custom.get("hidden_sectors", []))
    _extra_map = _custom.get("extra_stocks", {})      # {타일명: [{name, code}]}
    _custom_tile_conf = _custom.get("custom_tiles", [])

    tiles_raw = [t for t in tiles_raw if t["name"] not in _hidden_set]

    for _tile in tiles_raw:
        _existing_codes = {s["code"] for s in _tile["stocks"]}
        for _es in _extra_map.get(_tile["name"], []):
            _c = _es.get("code", "")
            if _c and _c not in _existing_codes:
                _tile["stocks"].append({"name": _es.get("name", _c), "code": _c})
                _existing_codes.add(_c)

    _existing_tile_names = {t["name"] for t in tiles_raw}
    for _ct in _custom_tile_conf:
        _ct_name = _ct.get("name", "")
        if _ct_name and _ct_name not in _existing_tile_names:
            tiles_raw.append({
                "name": _ct_name,
                "parent": _ct.get("parent", "커스텀"),
                "stocks": [{"name": s.get("name", ""), "code": s.get("code", "")}
                           for s in _ct.get("stocks", []) if s.get("code")],
            })
    # ─────────────────────────────────────────────────────────

    # 병렬 가격 조회
    all_codes = list({s["code"] for t in tiles_raw for s in t["stocks"]})
    prices = kis_api.get_prices_batch_parallel(all_codes)

    result = []
    for tile in tiles_raw:
        rates = [float((prices.get(s["code"]) or {}).get("change_rate", 0) or 0)
                 for s in tile["stocks"]]
        avg = round(sum(rates) / len(rates), 2) if rates else 0
        stocks_out = []
        for s in tile["stocks"][:3]:
            p = prices.get(s["code"]) or {}
            stocks_out.append({
                "name": s["name"], "code": s["code"],
                "change_rate": float(p.get("change_rate", 0) or 0),
                "price": p.get("price", 0),
            })
        result.append({
            "name": tile["name"],
            "parent": tile["parent"],
            "avg_rate": avg,
            "stocks": stocks_out,
        })

    result.sort(key=lambda x: x["avg_rate"], reverse=True)

    return {
        "tab_id": tab_id,
        "tab_label": tab.get("label", ""),
        "tiles": result,
        "updated_at": datetime.now().strftime("%H:%M:%S"),
    }


# ── 전체 탭 히트맵 (overview) ─────────────────────────────
def _collect_surfaced_subs(raw_full: dict) -> list:
    """제외 테마 부모들의 세부섹터 목록 → [(표시명, parent, [stocks])].
    build_heatmap(타일생성)과 surfaced_sub_names(편집목록)가 같은 결과를 쓰도록 공용화."""
    extract_subs = {e["sub"] for e in SECTOR_EXTRACT}   # 이미 메인추출(광통신6G·STO·해운·전쟁) 중복방지
    out, seen = [], set()
    for parent in SURFACE_THEME_PARENTS:
        for sub in raw_full.get(parent, []):
            nm = (sub.get("name") or "").strip() or parent
            if nm in seen or nm in extract_subs:
                continue
            if re.sub(r"[\d\s]", "", nm) == "":   # 이름이 숫자·코드뿐(깨진 서브) → 제외
                continue
            sts = _dedup_stocks(sub.get("stocks") or [])
            if not sts:
                continue
            seen.add(nm)
            out.append((nm, parent, sts))
    return out


def _collect_split_subs(raw_full: dict) -> list:
    """분할 대상 대섹터(SPLIT_SECTORS)의 서브섹터 목록 → [(표시명, 대섹터, [stocks])].
    무명 서브는 대섹터명 사용. 이름이 코드뿐이면 제외."""
    out, seen = [], set()
    for disp, parents in SPLIT_SECTORS.items():
        for parent in parents:
            for sub in raw_full.get(parent, []):
                nm = (sub.get("name") or "").strip() or disp
                if nm in seen:
                    continue
                if re.sub(r"[\d\s]", "", nm) == "":
                    continue
                sts = _dedup_stocks(sub.get("stocks") or [])
                if not sts:
                    continue
                seen.add(nm)
                out.append((nm, disp, sts))
    return out


def surfaced_sub_names() -> list:
    """전체 탭에 꺼낸 세부테마+분할 서브 이름들(편집탭 목록·복원용). 가격조회 없이 빠름."""
    try:
        raw = _parse_raw_full()
        names = [nm for nm, _, _ in _collect_surfaced_subs(raw)]
        names += [nm for nm, _, _ in _collect_split_subs(raw) if nm not in names]
        return names
    except Exception:
        return []


def build_heatmap(top_n: int = 3, mode: str = "regular") -> dict:
    """mode='nxt' — 15:30~20:00 애프터아워 전용. 정규장 마감가(15:30) 대비 등락률로
    재계산해 NXT 세션만의 움직임을 보여준다(종가베팅 후보 스크리닝용)."""
    import sys
    sys.path.insert(0, str(ROOT / "scripts"))
    import kis_api

    # 커스텀 오버레이 로드
    custom = _load_sector_custom()
    hidden_set = set(custom.get("hidden_sectors", []))
    extra_map = custom.get("extra_stocks", {})       # {섹터명: [{name, code}]}
    removed_map = custom.get("removed_stocks", {})   # {섹터명: [code, ...]}
    custom_tile_conf = custom.get("custom_tiles", [])

    sections = parse_watchlist(999)  # 전 종목 조회 (top_n은 표시용, 조회는 전체)

    raw_full = _parse_raw_full()
    # ── 제외됐던 테마의 세부섹터를 전체 탭에도 별도 타일로 '다 꺼내기'(사용자가 편집탭에서 숨김) ──
    surfaced_subs = _collect_surfaced_subs(raw_full)   # [(표시명, parent, [stocks])]
    # ── 특정 대섹터를 서브섹터 타일로 분할(에너지신재생·로봇·전선변압기·AI보안양자·우주방산) ──
    split_subs = _collect_split_subs(raw_full)

    # 기본 + extra + custom + 세부테마 코드 일괄 수집 → 단일 배치 조회
    base_codes = {s["code"] for sec in sections for s in sec["stocks"]}
    extra_codes = {s["code"] for lst in extra_map.values() for s in lst if s.get("code")}
    ct_codes = {s["code"] for ct in custom_tile_conf for s in ct.get("stocks", []) if s.get("code")}
    sub_codes = {s["code"] for _, _, sts in (surfaced_subs + split_subs) for s in sts if s.get("code")}
    codes = list(base_codes | extra_codes | ct_codes | sub_codes)
    prices = kis_api.get_prices_batch_parallel(codes)
    # KIS가 완전히 죽어있으면(2026-07-04 KIS 서버 장애처럼) 전종목이 price=0으로 돌아온다 —
    # 그럴 땐 네이버 공개 API(인증불필요)로 마지막 거래일 종가·등락률을 대신 채운다.
    if not any((p or {}).get("price") for p in prices.values()):
        import naver_api
        naver = naver_api.last_session_batch(codes)
        prices = {c: (r and {"price": r["price"], "change_rate": r["change_rate"]}) or {}
                  for c, r in naver.items()}

    _maybe_save_close_snapshot(prices)
    close_snap = _load_close_snapshot() if mode == "nxt" else {}

    def _rate_for(code: str, p: dict) -> float:
        if mode == "nxt":
            cp = close_snap.get(code, 0)
            cur = (p or {}).get("price", 0)
            if cp and cur:
                return round((cur - cp) / cp * 100, 2)
            return 0.0
        return float((p or {}).get("change_rate", 0) or 0)

    sectors = []
    for sec in sections:
        if sec["sector"] in hidden_set or sec["sector"] in SPLIT_SECTORS:  # 분할대상은 뭉친타일 생략
            continue
        removed_codes = set(removed_map.get(sec["sector"], []))
        items = []
        for s in sec["stocks"]:
            if s["code"] in removed_codes:
                continue
            p = prices.get(s["code"]) or {}
            rate = _rate_for(s["code"], p)
            items.append({"name": s["name"], "code": s["code"],
                          "change_rate": rate, "price": p.get("price", 0)})
        # extra_stocks 추가
        existing_codes = {s["code"] for s in items}
        for es in extra_map.get(sec["sector"], []):
            c = es.get("code", "")
            if c and c not in existing_codes:
                p = prices.get(c) or {}
                rate = _rate_for(c, p)
                items.append({"name": es.get("name", c), "code": c,
                              "change_rate": rate, "price": p.get("price", 0)})
                existing_codes.add(c)
        # 등락률 내림차순 정렬
        items = sorted(items, key=lambda x: x["change_rate"], reverse=True)
        # avg: 실제 가격 데이터 있는 종목만 (0% = 데이터 없음 제외)
        valid_rates = [x["change_rate"] for x in items if x.get("price", 0) != 0]
        avg = round(sum(valid_rates) / len(valid_rates), 2) if valid_rates else 0
        sectors.append({"name": sec["sector"], "avg_rate": avg, "stocks": items})

    # 커스텀 타일 추가
    existing_names = {s["name"] for s in sectors}
    for ct in custom_tile_conf:
        ct_name = ct.get("name", "")
        if not ct_name or ct_name in existing_names:
            continue
        ct_items, ct_rates = [], []
        for s in ct.get("stocks", []):
            c = s.get("code", "")
            if not c:
                continue
            p = prices.get(c) or {}
            rate = _rate_for(c, p)
            ct_items.append({"name": s.get("name", c), "code": c,
                             "change_rate": rate, "price": p.get("price", 0)})
            ct_rates.append(rate)
        avg = round(sum(ct_rates) / len(ct_rates), 2) if ct_rates else 0
        sectors.append({"name": ct_name, "avg_rate": avg,
                        "stocks": ct_items, "parent": ct.get("parent", "커스텀")})

    # 세부섹터 타일(꺼낸 테마 + 분할 대섹터) 추가. 이미 있는 이름은 skip. extra/removed 오버레이 반영.
    existing_names = {s["name"] for s in sectors}

    def _emit_sub_tiles(sub_list):
        for nm, parent, sts in sub_list:
            if nm in existing_names or nm in hidden_set:
                continue
            removed_c = set(removed_map.get(nm, []))
            s_items = []
            for s in sts:
                c = s.get("code", "")
                if not c or c in removed_c:
                    continue
                p = prices.get(c) or {}
                s_items.append({"name": s.get("name", c), "code": c,
                                "change_rate": _rate_for(c, p),
                                "price": p.get("price", 0)})
            exist_c = {x["code"] for x in s_items}
            for es in extra_map.get(nm, []):
                c = es.get("code", "")
                if c and c not in exist_c and c not in removed_c:
                    p = prices.get(c) or {}
                    s_items.append({"name": es.get("name", c), "code": c,
                                    "change_rate": _rate_for(c, p),
                                    "price": p.get("price", 0)})
                    exist_c.add(c)
            if not s_items:
                continue
            s_items.sort(key=lambda x: x["change_rate"], reverse=True)
            valid = [x["change_rate"] for x in s_items if x.get("price", 0) != 0]
            avg = round(sum(valid) / len(valid), 2) if valid else 0
            sectors.append({"name": nm, "avg_rate": avg, "stocks": s_items, "parent": parent})
            existing_names.add(nm)

    _emit_sub_tiles(surfaced_subs)
    _emit_sub_tiles(split_subs)

    sectors.sort(key=lambda x: x["avg_rate"], reverse=True)

    # 각 섹터 대표 종목(가격 데이터 있는 첫 번째) 일봉 bars 병렬 조회
    try:
        from concurrent.futures import ThreadPoolExecutor as _TPE
        rep = {}  # {섹터명: code}
        for sec in sectors:
            for s in sec["stocks"]:
                if s.get("price", 0) > 0:
                    rep[sec["name"]] = s["code"]
                    break
        bars_map = {}
        if rep:
            with _TPE(max_workers=10) as ex:
                futs = {ex.submit(kis_api.get_daily_bars, code, 20): name
                        for name, code in rep.items()}
                for fut, name in futs.items():
                    try:
                        bars_map[name] = fut.result(timeout=12)
                    except Exception:
                        bars_map[name] = []
        for sec in sectors:
            sec["bars"] = bars_map.get(sec["name"], [])
    except Exception:
        for sec in sectors:
            sec.setdefault("bars", [])

    # 표시이름 변경(편집탭): disp에만 반영, name은 원본 유지(숨김·추가·삭제 키 일관성)
    rename_map = custom.get("sector_rename", {})
    for sec in sectors:
        sec["disp"] = rename_map.get(sec["name"], sec["name"])

    # 종목 표시명 변경(코드 기준): 어느 섹터에 있든 동일 적용 (예: 삼화콘덴서공업→삼화콘덴서)
    stock_rename = custom.get("stock_rename", {})
    if stock_rename:
        for sec in sectors:
            for st in sec.get("stocks", []):
                nn = stock_rename.get(st.get("code"))
                if nn:
                    st["name"] = nn

    return {
        "sectors": sectors,
        "updated_at": datetime.now().strftime("%H:%M:%S"),
        "source": "내 관심종목.xlsx",
    }

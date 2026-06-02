"""
viz_card.py — 카드뉴스 스타일 목표주가 카드 생성

사용법:
    python scripts/viz_card.py LG이노텍
    python scripts/viz_card.py LG이노텍 --png     ← PNG 저장
    python scripts/viz_card.py LG이노텍 SK하이닉스  ← 여러 종목
"""

import sys, argparse, subprocess
from pathlib import Path
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

ROOT      = Path(__file__).parent.parent
EXCEL_DIR = ROOT / "raw" / "매일 엑셀넣을것"
OUT_DIR   = ROOT / "out"
OUT_DIR.mkdir(exist_ok=True)
TODAY     = datetime.today().strftime("%Y-%m-%d")
TODAY_KR  = datetime.today().strftime("%Y년 %m월 %d일")
RECENT    = 45

RANK_COLORS = ["#c0392b", "#1565c0", "#1a7a4a", "#7b1fa2", "#e65100"]
RANK_LABEL  = ["1위", "2위", "3위", "4위", "5위"]

# ── 엑셀 파싱 ─────────────────────────────────────────────────────

def find_excel(pattern):
    for f in EXCEL_DIR.iterdir():
        if f.name.startswith("~$"): continue
        if pattern in f.name and f.suffix in (".xlsx", ".xlsm"):
            return f
    return None

def load_all(wb):
    records = []
    sheet_types = {
        "Rating_Up":"rating_up","Rating_Down":"rating_down",
        "TP_Up":"tp_up","TP_Down":"tp_down",
        "New_KOSPI":"new","New_KOSDAQ":"new",
    }
    for sname, stype in sheet_types.items():
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        rows = [list(r) for r in ws.iter_rows(min_row=1, max_row=300, max_col=16, values_only=True)
                if any(v is not None for v in r)]
        hi = next((i for i,r in enumerate(rows)
                   if any(str(v) in ["일자","코드","From"] for v in r if v)), None)
        if hi is None: continue
        for r in rows[hi+2:]:
            dv = r[1] if len(r)>1 else None
            cv = r[2] if len(r)>2 else None
            nv = r[3] if len(r)>3 else None
            if not isinstance(dv, datetime): continue
            if not isinstance(cv, str) or not cv.startswith("A"): continue
            if not nv: continue
            tp_old = int(r[13]) if len(r)>13 and isinstance(r[13],(int,float)) else None
            tp_new = int(r[14]) if len(r)>14 and isinstance(r[14],(int,float)) else None
            if not tp_new: continue
            op_old = str(r[11]).strip() if len(r)>11 and r[11] else "—"
            op_new = str(r[12]).strip() if len(r)>12 and r[12] else "—"
            records.append({
                "date": dv.strftime("%Y-%m-%d"),
                "name": str(nv).strip(),
                "brokerage": str(r[5]).strip() if len(r)>5 and r[5] else "—",
                "op_old": op_old, "op_new": op_new,
                "tp_old": tp_old, "tp_new": tp_new, "stype": stype,
            })
    return records

def get_top3(records, stock_name, n=3):
    cutoff = (datetime.today()-timedelta(days=RECENT)).strftime("%Y-%m-%d")
    items  = [r for r in records if stock_name in r["name"] or r["name"] in stock_name]
    latest = {}
    for it in sorted(items, key=lambda x: x["date"]):
        latest[it["brokerage"]] = it
    recent = {b:it for b,it in latest.items() if it["date"] >= cutoff} or latest
    ranked = sorted(recent.values(), key=lambda x: x["tp_new"], reverse=True)[:n]
    all_tp = [it["tp_new"] for it in recent.values()]
    consensus = int(sum(all_tp)/len(all_tp)) if all_tp else 0
    return ranked, consensus, len(recent)

# ── 코멘터리 자동 생성 ──────────────────────────────────────────────

def make_comment(broker, tp_new, tp_old, pct, rank):
    if tp_old:
        updown = f"{abs(pct):.0f}% 상승여력"
    else:
        updown = "신규 커버리지"
    if rank == 0:
        return f"{broker}은 압도적인 최선호주로 {updown} 제시."
    elif pct >= 80:
        return f"{broker}은 구조적 성장 반영, 목표가 대폭 상향으로 {updown} 제시."
    elif pct >= 30:
        return f"{broker}은 실적 모멘텀을 근거로 {updown} 제시."
    elif pct >= 0:
        return f"{broker}은 안정적인 이익 성장을 근거로 {updown} 제시."
    else:
        return f"{broker}은 보수적인 관점에서도 목표가 유지."

# ── HTML 빌드 ──────────────────────────────────────────────────────

def build_card(stock_name, top3, consensus, n_brokers):
    if not top3:
        return "<h2>데이터 없음</h2>"

    max_tp = top3[0]["tp_new"]

    # 랭킹 카드 + 코멘터리 HTML
    rank_cards = ""
    comments   = ""

    for i, it in enumerate(top3):
        color  = RANK_COLORS[i]
        pct    = round((it["tp_new"]-it["tp_old"])/it["tp_old"]*100) if it["tp_old"] else 0
        bar_w  = round(it["tp_new"] / max_tp * 100)
        old_str = f"{it['tp_old']//10000}만원<br>→" if it["tp_old"] else "신규<br>→"
        sign   = "▲" if pct >= 0 else "▼"
        pct_color = "#c0392b" if pct >= 0 else "#1565c0"
        rating_change = ""
        if it["op_old"] != it["op_new"] and it["op_old"] != "—":
            rating_change = f'<div class="rating-chg">{it["op_old"]} → {it["op_new"]}</div>'

        rank_cards += f"""
        <div class="rank-card" style="border-left:5px solid {color}">
          <div class="rank-left">
            <div class="rank-num" style="color:{color}">{RANK_LABEL[i]}</div>
            <div class="rank-prev">{old_str}</div>
          </div>
          <div class="rank-right">
            <div class="broker-row">
              <span class="flag">🇰🇷</span>
              <span class="broker-name">{it['brokerage']}</span>
              {rating_change}
            </div>
            <div class="tp-value" style="color:{color}">{it['tp_new']:,}원</div>
            <div class="pct-row">
              <span class="pct-badge" style="color:{pct_color}">{sign}+{abs(pct)}%</span>
            </div>
            <div class="bar-track">
              <div class="bar-fill" style="width:{bar_w}%;background:{color}"></div>
            </div>
          </div>
        </div>"""

        comments += f'<p>{make_comment(it["brokerage"], it["tp_new"], it["tp_old"], pct, i)}</p>\n'

    return f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8">
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;900&display=swap');
*{{box-sizing:border-box;margin:0;padding:0}}
html{{margin:0;padding:0;background:#1a1a2e}}
body{{
  margin:0;padding:0;
  background:#1a1a2e;
  font-family:'Noto Sans KR','Malgun Gothic',sans-serif;
  -webkit-font-smoothing:antialiased;
  line-height:1;
  display:inline-block;
}}
.card{{
  width:420px;background:#fff;
  display:flex;flex-direction:column;
}}

/* ── 희생 구역 (텔레그램 상단 크롭 흡수) ── */
.sacrifice-zone{{
  background:#1a1a2e;
  height:52px;
  display:flex;align-items:center;
  padding:0 16px;gap:8px;
  flex-shrink:0;
}}
.brand-text{{font-size:12px;font-weight:800;color:#fff;letter-spacing:.5px}}
.brand-div{{font-size:11px;color:#475569}}
.brand-sub{{font-size:11px;color:#94a3b8}}

/* ── 타이틀 영역 ── */
.title-area{{
  padding:10px 16px 8px;
  background:#fff;
  border-bottom:3px solid #1a1a2e;
}}
.stock-name{{
  font-size:26px;font-weight:900;color:#1a1a2e;
  letter-spacing:-1px;line-height:1.1;
}}
.card-subtitle{{
  font-size:16px;font-weight:700;color:#1a1a2e;
  margin-top:1px;
}}
.meta-line{{font-size:10px;color:#888;margin-top:4px}}
.date-line{{font-size:9px;color:#aaa;margin-top:1px}}

/* ── 종합통계 박스 ── */
.stats-box{{
  background:#f8f9fa;
  margin:8px 14px;
  border-radius:5px;
  overflow:hidden;
  border:1px solid #e9ecef;
}}
.stats-header{{
  background:#e9ecef;text-align:center;
  font-size:11px;font-weight:700;color:#495057;
  padding:4px;letter-spacing:1px;
}}
.stats-row{{
  display:flex;justify-content:space-between;
  padding:5px 12px;border-bottom:1px solid #e9ecef;
  font-size:11px;
}}
.stats-row:last-child{{border-bottom:none}}
.stats-label{{color:#666}}
.stats-val{{font-weight:700;color:#1a1a2e}}
.stats-val.red{{color:#c0392b}}

/* ── 랭킹 카드 ── */
.rank-card{{
  display:flex;
  margin:5px 14px 0;
  background:#fff;
  border:1px solid #e9ecef;
  border-radius:4px;
  overflow:hidden;
  border-left-width:5px !important;
}}
.rank-left{{
  width:44px;flex-shrink:0;
  background:#fafafa;
  border-right:1px solid #f0f0f0;
  display:flex;flex-direction:column;
  align-items:center;justify-content:center;
  padding:5px 3px;text-align:center;
}}
.rank-num{{font-size:12px;font-weight:900;margin-bottom:3px}}
.rank-prev{{font-size:9px;color:#aaa;line-height:1.4}}
.rank-right{{flex:1;padding:5px 10px 5px}}
.broker-row{{display:flex;align-items:center;gap:4px;margin-bottom:1px}}
.flag{{font-size:11px}}
.broker-name{{font-size:11px;font-weight:700;color:#333}}
.rating-chg{{font-size:9px;color:#888;margin-left:3px}}
.tp-value{{font-size:21px;font-weight:900;letter-spacing:-1px;line-height:1.1}}
.pct-row{{margin-top:1px}}
.pct-badge{{font-size:12px;font-weight:700}}
.bar-track{{height:4px;background:#f0f0f0;border-radius:0;margin-top:5px}}
.bar-fill{{height:4px;border-radius:0}}

/* ── 코멘터리 ── */
.commentary{{
  margin:8px 14px 0;
  padding:8px 12px;
  background:#f8f9fa;
  border-radius:4px;
  border-left:3px solid #dee2e6;
}}
.commentary p{{
  font-size:10.5px;color:#444;line-height:1.6;
  margin-bottom:4px;
}}
.commentary p:last-child{{margin-bottom:0}}

/* ── 하단 푸터 ── */
.footer{{
  margin-top:8px;
  background:#1a1a2e;
  padding:7px 14px;
  display:flex;align-items:center;justify-content:space-between;
}}
.footer-brand{{display:flex;align-items:center;gap:7px}}
.footer-icon{{
  width:22px;height:22px;background:#c0392b;
  border-radius:3px;display:flex;align-items:center;justify-content:center;
  font-size:11px;color:#fff;font-weight:900;
}}
.footer-text{{color:#fff;font-size:10px;font-weight:700}}
.footer-sub{{color:#64748b;font-size:8px;margin-top:1px}}
.footer-url{{color:#64748b;font-size:9px}}
</style>
</head>
<body>
<div class="card">

  <!-- 희생 구역: 텔레그램이 항상 위 ~50px 자름 → 여기서 흡수 -->
  <div class="sacrifice-zone">
    <span class="brand-text">STOCK BRAIN</span>
    <span class="brand-div">|</span>
    <span class="brand-sub">금융 전문가 분석</span>
  </div>

  <!-- 타이틀 (희생 구역 아래부터 보임) -->
  <div class="title-area">
    <div class="stock-name">{stock_name}</div>
    <div class="card-subtitle">추정이익 카드뉴스</div>
    <div class="meta-line">최근 {RECENT}일 기준, {n_brokers}개 증권사 분석 결과 · {TODAY_KR}</div>
  </div>

  <!-- 종합 통계 -->
  <div class="stats-box">
    <div class="stats-header">종합 통계</div>
    <div class="stats-row">
      <span class="stats-label">전문가 평균 목표주가</span>
      <span class="stats-val">{consensus:,}원</span>
    </div>
    <div class="stats-row">
      <span class="stats-label">최고 목표가 (신고가)</span>
      <span class="stats-val red">{top3[0]['tp_new']:,}원 ({top3[0]['brokerage']})</span>
    </div>
  </div>

  <!-- 랭킹 카드 -->
  {rank_cards}

  <!-- 푸터 -->
  <div class="footer">
    <div class="footer-brand">
      <div class="footer-icon">S</div>
      <div>
        <div class="footer-text">STOCK BRAIN INTELLIGENCE</div>
        <div class="footer-sub">카드뉴스 제공: SBINTEL.COM</div>
      </div>
    </div>
    <span class="footer-url">sbintel.com</span>
  </div>

</div>
</body></html>"""


def save_png(html_path: Path, png_path: Path):
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  ⚠️  playwright 없음: pip install playwright && playwright install chromium")
        return False
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 420, "height": 900})
            page.goto(html_path.as_uri(), wait_until="networkidle")
            page.wait_for_timeout(800)
            el = page.query_selector(".card")
            box = el.bounding_box()
            page.screenshot(
                path=str(png_path),
                clip={"x": box["x"], "y": box["y"],
                      "width": box["width"], "height": box["height"]}
            )
            browser.close()
        print(f"  🖼  PNG: {png_path.name}  ({int(box['width'])}×{int(box['height'])}px)")
        return True
    except Exception as e:
        print(f"  ⚠️  PNG 실패: {e}")
        return False


def send_telegram_photo(png_path: Path, caption: str = ""):
    import json
    env = {}
    ep = ROOT / ".env"
    if ep.exists():
        for line in ep.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    token   = env.get("BOT_TOKEN", "")
    chat_id = env.get("CHAT_ID", "")
    if not token or not chat_id:
        print("  ⚠️  .env BOT_TOKEN/CHAT_ID 없음"); return False

    url = f"https://api.telegram.org/bot{token}/sendPhoto"
    try:
        import requests as _req
        with open(png_path, "rb") as f:
            resp = _req.post(url, data={"chat_id": chat_id, "caption": caption},
                             files={"photo": f}, timeout=30)
        res = resp.json()
    except ImportError:
        # requests 없으면 urllib multipart
        import urllib.request
        with open(png_path, "rb") as f:
            img_data = f.read()
        boundary = b"----TGBoundary"
        body = (
            b"--" + boundary + b"\r\n"
            b'Content-Disposition: form-data; name="chat_id"\r\n\r\n' +
            chat_id.encode() + b"\r\n"
            b"--" + boundary + b"\r\n"
            b'Content-Disposition: form-data; name="caption"\r\n\r\n' +
            caption.encode("utf-8") + b"\r\n"
            b"--" + boundary + b"\r\n" +
            f'Content-Disposition: form-data; name="photo"; filename="{png_path.name}"\r\n'.encode() +
            b"Content-Type: image/png\r\n\r\n" +
            img_data + b"\r\n--" + boundary + b"--\r\n"
        )
        req = urllib.request.Request(url, data=body,
              headers={"Content-Type": f"multipart/form-data; boundary={boundary.decode()}"})
        with urllib.request.urlopen(req, timeout=30) as r:
            res = json.loads(r.read())

    if res.get("ok"):
        print("  ✅ 텔레그램 전송 완료")
        return True
    else:
        print(f"  ❌ 텔레그램 오류: {res}")
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("stocks", nargs="*")
    parser.add_argument("--png", action="store_true")
    parser.add_argument("--tg",  action="store_true", help="텔레그램으로 전송")
    parser.add_argument("--top", type=int, default=3, help="상위 N개 (기본 3)")
    args = parser.parse_args()

    path = find_excel("추정이익 변경")
    if not path:
        print("❌ 추정이익 변경 파일 없음"); sys.exit(1)

    print(f"📂 {path.name} 읽는 중...")
    wb = openpyxl.load_workbook(str(path), read_only=True, keep_vba=False, data_only=True)
    records = load_all(wb)
    wb.close()
    print(f"   {len(records)}건 로드\n")

    for stock in (args.stocks or []):
        top3, consensus, n = get_top3(records, stock, args.top)
        if not top3:
            print(f"⚠️  {stock} — 데이터 없음"); continue

        html = build_card(stock, top3, consensus, n)
        slug = f"card_{stock}_{TODAY}"
        hp   = OUT_DIR / f"{slug}.html"
        hp.write_text(html, encoding="utf-8")
        print(f"✅ {stock} (TOP{args.top}) → {hp.name}")

        pp = OUT_DIR / f"{slug}.png"
        if args.png or args.tg:
            save_png(hp, pp)

        if args.tg and pp.exists():
            caption = f"📊 {stock} 목표주가 분석\n전문가 평균: {consensus:,}원 | 신고가: {top3[0]['tp_new']:,}원\n{TODAY_KR}"
            send_telegram_photo(pp, caption)

        subprocess.Popen(["start", str(hp)], shell=True)

if __name__ == "__main__":
    main()

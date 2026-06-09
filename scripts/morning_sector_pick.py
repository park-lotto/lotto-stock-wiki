"""
morning_sector_pick.py — 유동성 컨셉 교집합 섹터 → 빈집 종목 추출

파이프라인:
  1. 유동성 컨셉 날짜별 시트 → 사모/투신/연금 공통 섹터 N일 연속 확인
  2. 해당 섹터 종목 중 수급 빈집(A/B) 추출

사용법:
    python scripts/morning_sector_pick.py
    python scripts/morning_sector_pick.py --days 3
    python scripts/morning_sector_pick.py --tg
"""

import sys, re, json, argparse
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

ROOT      = Path(__file__).parent.parent
XLSM_DIR  = ROOT / "raw" / "매일 엑셀넣을것"
WLIST_JSON = ROOT / "watchlist.json"

# calc_oscillator.py 상수 재사용
DATA_ROW_START = 15
DATA_ROW_END   = 91
STAT_ROW_P90   = 7
STAT_ROW_P75   = 8
STAT_ROW_P25   = 10
STAT_ROW_P10   = 11
STAT_COL       = 12
K12 = 2 / 13
K26 = 2 / 27
K9  = 2 / 10

# 유동성 컨셉 텍스트 → watchlist 섹터 키 매핑
LIQUIDITY_TO_SECTOR: dict[str, list[str]] = {
    "반도체":     ["반도체"],
    "피지컬ai":   ["로봇", "자동차"],
    "피지컬 ai":  ["로봇", "자동차"],
    "전력":       ["전력", "원전"],
    "조선":       ["조선"],
    "방산":       ["방산"],
    "바이오":     ["바이오"],
    "우주":       ["우주"],
    "lng":        ["LNG"],
    "이차전지":   ["이차전지"],
    "화장품":     ["화장품"],
    "소비재":     ["화장품", "엔터"],
    "엔터":       ["엔터"],
    "통신":       [],   # watchlist 없음 — 무시
    "보험":       [],
    "금융":       [],
}

GRADE_ICON = {"A 완전빈집": "🔴", "B 반빈집": "🟠"}


# ── 유틸 ─────────────────────────────────────────────────────────

def ema_series(values, k):
    r = [values[0]]
    for v in values[1:]:
        r.append(v * k + r[-1] * (1 - k))
    return r


def grade(v, p10, p25, p75):
    if v <= p10: return "A 완전빈집"
    if v <= p25: return "B 반빈집"
    if v <= p75: return "C 정상"
    return "D 과매수"



# ── 1단계: 유동성 컨셉 파싱 ──────────────────────────────────────

def find_liquidity_file() -> Path:
    for f in XLSM_DIR.iterdir():
        if "유동성" in f.name and f.suffix in (".xlsm", ".xlsx") and not f.name.startswith("~$"):
            return f
    raise FileNotFoundError("유동성 파일 없음")


def parse_sector_text(text: str) -> list[str]:
    """'반도체, 전력(태양광), 피지컬 ai' → ['반도체', '전력', '피지컬 ai']"""
    cleaned = re.sub(r"\([^)]*\)", "", text)
    terms = [t.strip().lower() for t in cleaned.split(",")]
    return [t for t in terms if t]


def parse_one_sheet(wb, sheet_name: str) -> dict[str, list[str]]:
    """시트 하나에서 사모/투신/연금 시총대비 섹터 추출."""
    ws = wb[sheet_name]
    result: dict[str, list[str]] = {"사모": [], "투신": [], "연금": []}
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if not cell or not isinstance(cell, str):
                continue
            for 기관 in ["사모", "투신", "연금"]:
                if f"{기관} 시총대비" in cell and ":" in cell:
                    body = cell.split(":", 1)[1]
                    result[기관] = parse_sector_text(body)
    return result


def get_consecutive_sectors(wb, min_days: int) -> tuple[list[str], dict]:
    """
    날짜별 시트를 최신 순으로 읽어 사모+투신+연금 공통 섹터가
    min_days 연속 등장하는 섹터 반환.
    returns: (공통섹터_raw_list, {섹터: 연속일수})
    """
    import re as _re
    date_sheets = []
    for name in wb.sheetnames:
        m = _re.search(r"(\d{8})", name)
        if m and "유동성" in name:
            date_sheets.append((m.group(1), name))
    date_sheets.sort(reverse=True)  # 최신 먼저

    # 날짜별 교집합 섹터 수집
    daily_common: list[set[str]] = []
    for _, sname in date_sheets:
        parsed = parse_one_sheet(wb, sname)
        sets = [set(parsed[k]) for k in ["사모", "투신", "연금"] if parsed[k]]
        if len(sets) >= 2:
            common = sets[0].intersection(*sets[1:])
        elif sets:
            common = sets[0]
        else:
            common = set()
        daily_common.append(common)

    if not daily_common:
        return [], {}

    # 섹터별 연속 등장일수 계산
    all_sectors = set().union(*daily_common)
    streak: dict[str, int] = {}
    for sec in all_sectors:
        count = 0
        for day_set in daily_common:
            if sec in day_set:
                count += 1
            else:
                break
        streak[sec] = count

    # min_days 이상 연속 등장한 섹터
    valid = [s for s, cnt in streak.items() if cnt >= min_days]
    return valid, streak


def sectors_to_watchlist_keys(raw_sectors: list[str]) -> list[str]:
    """유동성 텍스트 섹터명 → watchlist.json 키 목록"""
    keys = []
    for sec in raw_sectors:
        mapped = LIQUIDITY_TO_SECTOR.get(sec, [])
        for k in mapped:
            if k not in keys:
                keys.append(k)
    return keys


# ── 2+3단계: 오실레이터 계산 ────────────────────────────────────

def find_osc_file() -> Path:
    # 700 종목 파일 (개별종목) 우선
    candidates = [f for f in XLSM_DIR.iterdir()
                  if "수급오실레이터" in f.name and "(700)" in f.name
                  and not f.name.startswith("~$")]
    if candidates:
        return max(candidates, key=lambda p: p.stat().st_mtime)
    # fallback: 아무 오실레이터 파일
    candidates = [f for f in XLSM_DIR.iterdir()
                  if "수급오실레이터" in f.name and "업종" not in f.name
                  and not f.name.startswith("~$")]
    if candidates:
        return max(candidates, key=lambda p: p.stat().st_mtime)
    raise FileNotFoundError("수급오실레이터 파일 없음")


def scan_stocks(sector_keys: list[str]) -> list[dict]:
    """
    sector_keys 내 종목을 오실레이터 파일에서 계산.
    returns: [{name, code, sector, consec_days, grade, osc}] — A/B + consec >= min_consec만
    """
    import win32com.client as win32

    watchlist = json.loads(WLIST_JSON.read_text(encoding="utf-8"))

    # 대상 종목 수집
    target_stocks: list[dict] = []  # {name, code, sector}
    seen = set()
    for sk in sector_keys:
        if sk not in watchlist:
            continue
        for sub_stocks in watchlist[sk].values():
            for st in sub_stocks:
                if st["name"] not in seen:
                    target_stocks.append({**st, "sector": sk})
                    seen.add(st["name"])

    if not target_stocks:
        return []

    xlsm_path = find_osc_file()
    print(f"  오실레이터 파일: {xlsm_path.name}")

    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.AutomationSecurity = 3
    wb_osc = xl.Workbooks.Open(str(xlsm_path))

    try:
        ws_osc  = wb_osc.Sheets("수급오실레이터")
        ws_size = wb_osc.Sheets("시가총액")
        ws_forn = wb_osc.Sheets("외국인매수데이터")
        ws_inst = wb_osc.Sheets("기관매수데이터")

        p90 = ws_osc.Cells(STAT_ROW_P90, STAT_COL).Value2
        p75 = ws_osc.Cells(STAT_ROW_P75, STAT_COL).Value2
        p25 = ws_osc.Cells(STAT_ROW_P25, STAT_COL).Value2
        p10 = ws_osc.Cells(STAT_ROW_P10, STAT_COL).Value2

        # 종목명 → 컬럼 매핑
        max_col = ws_forn.UsedRange.Columns.Count
        xl_stocks: dict[str, int] = {}
        for c in range(2, max_col + 1):
            name = ws_forn.Cells(9, c).Value2
            if name:
                xl_stocks[name] = c

        # 데이터 일괄 로딩
        def to_matrix(ws):
            return ws.Range(ws.Cells(DATA_ROW_START, 1),
                            ws.Cells(DATA_ROW_END, max_col)).Value

        mat_size = to_matrix(ws_size)
        mat_forn = to_matrix(ws_forn)
        mat_inst = to_matrix(ws_inst)

    finally:
        wb_osc.Close(False)
        xl.Quit()

    results = []
    for st in target_stocks:
        name = st["name"]
        if name not in xl_stocks:
            continue
        c = xl_stocks[name] - 1
        size_v = [float(mat_size[r][c]) if mat_size[r][c] else None for r in range(len(mat_size))]
        forn_v = [float(mat_forn[r][c]) if mat_forn[r][c] else None for r in range(len(mat_forn))]
        inst_v = [float(mat_inst[r][c]) if mat_inst[r][c] else None for r in range(len(mat_inst))]

        pairs = [((iv or 0) + (fv or 0), sv)
                 for iv, fv, sv in zip(inst_v, forn_v, size_v)
                 if sv and sv > 0]
        if len(pairs) < 27:
            continue

        signal = [net / sz for net, sz in pairs]

        ema12 = ema_series(signal, K12)
        ema26 = ema_series(signal, K26)
        macd  = [e12 - e26 for e12, e26 in zip(ema12, ema26)]
        sig_l = ema_series(macd, K9)
        osc   = macd[-1] - sig_l[-1]

        g = grade(osc, p10, p25, p75)

        if g in ("A 완전빈집", "B 반빈집"):
            results.append({
                "name":   name,
                "code":   st["code"],
                "sector": st["sector"],
                "grade":  g,
                "osc":    osc,
            })

    return sorted(results, key=lambda x: x["osc"])


# ── 출력/텔레그램 ─────────────────────────────────────────────────

def build_message(streak: dict, results: list[dict], min_days: int) -> str:
    today = datetime.now().strftime("%Y-%m-%d")
    lines = [f"[유동성x빈집 픽] {today}", ""]

    # 연속 섹터 현황
    lines.append(f"연속 {min_days}일+ 교집합 섹터:")
    for sec, cnt in sorted(streak.items(), key=lambda x: -x[1]):
        if cnt >= min_days:
            lines.append(f"  {sec} {cnt}일 연속")
    lines.append("")

    if not results:
        lines.append("조건 충족 종목 없음 (연속매수 + 빈집)")
        return "\n".join(lines)

    lines.append(f"빈집 종목 ({len(results)}개):")
    cur_sector = None
    for r in results:
        if r["sector"] != cur_sector:
            cur_sector = r["sector"]
            lines.append(f"\n[{cur_sector}]")
        icon = GRADE_ICON[r["grade"]]
        lines.append(f"  {icon} {r['name']}  {r['grade']}")

    return "\n".join(lines)


def send_telegram(text: str):
    import urllib.request, urllib.parse
    env = {}
    ep = ROOT / ".env"
    if ep.exists():
        for line in ep.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    token = env.get("BOT_TOKEN", "")
    chat_id = env.get("CHAT_ID", "")
    if not token or not chat_id:
        print("텔레그램 설정 없음")
        return
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        url  = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode({"chat_id": chat_id, "text": chunk}).encode()
        with urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=10) as r:
            res = json.loads(r.read())
            print("텔레그램 전송 완료" if res.get("ok") else f"실패: {res}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--days", type=int, default=2, help="유동성 연속 등장 기준일 (기본 2)")
    ap.add_argument("--tg", action="store_true")
    args = ap.parse_args()

    import openpyxl
    liq_file = find_liquidity_file()
    print(f"유동성 파일: {liq_file.name}")
    wb_liq = openpyxl.load_workbook(str(liq_file), read_only=True, keep_vba=True, data_only=True)

    print(f"유동성 컨셉 파싱 중 (연속 {args.days}일 기준)...")
    raw_sectors, streak = get_consecutive_sectors(wb_liq, args.days)
    wb_liq.close()

    if not raw_sectors:
        print("연속 교집합 섹터 없음")
        return

    sector_keys = sectors_to_watchlist_keys(raw_sectors)
    print(f"교집합 섹터: {raw_sectors}")
    print(f"watchlist 키: {sector_keys}")

    if not sector_keys:
        print("매핑된 watchlist 섹터 없음")
        return

    print(f"\n오실레이터 계산 중 (빈집 A/B 필터)...")
    results = scan_stocks(sector_keys)

    msg = build_message(streak, results, args.days)
    print("\n" + msg)

    if args.tg:
        send_telegram(msg)


if __name__ == "__main__":
    main()

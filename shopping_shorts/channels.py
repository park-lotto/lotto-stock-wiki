"""벤치마킹 엑셀 → 채널 목록 로더 + 활동성 선별(2026-07-22)."""
import re
from openpyxl import load_workbook
from shopping_shorts.config import EXCEL_PATH, MAX_CHANNELS, MAX_TRACKED

_IG_RE = re.compile(r"instagram\.com/([A-Za-z0-9_.]+)")
_RESERVED = {"p", "reel", "reels", "stories", "explore", "tv", "s", "accounts"}


def username_from_url(url):
    """인스타 URL에서 username 추출. 인스타 URL 아니면 None. 예약 경로는 거부."""
    if not url:
        return None
    m = _IG_RE.search(str(url).strip())
    if not m:
        return None
    name = m.group(1).strip("/")
    if name.lower() in _RESERVED:
        return None
    return name


def parse_rows(rows):
    """엑셀 rows(values_only) → 채널 dict 리스트. 헤더 1행 스킵, 무효 URL 제외."""
    channels = []
    for row in rows[1:]:
        name = row[0]
        url = row[1] if len(row) > 1 else None
        followers = row[2] if len(row) > 2 else None
        inpock = row[3] if len(row) > 3 else None
        username = username_from_url(url)
        if not username:
            continue
        channels.append({
            "name": name,
            "username": username,
            "followers": int(followers) if followers else 0,
            "inpock": inpock or "",
        })
    return channels


def cap_channels(channels, max_channels=MAX_CHANNELS):
    """채널 수가 max_channels 초과면 팔로워 낮은 순으로 잘라낸다(2026-07-13,
    채널당 수집비 고정이라 리스트가 커질수록 매일 비용이 그대로 늘어남 —
    비용 관리용 상한). max_channels 이하면 순서 그대로 반환."""
    if len(channels) <= max_channels:
        return channels
    return sorted(channels, key=lambda c: c.get("followers", 0), reverse=True)[:max_channels]


def _norm(c):
    return c["username"].strip().lstrip("@").lower()


def merge_tracked(excel_channels, discovered, removed=None, max_channels=MAX_CHANNELS):
    """엑셀 채널 + 발굴/등록 채널을 union하고 제외(removed) 채널을 뺀 최종 추적 목록.

    손으로 등록/발굴한 채널(discovered)을 union 앞쪽에 둔다(2026-07-18). 엑셀은
    이미 MAX_CHANNELS(팔로워 상위)로 꽉 차 있어서, 발굴채널을 뒤에 붙이고 cap을
    자르면 등록해도 전부 잘려나가(수집 안 됨) 버그였다. 앞에 두면 cap에서 엑셀의
    팔로워 최저 채널부터 밀려나고 사장님이 고른 채널은 살아남는다 — 수집 총량
    (=수집비 상한)은 그대로. 엑셀에 이미 있는 발굴채널은 중복 제외한다.

    discovered: [{username, ...}] 최신 추가순. removed: 소문자 정규화된 username set."""
    known = {_norm(c) for c in excel_channels}
    fresh = [d for d in discovered if _norm(d) not in known]
    merged = fresh + list(excel_channels)
    if removed:
        merged = [c for c in merged if _norm(c) not in removed]
    return merged[:max_channels]


def select_tracked(excel_channels, discovered, removed=None, dead=None,
                   max_channels=MAX_TRACKED):
    """활동성 기반 최종 추적 목록(2026-07-22, 팔로워 컷 대체).

    포함: 생존(alive) + 측정대기(pending=활동기록 없음). 제외: 사망(dead) + 수동제거(removed).
    팔로워로 자르지 않는다 — 2일 게이트가 세트를 좁히므로 max_channels(안전상한)는 보통
    안 걸린다. 초과 시엔 앞쪽(discovered 먼저)부터 남긴다.

    dead: 센서스가 사망 판정한 소문자 username set. removed: 수동제거 소문자 set.
    측정대기 채널(dead·alive 어디에도 없음)은 포함 — 신규/등록 채널이 여기라 다음 수집에서
    바로 긁혀 측정된다(초기 안전동작: 센서스 전엔 전원 pending이라 아무도 안 잘림)."""
    known = {_norm(c) for c in excel_channels}
    fresh = [d for d in discovered if _norm(d) not in known]
    merged = fresh + list(excel_channels)
    removed = removed or set()
    dead = dead or set()
    merged = [c for c in merged if _norm(c) not in removed and _norm(c) not in dead]
    return merged[:max_channels]


def load_channels(excel_path=EXCEL_PATH):
    """엑셀 파일 열어 채널 리스트 전체 반환. 팔로워 컷 없음(2026-07-22 — 선별은
    select_tracked가 활동성으로 담당). cap_channels는 기존 테스트 보존용으로만 남김."""
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return parse_rows(rows)

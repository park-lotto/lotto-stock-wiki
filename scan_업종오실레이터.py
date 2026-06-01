#!/usr/bin/env python3
"""
scan_업종오실레이터.py — 업종별 수급 빈집 탐지 + 텔레그램 전송

사용법:
  python scan_업종오실레이터.py          ← 콘솔 출력
  python scan_업종오실레이터.py --tg     ← 텔레그램 전송

업종 파일: raw/매일 엑셀넣을것/외국인기관수급오실레이터(업종).xlsm
  오실 시트: 9행=업종명, 15~91행=날짜별 오실레이터값
  컬럼 구조: 업종당 2열 (짝수=원시금액, 홀수=정규화 오실레이터)
  → 홀수 열(3,5,7…)만 사용
"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from pathlib import Path
from datetime import date

BASE     = Path(__file__).parent
XLSM_DIR = BASE / 'raw' / '매일 엑셀넣을것'
WLIST    = BASE / 'watchlist.json'

# 관심 섹터 키워드 (강조 표시용)
WATCH_KEYS = {
    '반도체', '조선', '방산', '바이오', '우주', '로봇', '자동차',
    '이차전지', '2차전지', '전력', '원전', '신재생', '화장품',
    'LNG', '철강', '엔터', 'AI', '건설', '은행', '보험',
}

GRADE_ICON = {'A': '🔴', 'B': '🟠', 'C': '⚪', 'D': '🟡'}


def find_file() -> Path:
    hits = sorted(XLSM_DIR.glob('*오실레이터*(업종)*.xlsm'), key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        hits = sorted(XLSM_DIR.glob('*업종*오실레이터*.xlsm'), key=lambda p: p.stat().st_mtime, reverse=True)
    if not hits:
        raise FileNotFoundError('업종 오실레이터 파일 없음 — raw/매일 엑셀넣을것/ 확인')
    return hits[0]


def load_data(path: Path):
    """오실 시트에서 업종명 + 시계열 + 최신 오실레이터 읽기"""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True, keep_vba=False)
    ws = wb['오실']

    # 9행: 업종명 (짝이 되는 첫 번째 컬럼만, 즉 홀수 위치 컬럼)
    # 구조: col2=KOSPI(원시), col3=KOSPI(정규화), col4=KOSDAQ(원시), col5=KOSDAQ(정규화)…
    # 홀수 열 = 정규화 오실레이터 (col 3,5,7…)
    업종_cols = {}  # name → col
    c = 2
    while c <= ws.max_column:
        name = ws.cell(9, c).value
        if name and isinstance(name, str):
            osc_col = c + 1  # 바로 다음 홀수 열
            if osc_col <= ws.max_column:
                if name not in 업종_cols:  # 중복 방지 (이미 등록된 이름은 건너뜀)
                    업종_cols[name] = osc_col
        c += 2

    # 15~91행: 날짜별 데이터
    results = {}
    for name, osc_col in 업종_cols.items():
        series = []
        for r in range(15, 92):
            v = ws.cell(r, osc_col).value
            try:
                series.append(float(v))
            except (TypeError, ValueError):
                series.append(0.0)
        results[name] = series

    # 날짜 (1열 15~91행)
    dates = []
    for r in range(15, 92):
        v = ws.cell(15 if r == 15 else r, 1).value
        dates.append(str(v)[:10] if v else '')

    last_date = str(ws.cell(91, 1).value)[:10] if ws.cell(91, 1).value else str(date.today())

    wb.close()
    return results, last_date


def grade(v, p10, p25, p75):
    if v <= p10: return 'A'
    if v <= p25: return 'B'
    if v >= p75: return 'D'
    return 'C'


def trend(series):
    if len(series) < 5: return '—'
    avg = sum(series[-5:-1]) / 4
    latest = series[-1]
    if avg == 0: return '→ 횡보'
    if latest > avg * 1.05: return '↑ 회복'
    if latest < avg * 0.95: return '↓ 심화'
    return '→ 횡보'


def send_telegram(text: str):
    import urllib.request, urllib.parse
    env = {}
    env_path = BASE / '.env'
    if env_path.exists():
        for line in env_path.read_text(encoding='utf-8').splitlines():
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip()
    token   = env.get('BOT_TOKEN', '')
    chat_id = env.get('CHAT_ID', '')
    if not token or not chat_id:
        print('텔레그램 설정 없음 (.env BOT_TOKEN/CHAT_ID)')
        return
    url  = f'https://api.telegram.org/bot{token}/sendMessage'
    data = urllib.parse.urlencode({'chat_id': chat_id, 'text': text, 'parse_mode': 'HTML'}).encode()
    with urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=10) as r:
        res = json.loads(r.read())
    print('✅ 텔레그램 전송 완료' if res.get('ok') else f'❌ 전송 실패: {res}')


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--tg', action='store_true')
    args = parser.parse_args()

    path = find_file()
    print(f'파일: {path.name}\n')

    data, last_date = load_data(path)

    # 최신값 추출 및 백분위 계산
    latest = {name: s[-1] for name, s in data.items() if s}

    # KOSPI/KOSDAQ 제외하고 업종만 (지수 제외)
    EXCLUDE = {'KOSPI', 'KOSDAQ'}
    업종만 = {n: v for n, v in latest.items() if n not in EXCLUDE}

    all_vals = sorted(업종만.values())
    n = len(all_vals)
    p10 = all_vals[int(n * 0.10)]
    p25 = all_vals[int(n * 0.25)]
    p75 = all_vals[int(n * 0.75)]

    # 등급별 분류
    A = sorted([(v, nm) for nm, v in 업종만.items() if grade(v, p10, p25, p75) == 'A'])
    B = sorted([(v, nm) for nm, v in 업종만.items() if grade(v, p10, p25, p75) == 'B'])
    D = sorted([(v, nm) for nm, v in 업종만.items() if grade(v, p10, p25, p75) == 'D'], reverse=True)

    today = str(date.today())

    # ── 콘솔 출력 ───────────────────────────────────────────────
    print(f'[ 업종 수급오실레이터 ] {last_date}')
    print(f'총 {n}개 업종 | 🔴A:{len(A)} 🟠B:{len(B)} 🟡D:{len(D)}\n')

    print('🔴 빈집A (하위10%):')
    for v, nm in A:
        tr = trend(data[nm])
        watch = '⭐' if any(k in nm for k in WATCH_KEYS) else ''
        print(f'  {watch}{nm}  {v:+.6f}  {tr}')

    print('\n🟠 빈집B (하위25%):')
    for v, nm in B:
        tr = trend(data[nm])
        watch = '⭐' if any(k in nm for k in WATCH_KEYS) else ''
        print(f'  {watch}{nm}  {v:+.6f}  {tr}')

    print(f'\n🟡 과매수D (상위25%):')
    for v, nm in D[:5]:
        tr = trend(data[nm])
        print(f'  {nm}  {v:+.6f}  {tr}')

    # ── 텔레그램 메시지 ─────────────────────────────────────────
    if args.tg:
        lines = [f'<b>[ 업종 수급오실레이터 ] {last_date}</b>']
        lines.append(f'총 {n}개 업종 | 🔴A:{len(A)} 🟠B:{len(B)} 🟡D:{len(D)}\n')

        lines.append('<b>🔴 빈집A (하위10%)</b>')
        for v, nm in A:
            tr = trend(data[nm])
            watch = '⭐' if any(k in nm for k in WATCH_KEYS) else ''
            lines.append(f'{GRADE_ICON["A"]} {watch}{nm}  <code>{v:+.6f}</code>  {tr}')

        lines.append('\n<b>🟠 빈집B (하위25%)</b>')
        for v, nm in B:
            tr = trend(data[nm])
            watch = '⭐' if any(k in nm for k in WATCH_KEYS) else ''
            lines.append(f'{GRADE_ICON["B"]} {watch}{nm}  <code>{v:+.6f}</code>  {tr}')

        lines.append(f'\n<b>🟡 과매수 Top5</b>')
        for v, nm in D[:5]:
            tr = trend(data[nm])
            lines.append(f'{GRADE_ICON["D"]} {nm}  <code>{v:+.6f}</code>  {tr}')

        msg = '\n'.join(lines)
        if len(msg) > 3900:
            msg = msg[:3890] + '\n...(이하 생략)'
        send_telegram(msg)


if __name__ == '__main__':
    main()

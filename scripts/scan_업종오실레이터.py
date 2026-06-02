"""
scan_업종오실레이터.py — 업종별 외인+기관 수급 오실레이터 분석

원리: 외인+기관 순매수 / 시가총액 의 EMA MACD
  양수(+) = 수급 유입 강화 중 → 주도섹터 후보
  음수(-) = 수급 이탈 중     → 약한 섹터

사용법:
    python scripts/scan_업종오실레이터.py
    python scripts/scan_업종오실레이터.py --tg
    python scripts/scan_업종오실레이터.py --top 15
"""

import sys, argparse, json, urllib.request, urllib.parse
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

ROOT      = Path(__file__).parent.parent
EXCEL_DIR = ROOT / "raw" / "매일 엑셀넣을것"
TODAY     = datetime.today().strftime("%Y-%m-%d")
TODAY_KR  = datetime.today().strftime("%Y년 %m월 %d일")

# 주요 섹터만 필터 (KRX/WISE 지수 제외, 업종명만)
SECTOR_FILTER = {
    "반도체", "반도체와반도체장비", "IT하드웨어", "IT가전", "디스플레이",
    "자동차", "자동차부품", "조선", "기계", "전기장비",
    "화학", "에너지", "철강", "비철,목재등",
    "소프트웨어", "미디어", "미디어,교육",
    "건강관리", "제약", "생명과학",
    "은행", "증권", "보험", "생명보험",
    "건설,건축관련", "운송", "항공사", "해운사",
    "호텔,레저서비스", "소매(유통)", "식품,음료", "필수소비재",
    "화장품", "복합기업", "상사,자본재", "가스유틸리티", "전기유틸리티",
    "우주항공과국방",
}


def find_excel(pattern: str) -> Path | None:
    for f in EXCEL_DIR.iterdir():
        if f.name.startswith("~$"): continue
        if pattern in f.name and f.suffix in (".xlsx", ".xlsm"):
            return f
    return None


def load_sector_osc() -> list[dict]:
    path = find_excel("수급오실레이터(업종)")
    if not path:
        print("❌ 수급오실레이터(업종) 파일 없음"); sys.exit(1)

    print(f"📂 {path.name} 로딩...")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["오실"]

    # 헤더: row 9 = Name, row 12 = 시가총액/오실레이터 구분
    header_row = list(ws.iter_rows(min_row=9, max_row=9, max_col=200, values_only=True))[0]

    # 오실레이터 컬럼 위치 (짝수: 2,4,6...) — 홀수는 시가총액
    osc_cols = {}  # col_idx → sector_name
    for j in range(2, len(header_row), 2):
        name = header_row[j]
        if name and isinstance(name, str):
            osc_cols[j] = name

    # 데이터 읽기 (row 13~)
    all_data = list(ws.iter_rows(min_row=13, max_row=5000, max_col=max(osc_cols.keys())+2, values_only=True))
    valid = [r for r in all_data if r[0] and isinstance(r[0], datetime)]
    wb.close()

    if not valid:
        print("❌ 데이터 없음"); sys.exit(1)

    print(f"   {len(valid)}일 데이터 / {len(osc_cols)}개 섹터")

    # 최신값 + 방향 (5일 평균 대비)
    last = valid[-1]
    prev5_avg = {}

    for r in valid[-6:-1]:
        for j, name in osc_cols.items():
            if j < len(r) and isinstance(r[j], (int, float)):
                prev5_avg.setdefault(name, []).append(r[j])

    seen = {}  # 중복 섹터: 절대값 가장 큰 것만 유지
    for j, name in osc_cols.items():
        if j >= len(last): continue
        val = last[j]
        if not isinstance(val, (int, float)): continue

        p5 = prev5_avg.get(name, [])
        avg5 = sum(p5) / len(p5) if p5 else val
        trend = "↑" if val > avg5 * 1.05 else "↓" if val < avg5 * 0.95 else "→"

        entry = {
            "name":  name,
            "osc":   round(float(val), 6),
            "trend": trend,
            "pure":  name in SECTOR_FILTER,
        }
        if name not in seen or abs(entry["osc"]) > abs(seen[name]["osc"]):
            seen[name] = entry

    return sorted(seen.values(), key=lambda x: x["osc"], reverse=True)


def send_tg(text: str):
    env = {}
    ep = ROOT / ".env"
    if ep.exists():
        for line in ep.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    token, chat_id = env.get("BOT_TOKEN",""), env.get("CHAT_ID","")
    if not token or not chat_id:
        print("  ⚠️  .env BOT_TOKEN/CHAT_ID 없음"); return
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        url  = f"https://api.telegram.org/bot{token}/sendMessage"
        data = urllib.parse.urlencode({
            "chat_id": chat_id, "text": chunk,
            "parse_mode": "HTML", "disable_web_page_preview": "true"
        }).encode()
        try:
            with urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=10) as r:
                res = json.loads(r.read())
                if res.get("ok"):
                    print("  ✅ 텔레그램 발송")
                else:
                    print(f"  ❌ {res}")
        except Exception as e:
            print(f"  ❌ 발송 실패: {e}")


def build_message(results: list, top_n: int) -> str:
    # 주요 업종만 필터
    pure = [r for r in results if r["pure"]]
    top  = pure[:top_n]
    bot  = list(reversed(pure))[:5]

    lines = [
        f"📊 <b>업종 수급 오실레이터</b> — {TODAY_KR}",
        f"<i>외인+기관 순매수/시가총액 MACD | 양수=수급유입 강화</i>",
        "",
        f"<b>🔴 수급 유입 중 업종 Top {top_n}</b>",
    ]
    for r in top:
        arrow = "🔺" if r["trend"] == "↑" else ("🔻" if r["trend"] == "↓" else "➡️")
        lines.append(f"  {arrow} {r['name']}  <code>{r['osc']:+.5f}</code>")

    lines += [
        "",
        f"<b>🔵 수급 약한 업종 Bottom 5</b>",
    ]
    for r in bot:
        arrow = "🔺" if r["trend"] == "↑" else ("🔻" if r["trend"] == "↓" else "➡️")
        lines.append(f"  {arrow} {r['name']}  <code>{r['osc']:+.5f}</code>")

    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tg",  action="store_true")
    ap.add_argument("--top", type=int, default=10)
    ap.add_argument("--all", action="store_true", help="KRX/WISE 지수 포함 전체 출력")
    args = ap.parse_args()

    results = load_sector_osc()

    # 출력
    display = results if args.all else [r for r in results if r["pure"]]

    print(f"\n{'='*55}")
    print(f"📊 업종 수급 오실레이터 — {TODAY}")
    print(f"{'='*55}")
    print(f"\n🔴 수급 유입 중 업종 Top {args.top}:")
    for r in display[:args.top]:
        arrow = "▲" if r["trend"]=="↑" else ("▼" if r["trend"]=="↓" else "→")
        print(f"  {arrow} {r['name']:25s}  {r['osc']:+.6f}")

    print(f"\n🔵 수급 약한 업종 Bottom 5:")
    for r in list(reversed(display))[:5]:
        arrow = "▲" if r["trend"]=="↑" else ("▼" if r["trend"]=="↓" else "→")
        print(f"  {arrow} {r['name']:25s}  {r['osc']:+.6f}")

    if args.tg:
        msg = build_message(results, args.top)
        print("\n📲 텔레그램 발송 중...")
        send_tg(msg)


if __name__ == "__main__":
    main()

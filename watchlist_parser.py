"""
watchlist_parser.py
raw/내 관심종목.xlsx → watchlist.json 변환 스크립트

실행: python watchlist_parser.py
출력: watchlist.json  {섹터: {서브섹터: [{name, code}]}}

종목 추가/수정: 엑셀만 편집 후 이 스크립트 다시 실행.
"""

import json
import re
from pathlib import Path

from openpyxl import load_workbook

WATCHLIST_XLSX = Path(r"C:\Users\TheRose\Desktop\로또의 주식\raw\내 관심종목.xlsx")
OUTPUT_JSON    = Path(r"C:\Users\TheRose\Desktop\로또의 주식\watchlist.json")

# 16개 일일 섹터. Excel A열 헤더 텍스트가 아래 키(정확히 일치 or 포함)이면 섹터 전환.
# "포함" 방식이므로 순서 중요 — 긴 것을 앞에.
SECTOR_TRIGGERS: dict[str, str] = {
    # Excel 헤더 텍스트            → 섹터명
    "바이오 섹터별-2":              "바이오",
    "바이오 섹터별":                "바이오",
    "바이오 비만당뇨치매":           "바이오",
    "반도체 테마별":                "반도체",
    "반도체 설계 및 파운드리":        "반도체",
    "반도체 전공정":                "반도체",
    "반도체 후공정":                "반도체",
    "반도체 디스플레이":             "반도체",
    "반도체 데이터센터":             "반도체",
    "에너지 전선/변압기":            "전력",
    "에너지 원전":                  "원전",
    "에너지 신재생":                "신재생",
    "뷰티 화장품":                  "화장품",
    "뷰티 미용":                    "미용",
    "이차전지-1":                   "이차전지",
    "이차전지-2":                   "이차전지",
    "로봇-1":                      "로봇",
    "로봇-2":                      "로봇",
    "조선":                        "조선",
    "LNG":                         "LNG",
    "방산":                        "방산",
    "우주":                        "우주",
    "자동차":                      "자동차",
    "철강":                        "철강",
    "엔터":                        "엔터",
}

# 이 텍스트는 섹터/서브섹터 구분 없이 무시할 행
# 헤더 행이지만 섹터/서브섹터 변경 없이 무시
SKIP_PATTERNS = {"0 0 0 0", "ETF 시장전체", "ETF시장 체크", "우주방산", "원자재"}

# 이 헤더를 만나면 current_sector=None (16개 외 섹션 진입 신호)
SECTOR_RESET_TRIGGERS = {
    "AI 소프트웨어", "AI 의료", "AI 보안/양자", "메타버스/6G",
    "시멘트/페인트", "가구 인테리어", "코인/STO",
    "대왕고래", "트럼프<재건/경협>",
    "식품", "여행/항공", "웹툰/게임",
    "미디어 콘텐츠", "소비재",
    "블랙테마 원자재", "블랙테마 전쟁",
    "코로나 전염병", "육계/수산", "정책",
}

# 코드 패턴 (6자리 숫자)
CODE_RE = re.compile(r"^\d{6}$")

DAILY_SECTORS = {
    "반도체", "방산", "조선", "LNG", "바이오", "우주", "로봇",
    "자동차", "이차전지", "전력", "원전", "신재생", "화장품", "미용", "철강", "엔터",
}


def is_code(val) -> bool:
    if val is None:
        return False
    return bool(CODE_RE.match(str(val).strip()))


def find_sector_trigger(text: str) -> str | None:
    """A열 헤더 → 16개 섹터 매핑 (정확한 텍스트 일치만). 없으면 None."""
    return SECTOR_TRIGGERS.get(text.strip())


def parse() -> dict:
    wb = load_workbook(WATCHLIST_XLSX, read_only=True, data_only=True)
    ws = wb.active

    result: dict[str, dict[str, list]] = {}
    current_sector: str | None = None
    current_sub: str | None = None

    for row in ws.iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] is not None else ""
        b = str(row[1]).strip() if row[1] is not None else ""
        e = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""

        if is_code(e):
            # 주식 행
            name = b
            if not name or not current_sector:
                continue
            sub = current_sub or "기타"
            result.setdefault(current_sector, {}).setdefault(sub, [])
            result[current_sector][sub].append({"name": name, "code": e})

        else:
            # 헤더/구분 행
            if not a or a in SKIP_PATTERNS:
                continue

            # 16개 외 섹션 진입 → 섹터 초기화 (이후 stocks 제외)
            if a in SECTOR_RESET_TRIGGERS:
                current_sector = None
                current_sub = None
                continue

            sector = find_sector_trigger(a)
            if sector:
                current_sector = sector
                current_sub = None
            elif current_sector:
                # 서브섹터명
                current_sub = a

    wb.close()
    return result


def main():
    print(f"파싱 중: {WATCHLIST_XLSX.name}")
    data = parse()

    total_stocks = 0
    for sector in sorted(data.keys()):
        subs = data[sector]
        n = sum(len(v) for v in subs.values())
        total_stocks += n
        sub_names = list(subs.keys())
        print(f"  {sector:8s}: {len(subs):2d}개 서브섹터  {n:3d}종목  ({', '.join(sub_names[:4])}{'...' if len(sub_names)>4 else ''})")

    missing = DAILY_SECTORS - set(data.keys())
    if missing:
        print(f"\n⚠️  미감지 섹터: {missing}")

    extra = set(data.keys()) - DAILY_SECTORS
    if extra:
        print(f"ℹ️  추가 감지(미포함): {extra}")

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\nwatchlist.json 저장 ({total_stocks}종목, {len(data)}섹터) -> {OUTPUT_JSON}")


if __name__ == "__main__":
    main()

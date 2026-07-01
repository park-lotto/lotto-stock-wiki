"""
ingest_excel.py — 매일 엑셀넣을것 → wiki 자동 반영

실행:
    python scripts/ingest_excel.py              # 전체 처리
    python scripts/ingest_excel.py --dry-run    # 실제 파일 수정 없이 리포트만 출력

출력:
    raw/ingest_report_{날짜}.md   — 처리 결과 요약
    wiki 파일 직접 수정           (--dry-run 아닐 때)
    wiki/log.md 자동 기록
"""

import sys, os, re, glob, argparse, subprocess, json
from datetime import datetime, date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("openpyxl 없음. 설치: python -m pip install openpyxl")
    sys.exit(1)

# ─── 경로 설정 ────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
EXCEL_DIR = ROOT / "raw" / "매일 엑셀넣을것"
WIKI_DIR  = ROOT / "wiki"
LOG_FILE  = WIKI_DIR / "log.md"
TODAY     = datetime.today().strftime("%Y-%m-%d")

# ─── 섹터/종목 매핑 ───────────────────────────────────────────

SECTOR_FOLDER_MAP: dict[str, Path] = {
    "반도체":         WIKI_DIR / "L5_섹터" / "반도체",
    "반도체부품":     WIKI_DIR / "L5_섹터" / "반도체",
    "기판":           WIKI_DIR / "L5_섹터" / "반도체",
    "2차전지":        WIKI_DIR / "L5_섹터" / "2차전지ESS",
    "2차전지ESS":     WIKI_DIR / "L5_섹터" / "2차전지ESS",
    "조선":           WIKI_DIR / "L5_섹터" / "조선",
    "전력기기":       WIKI_DIR / "L5_섹터" / "전력기기",
    "로봇":           WIKI_DIR / "L5_섹터" / "로봇",
    "바이오":         WIKI_DIR / "L5_섹터" / "바이오",
    "방산":           WIKI_DIR / "L5_섹터" / "방산",
    "자동차":         WIKI_DIR / "L5_섹터" / "자동차",
    "화장품":         WIKI_DIR / "L5_섹터" / "화장품",
    "소비내수":       WIKI_DIR / "L5_섹터" / "소비내수",
    "AI소프트웨어":   WIKI_DIR / "L5_섹터" / "AI소프트웨어",
    "신재생":         WIKI_DIR / "L5_섹터" / "신재생",
    "원전":           WIKI_DIR / "L5_섹터" / "원전",
    "LNG":            WIKI_DIR / "L5_섹터" / "LNG",
    "우주":           WIKI_DIR / "L5_섹터" / "우주",
    "철강":           WIKI_DIR / "L5_섹터" / "철강",
    "전력":           WIKI_DIR / "L5_섹터" / "전력",
    "전자전기":       WIKI_DIR / "L5_섹터" / "전자전기",
    "테마이벤트":     WIKI_DIR / "L5_섹터" / "테마이벤트",
    "미용":           WIKI_DIR / "L5_섹터" / "미용",
    "엔터":           WIKI_DIR / "L5_섹터" / "엔터",
}

# 투자아이디어정리.xlsx 시트명 → 섹터 키 (None=제외)
IDEA_SHEET_SECTOR_MAP: dict[str, str | None] = {
    "반도체소부장":          "반도체",
    "리노공업이오테크닉스":  "반도체",
    "리노공업·이오테크닉스": "반도체",
    "바이오":                "바이오",
    "중소바이오":            "바이오",
    "전력기기":              "전력기기",
    "SNT에너지":             "전력기기",
    "조선":                  "조선",
    "동성화인텍한국카본":    "조선",
    "동성화인텍·한국카본":   "조선",
    "로봇":                  "로봇",
    "로보티즈고영":          "로봇",
    "로보티즈·고영":         "로봇",
    "방산":                  "방산",
    "자동차":                "자동차",
    "화장품":                "화장품",
    "2차전지":               "2차전지",
    "해외주식":              None,
}

# ─── 헬퍼 ────────────────────────────────────────────────────

def load_wb(path: Path):
    """data_only=True로 캐시값 읽기 (수식 결과값)"""
    try:
        return openpyxl.load_workbook(str(path), read_only=True, keep_vba=False, data_only=True)
    except Exception as e:
        print(f"  ⚠️ {path.name} 열기 실패: {e}")
        return None

def find_excel(pattern: str) -> Path | None:
    """패턴 문자열로 EXCEL_DIR에서 파일 찾기"""
    for f in EXCEL_DIR.iterdir():
        if f.name.startswith("~$"): continue
        if pattern in f.name and f.suffix in (".xlsx", ".xlsm"):
            return f
    return None

_STOCK_PAGE_CACHE: dict[str, Path] | None = None

def _build_stock_cache() -> dict[str, Path]:
    """wiki/L5_섹터 1회 스캔 → 종목명(공백제거) : Path 캐시"""
    cache = {}
    for p in WIKI_DIR.rglob("stock_*.md"):
        name = p.stem.replace("stock_", "").replace(" ", "")
        cache[name] = p
    return cache

def find_stock_page(name: str) -> Path | None:
    """종목명으로 wiki stock 페이지 경로 찾기 (캐시 사용)"""
    global _STOCK_PAGE_CACHE
    if _STOCK_PAGE_CACHE is None:
        _STOCK_PAGE_CACHE = _build_stock_cache()
    name_clean = name.strip().replace(" ", "")
    if name_clean in _STOCK_PAGE_CACHE:
        return _STOCK_PAGE_CACHE[name_clean]
    for k, v in _STOCK_PAGE_CACHE.items():
        if name_clean in k or k in name_clean:
            return v
    return None

def get_sector_folder(sector_name: str) -> Path | None:
    """섹터 키워드 → wiki L5_섹터 폴더 Path"""
    return SECTOR_FOLDER_MAP.get(sector_name)

def get_stock_sector(stock_name: str) -> str | None:
    """종목 wiki 페이지 경로에서 섹터 폴더명 추출"""
    page = find_stock_page(stock_name)
    if not page:
        return None
    parts = list(page.parts)
    try:
        idx = parts.index("L5_섹터")
        return parts[idx + 1]
    except (ValueError, IndexError):
        return None

def rows_from_sheet(ws, max_row=500, max_col=20):
    """시트에서 값 있는 행만 추출"""
    result = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        if any(v is not None for v in row):
            result.append(row)
    return result

def fmt_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        s = str(int(v))
        if len(s) == 8:
            return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return str(v) if v else ""

def fmt_pct(v) -> str:
    if v is None: return "—"
    if isinstance(v, (int, float)):
        return f"{v:+.1f}%"
    return str(v)

def fmt_num(v, unit="") -> str:
    if v is None: return "—"
    if isinstance(v, (int, float)):
        if abs(v) >= 1e12: return f"{v/1e12:.1f}조{unit}"
        if abs(v) >= 1e8:  return f"{v/1e8:.0f}억{unit}"
        if abs(v) >= 1e4:  return f"{int(v):,}{unit}"
        return f"{v:.2f}{unit}"
    return str(v)

# ─── 종목별 태린이 지표 집계 (대시보드 종목 카드용) ──────────────────
_KRX_NAME2CODE = None


def _load_name2code() -> dict:
    """pipeline/atoms/krx_codes.json ({"codes": {종목명: 코드}}) → {종목명: '6자리코드'}."""
    global _KRX_NAME2CODE
    if _KRX_NAME2CODE is not None:
        return _KRX_NAME2CODE
    m = {}
    try:
        path = ROOT / "pipeline" / "atoms" / "krx_codes.json"
        data = json.loads(path.read_text(encoding="utf-8"))
        for name, code in (data.get("codes") or {}).items():
            c = _norm_code(code)
            if c:
                m[name.strip()] = c
    except Exception:
        pass
    _KRX_NAME2CODE = m
    return m


def _norm_code(raw) -> str | None:
    """'A247540' / 247540 / '247540' → '247540'. 6자리 숫자 아니면 None."""
    if raw is None:
        return None
    s = str(raw).strip().upper()
    if s.startswith("A"):
        s = s[1:]
    if s.isdigit():
        s = s.zfill(6)
    return s if (s.isdigit() and len(s) == 6) else None


def _resolve_code(item: dict, name2code: dict, unmatched: list) -> str | None:
    """항목에서 코드 확보: code 필드 우선, 없으면 종목명(name/related)→코드 매핑."""
    c = _norm_code(item.get("code"))
    if c:
        return c
    name = (item.get("name") or item.get("related") or "").strip()
    if name and name in name2code:
        return name2code[name]
    if name:
        unmatched.append(name)
    return None


def build_stock_index(results: dict, dest=None) -> dict:
    """파서 results → 종목코드별 태린이 지표 스냅샷. pipeline/taerini_stock.json 저장."""
    name2code = _load_name2code()
    stocks: dict = {}
    unmatched: list = []

    def slot(code: str, name: str) -> dict:
        s = stocks.setdefault(code, {"name": name or ""})
        if name and not s.get("name"):
            s["name"] = name
        return s

    # 1) 오실레이터 (수급 + 중소형주) → osc
    for key in ("수급", "중소형주수급"):
        r = results.get(key) or {}
        for grp in ("빈집_A", "빈집_B", "과매수_상승", "과매수_하락"):
            for it in (r.get(grp) or []):
                code = _resolve_code(it, name2code, unmatched)
                if not code:
                    continue
                slot(code, it.get("name", ""))["osc"] = {
                    "group": grp, "osc": it.get("osc"),
                    "pct": it.get("pct"), "trend": it.get("trend"),
                }

    # 2) RS → rs (top30 / bottom10 만 존재)
    r = results.get("RS") or {}
    for bucket, tag in (("top30", "상위"), ("bottom10", "하위")):
        for it in (r.get(bucket) or []):
            code = _resolve_code(it, name2code, unmatched)
            if not code:
                continue
            slot(code, it.get("name", ""))["rs"] = {
                "rs_avg": it.get("RS_avg"), "norm": it.get("norm_RS_avg"),
                "bucket": tag,
            }

    # 3) 추정이익변경 → tp (TP_Up / TP_Down)
    r = (results.get("추정이익변경") or {}).get("results") or {}
    for bucket, d in (("TP_Up", "상향"), ("TP_Down", "하향")):
        for it in (r.get(bucket) or []):
            code = _resolve_code(it, name2code, unmatched)
            if not code:
                continue
            tp_new, tp_old = it.get("tp_new"), it.get("tp_old")
            pct = None
            if isinstance(tp_new, (int, float)) and isinstance(tp_old, (int, float)) and tp_old:
                pct = round((tp_new - tp_old) / tp_old * 100, 1)
            slot(code, it.get("name", ""))["tp"] = {
                "target": tp_new, "prev": tp_old, "change_pct": pct, "dir": d,
                "brokerage": it.get("brokerage"), "date": it.get("date"),
            }

    # 4) 컨센움직임 → consensus (집계 기준일 last_update 포함)
    cm = results.get("컨센움직임") or {}
    cm_date = cm.get("last_update")
    r = cm.get("results") or {}
    for bucket in ("컨센상향", "컨센하향", "서프라이즈", "쇼크"):
        for it in (r.get(bucket) or []):
            code = _resolve_code(it, name2code, unmatched)
            if not code:
                continue
            slot(code, it.get("name", ""))["consensus"] = {
                "type": bucket, "csen_chg": it.get("csen_chg"),
                "surprise_rate": it.get("surprise_rate"), "date": cm_date,
            }

    # 5) 가속화모멘텀 → accel (그룹별, 종목명만; 점수 높은 그룹 우선)
    r = (results.get("가속화모멘텀") or {}).get("results") or {}
    for grp, items in r.items():
        for it in (items or []):
            code = _resolve_code(it, name2code, unmatched)
            if not code:
                continue
            cur = slot(code, it.get("name", ""))
            prev = cur.get("accel")
            if prev and (prev.get("score") or 0) >= (it.get("score") or 0):
                continue
            cur["accel"] = {"group": grp, "score": it.get("score")}

    # 6) 액티브ETF → etf
    r = results.get("액티브ETF") or {}
    for bucket, act in (("increase", "비중증가"), ("decrease", "비중감소")):
        for it in (r.get(bucket) or []):
            code = _resolve_code(it, name2code, unmatched)
            if not code:
                continue
            slot(code, it.get("name", ""))["etf"] = {
                "action": act, "diff": it.get("diff"), "rate": it.get("rate"),
            }

    # 7) 일정 → schedule (related = 종목명)
    r = results.get("일정") or {}
    today = date.today()
    for bucket in ("d7", "d30"):
        for it in (r.get(bucket) or []):
            code = _resolve_code({"name": it.get("related")}, name2code, unmatched)
            if not code:
                continue
            dday = None
            try:
                ev = datetime.strptime((it.get("date") or "").strip(), "%Y-%m-%d").date()
                dday = (ev - today).days
            except Exception:
                pass
            slot(code, it.get("related", ""))["schedule"] = {
                "event": it.get("content", ""), "date": it.get("date"), "dday": dday,
            }

    out = {
        "date": TODAY,
        "stocks": stocks,
        "meta": {
            "built_at": datetime.now().isoformat(timespec="seconds"),
            "stock_count": len(stocks),
            "unmatched": sorted(set(unmatched)),
        },
    }
    dest = Path(dest) if dest else (ROOT / "pipeline" / "taerini_stock.json")
    tmp = dest.with_suffix(dest.suffix + ".tmp")
    tmp.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp, dest)
    return out

# ─── 파서 1: 추정이익변경 (Rating/TP 상향·하향) ────────────────

def parse_추정이익변경(dry_run: bool) -> dict:
    # 크롤링본(WiseReport 요약, TP시트 없음) 말고 날짜본(Rating_Up/TP_Up 시트)을 잡는다
    path = _find_excel_ex("추정이익 변경", exclude=["크롤링"]) or find_excel("추정이익 변경")
    if not path:
        return {"error": "추정이익 변경 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    results = {
        "Rating_Up": [], "Rating_Down": [],
        "TP_Up": [],     "TP_Down": [],
        "New_KOSPI": [],  "New_KOSDAQ": [],
    }

    sheet_map = {
        "Rating_Up":   results["Rating_Up"],
        "Rating_Down": results["Rating_Down"],
        "TP_Up":       results["TP_Up"],
        "TP_Down":     results["TP_Down"],
        "New_KOSPI":   results["New_KOSPI"],
        "New_KOSDAQ":  results["New_KOSDAQ"],
    }

    for sheet_name, bucket in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = rows_from_sheet(ws, max_row=200, max_col=15)

        # 헤더 행 찾기: '일자' 또는 '코드' 있는 행
        header_idx = None
        for i, r in enumerate(rows):
            vals = [str(v) for v in r if v is not None]
            if any(k in vals for k in ["일자", "코드", "From"]):
                header_idx = i
                break

        # 데이터 행 (헤더 다음 2행부터)
        if header_idx is None:
            continue

        for r in rows[header_idx + 2:]:
            # 날짜, 코드, 종목명 위치 탐색 (컬럼 1,2,3)
            vals = list(r)
            # 날짜가 datetime이고 코드가 'A'로 시작하는 행
            date_v = vals[1] if len(vals) > 1 else None
            code_v = vals[2] if len(vals) > 2 else None
            name_v = vals[3] if len(vals) > 3 else None

            if not isinstance(date_v, (datetime, date)): continue
            if not isinstance(code_v, str) or not code_v.startswith("A"): continue
            if not name_v: continue

            bucket.append({
                "date":      fmt_date(date_v),
                "code":      code_v,
                "name":      str(name_v).strip(),
                "brokerage": str(vals[5]).strip() if vals[5] else "—",
                "analyst":   str(vals[6]).strip() if vals[6] else "—",
                "op_old":    str(vals[11]).strip() if len(vals) > 11 and vals[11] else "—",
                "op_new":    str(vals[12]).strip() if len(vals) > 12 and vals[12] else "—",
                "tp_old":    int(vals[13]) if len(vals) > 13 and isinstance(vals[13], (int, float)) else None,
                "tp_new":    int(vals[14]) if len(vals) > 14 and isinstance(vals[14], (int, float)) else None,
            })

    wb.close()

    # wiki 업데이트
    updated = []
    if not dry_run:
        all_items = results["Rating_Up"] + results["TP_Up"] + results["Rating_Down"] + results["TP_Down"]
        for item in all_items:
            page = find_stock_page(item["name"])
            if not page: continue
            updated.append(_update_consensus(page, item))

    return {"file": path.name, "results": results, "updated": updated}


def _update_consensus(page: Path, item: dict) -> str:
    """stock 페이지의 증권사 컨센서스 섹션에 1행 덮어쓰기"""
    text = page.read_text(encoding="utf-8")

    tp_old_str = fmt_num(item["tp_old"], "원") if item["tp_old"] else "—"
    tp_new_str = fmt_num(item["tp_new"], "원") if item["tp_new"] else "—"

    if item["tp_old"] and item["tp_new"] and item["tp_old"] > 0:
        chg = (item["tp_new"] - item["tp_old"]) / item["tp_old"]
        direction = f"↑ {chg:+.0%}" if chg > 0 else f"↓ {chg:+.0%}"
    else:
        direction = "신규"

    new_row = f"| {item['brokerage']} | {tp_new_str} | {item['op_new']} | {item['date']} | {tp_old_str} | {direction} |"

    # 기존 해당 증권사 행 찾아 덮어쓰기
    pattern = re.compile(rf"^\| {re.escape(item['brokerage'])} \|.*$", re.MULTILINE)
    if pattern.search(text):
        new_text = pattern.sub(new_row, text)
    else:
        # 컨센서스 테이블 끝에 추가
        consensus_end = text.find("\n---", text.find("## 증권사 컨센서스"))
        if consensus_end > 0:
            new_text = text[:consensus_end] + "\n" + new_row + text[consensus_end:]
        else:
            new_text = text + f"\n{new_row}"

    if new_text != text:
        page.write_text(new_text, encoding="utf-8")
        return f"✅ {item['name']} — {item['brokerage']} TP {tp_old_str}→{tp_new_str}"
    return f"⏭ {item['name']} 변경 없음"


# ─── 파서 2: 컨센움직임서프쇼크 ─────────────────────────────────
#
# 실제 컬럼 구조 (0-based, 검증 완료):
#   컨센상향/하향: r[1]=종목코드 r[2]=종목명 r[3]=업종
#                  r[4]=1M수익률 r[5]=3M수익률 r[6]=2026OP
#                  r[14]=컨센변화(정렬기준) r[16]=30일전컨센변화
#   서프라이즈/쇼크: r[1]=종목코드 r[2]=종목명 r[3]=업종
#                    r[4]=1M수익률 r[5]=3M수익률
#                    r[7]=1Q실제발표 r[11]=어닝서프율(정렬기준) r[12]=컨센변화

def parse_컨센움직임(dry_run: bool) -> dict:
    # 태린이아빠 개편(2026-07): '컨센움직임서프쇼크' → '투자픽업용'(컨센상향/하향/서프라이즈/쇼크 시트)
    path = find_excel("투자픽업용")
    if not path:
        path = find_excel("컨센움직임서프쇼크")   # 구버전 호환
    if not path:
        return {"error": "컨센움직임(투자픽업용) 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    # 마지막 업데이트 날짜
    last_update = None
    if "1개월 컨센 1개월일 변화" in wb.sheetnames:
        ws2 = wb["1개월 컨센 1개월일 변화"]
        for r in ws2.iter_rows(min_row=1, max_row=5, max_col=5, values_only=True):
            for v in r:
                if v and "Last Update" in str(v):
                    last_update = str(v).replace("Last Update : ", "").strip()[:10]
                    break

    sheets = {"컨센상향": [], "컨센하향": [], "서프라이즈": [], "쇼크": []}

    for sheet_name, bucket in sheets.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        is_surf = sheet_name in ("서프라이즈", "쇼크")

        for r in ws.iter_rows(min_row=1, max_row=500, max_col=22, values_only=True):
            code = r[1] if len(r) > 1 else None
            if not isinstance(code, str) or not code.startswith("A"): continue
            name = r[2] if len(r) > 2 else None
            if not name or not isinstance(name, str): continue

            if is_surf:
                bucket.append({
                    "code":          code,
                    "name":          name.strip(),
                    "sector":        str(r[3]).strip() if r[3] else "—",
                    "ret_1m":        r[4],
                    "ret_3m":        r[5],
                    "op_1q_actual":  r[7],   # 1Q 실제발표
                    "surprise_rate": r[11],  # 어닝서프율 (정렬 기준)
                    "csen_chg":      r[12],  # 컨센변화
                })
            else:
                bucket.append({
                    "code":      code,
                    "name":      name.strip(),
                    "sector":    str(r[3]).strip() if r[3] else "—",
                    "ret_1m":    r[4],
                    "ret_3m":    r[5],
                    "op_2026":   r[6],   # 2026 영업이익(십억)
                    "csen_chg":  r[14],  # 컨센변화율 (정렬 기준)
                    "csen_30d":  r[16],  # 30일전 컨센변화율
                })

    wb.close()
    return {"file": path.name, "last_update": last_update, "results": sheets}


# ─── 파서 3: 핵심 수출 데이터 ────────────────────────────────────

# (섹터 키, 관련 종목) — 섹터 키는 SECTOR_FOLDER_MAP 키와 동일
EXPORT_KEY_SHEETS: dict[str, tuple[str, list[str]]] = {
    "디램":            ("반도체",   ["삼성전자", "SK하이닉스"]),
    "낸드":            ("반도체",   ["삼성전자", "SK하이닉스"]),
    "HBM":             ("반도체",   ["SK하이닉스"]),
    "동박":            ("2차전지",  ["롯데에너지머티리얼즈", "SKC"]),
    "양극재(NCM+NCA)": ("2차전지",  ["에코프로비엠", "포스코퓨처엠"]),
    "MLCC":            ("반도체부품", ["삼성전기"]),
    "CCL":             ("기판",     ["두산"]),
    "선박 엔진":       ("조선",     ["HD현대중공업", "한화오션"]),
    "고용량변압기":    ("전력기기", ["HD현대일렉트릭", "효성중공업"]),
}

def parse_수출정리(dry_run: bool) -> dict:
    path = find_excel("수출정리")
    if not path:
        return {"error": "수출정리 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    results = {}

    for sheet_name, (sector, stocks) in EXPORT_KEY_SHEETS.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        rows = rows_from_sheet(ws, max_row=100, max_col=12)

        # 헤더 행 찾기 ('년월' 있는 행)
        data_start = None
        for i, r in enumerate(rows):
            if r[0] == "년월":
                data_start = i + 1
                break
        if data_start is None: continue

        # 최근 3개월 데이터
        data_rows = [r for r in rows[data_start:] if r[0] and str(r[0]).startswith("20")][-3:]

        monthly = []
        for r in data_rows:
            monthly.append({
                "month":     str(r[0]),
                "daily_avg": r[5],       # 일평균수출 (달러)
                "mom":       r[6],       # MoM
                "yoy_vol":   r[7],       # 물량 YoY
                "price_abs": r[8],       # 판가 절대값
                "price_yoy": r[9],       # 판가 YoY
            })

        # 신호 판단
        if monthly:
            latest = monthly[-1]
            yoy = latest["yoy_vol"]
            price_yoy = latest["price_yoy"]
            if isinstance(yoy, float) and isinstance(price_yoy, float):
                if yoy > 0.15 and price_yoy > 0.10:
                    signal = "🔴"
                elif yoy > 0 or price_yoy > 0:
                    signal = "🟠"
                else:
                    signal = "🟡"
            else:
                signal = "—"
        else:
            signal = "—"

        results[sheet_name] = {
            "sector": sector,
            "stocks": stocks,
            "monthly": monthly,
            "signal": signal,
        }

    wb.close()
    return {"file": path.name, "results": results}


# ─── 파서 4: 유동성체크 (컨센신고가 + 가속화모멘텀) ────────────────

def parse_유동성체크(dry_run: bool) -> dict:
    # 태린이아빠 개편(2026-07): '유동성 체크' → '카페라떼 회원용 종목피킹'
    #   (컨센신고가·가속화모멘텀·유동성컨셉 시트를 한 파일에 통합)
    path = find_excel("카페라떼 회원용 종목피킹")
    if not path:
        path = find_excel("유동성 체크")   # 구버전 호환
    if not path:
        return {"error": "유동성체크(카페라떼 종목피킹) 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    results = {"컨센신고가": [], "가속화모멘텀": []}

    # 컨센신고가 시트
    for sheet_name in ["2027년 컨센 신고가", "2027년 컨센 신고가(2)"]:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        for r in ws.iter_rows(min_row=2, max_row=100, max_col=10, values_only=True):
            rank = r[0]
            name = r[1]
            if not isinstance(rank, (int, float)): continue
            if not name: continue
            results["컨센신고가"].append({
                "rank":       int(rank),
                "name":       str(name).strip(),
                "latest_date": fmt_date(r[4]),
                "brokerage":  str(r[5]).strip() if r[5] else "—",
                "op_2027":    r[7],
            })

    # 가속화모멘텀 시트
    if "가속화모멘텀" in wb.sheetnames:
        ws = wb["가속화모멘텀"]
        rows = rows_from_sheet(ws, max_row=200, max_col=12)
        for r in rows:
            # 종목코드 형태 'A######' 있는 행
            for v in r:
                if isinstance(v, str) and re.match(r"A\d{6}", v):
                    name_idx = list(r).index(v) + 1
                    name = r[name_idx] if name_idx < len(r) else None
                    if name:
                        results["가속화모멘텀"].append(str(name).strip())
                    break

    wb.close()
    return {"file": path.name, "results": results}


# ─── 파서 5: 수급오실레이터 (COM 기반 EMA MACD 방식) ─────────────

def _run_osc_com(path: Path) -> dict:
    """COM으로 xlsm 열어 EMA MACD 오실레이터 계산 (대형주·중소형주 공용)"""
    try:
        import win32com.client as win32
    except ImportError:
        return {"error": "win32com 없음 — pip install pywin32"}

    K12, K26, K9 = 2/13, 2/27, 2/10
    DATA_START, DATA_END = 15, 91
    STAT_COL = 12

    def ema(vals, k):
        r = [vals[0]]
        for v in vals[1:]:
            r.append(v * k + r[-1] * (1 - k))
        return r

    def calc_osc(signal):
        e12 = ema(signal, K12); e26 = ema(signal, K26)
        macd = [a - b for a, b in zip(e12, e26)]
        sig  = ema(macd, K9)
        return [m - s for m, s in zip(macd, sig)]

    print("    [COM] Excel 열기...")
    xl = win32.Dispatch("Excel.Application")
    xl.Visible = False
    xl.AutomationSecurity = 3
    wb_com = xl.Workbooks.Open(str(path))

    ws_size = wb_com.Sheets("시가총액")
    ws_forn = wb_com.Sheets("외국인매수데이터")
    ws_inst = wb_com.Sheets("기관매수데이터")

    # 기준값: 수급오실레이터 시트가 있으면 읽고, 없으면 계산값으로 대체
    p10 = p25 = p75 = p90 = None
    if "수급오실레이터" in [wb_com.Sheets(i+1).Name for i in range(wb_com.Sheets.Count)]:
        ws_osc = wb_com.Sheets("수급오실레이터")
        try:
            p10 = float(ws_osc.Cells(11, STAT_COL).Value2)
            p25 = float(ws_osc.Cells(10, STAT_COL).Value2)
            p75 = float(ws_osc.Cells(8,  STAT_COL).Value2)
            p90 = float(ws_osc.Cells(7,  STAT_COL).Value2)
        except (TypeError, ValueError):
            pass

    last_date = ws_forn.Cells(DATA_END, 1).Value2
    max_col = ws_forn.UsedRange.Columns.Count

    xl_stocks = {}
    for c in range(2, max_col + 1):
        name = ws_forn.Cells(9, c).Value2
        code = ws_forn.Cells(8, c).Value2
        if name:
            xl_stocks[str(name)] = {"col": c, "code": str(code or "").replace("A", "")}

    def to_matrix(ws):
        raw = ws.Range(ws.Cells(DATA_START, 1), ws.Cells(DATA_END, max_col)).Value
        return raw or []

    mat_size = to_matrix(ws_size)
    mat_forn = to_matrix(ws_forn)
    mat_inst = to_matrix(ws_inst)
    wb_com.Close(False); xl.Quit()
    print(f"    [COM] Excel 닫힘. {len(xl_stocks)}종목 계산 중...")

    results = []
    for name, info in xl_stocks.items():
        c = info["col"] - 1
        size_v = [float(mat_size[r][c]) if mat_size[r][c] else None for r in range(len(mat_size))]
        forn_v = [float(mat_forn[r][c]) if mat_forn[r][c] else None for r in range(len(mat_forn))]
        inst_v = [float(mat_inst[r][c]) if mat_inst[r][c] else None for r in range(len(mat_inst))]
        pairs  = [((fv or 0) + (iv or 0), sv)
                  for fv, iv, sv in zip(forn_v, inst_v, size_v) if sv and sv > 0]
        if len(pairs) < 27:
            continue
        series = calc_osc([net / sz for net, sz in pairs])
        osc = series[-1]
        trend = "↑ 재진입" if len(series) >= 5 and osc > sum(series[-5:-1])/4 * 1.05 else \
                "↓ 빈집심화" if len(series) >= 5 and osc < sum(series[-5:-1])/4 * 0.95 else "→ 횡보"
        results.append({"name": name, "code": info["code"], "osc": osc, "trend": trend})

    if not results:
        return {"file": path.name, "error": "계산 가능 종목 없음"}

    # 기준값이 없으면 계산값 기반으로 산출
    all_sorted = sorted(r["osc"] for r in results)
    n = len(all_sorted)
    if p10 is None: p10 = all_sorted[int(n * 0.10)]
    if p25 is None: p25 = all_sorted[int(n * 0.25)]
    if p75 is None: p75 = all_sorted[int(n * 0.75)]
    if p90 is None: p90 = all_sorted[int(n * 0.90)]

    def pct(v):
        if v <= p10: return 10 * (v / p10) if p10 else 5
        if v <= p25: return 10 + (v - p10) / (p25 - p10) * 15
        if v <= p75: return 25 + (v - p25) / (p75 - p25) * 50
        if v <= p90: return 75 + (v - p75) / (p90 - p75) * 15
        return min(90 + (v - p90) / abs(p90) * 5 if p90 else 95, 100)

    빈집_A    = sorted([r for r in results if r["osc"] <= p10], key=lambda x: x["osc"])
    빈집_B    = sorted([r for r in results if p10 < r["osc"] <= p25], key=lambda x: x["osc"])
    과매수_하락 = sorted(
        [r for r in results if r["osc"] >= p75 and r.get("trend", "") == "↓ 빈집심화"],
        key=lambda x: x["osc"], reverse=True
    )[:30]
    for r in results:
        r["pct"] = pct(r["osc"])

    if isinstance(last_date, datetime):
        last_date_str = last_date.strftime("%Y-%m-%d")
    else:
        last_date_str = TODAY

    def _row(r): return {"name": r["name"], "code": r["code"], "osc": r["osc"], "pct": r["pct"], "trend": r["trend"]}
    return {
        "file":       path.name,
        "date":       last_date_str,
        "total":      len(results),
        "빈집_A":     [_row(r) for r in 빈집_A],
        "빈집_B":     [_row(r) for r in 빈집_B],
        "과매수_하락": [_row(r) for r in 과매수_하락],
        "note": f"전체 {len(results)}종목 | 빈집A: {len(빈집_A)}개 | 빈집B: {len(빈집_B)}개 | 기준A≤{p10:.6f} B≤{p25:.6f}",
    }


def _run_osc_openpyxl(path: Path) -> dict:
    """openpyxl로 xlsm 읽어 EMA MACD 오실레이터 계산 (COM 없이 동작)"""
    K12, K26, K9 = 2/13, 2/27, 2/10
    DATA_START, DATA_END = 15, 91
    STAT_COL = 12

    def ema(vals, k):
        r = [vals[0]]
        for v in vals[1:]:
            r.append(v * k + r[-1] * (1 - k))
        return r

    def calc_osc(signal):
        e12 = ema(signal, K12); e26 = ema(signal, K26)
        macd = [a - b for a, b in zip(e12, e26)]
        sig  = ema(macd, K9)
        return [m - s for m, s in zip(macd, sig)]

    print("    [openpyxl] 열기...")
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True, keep_vba=False)
    except Exception as e:
        return {"error": f"openpyxl 열기 실패: {e}"}

    ws_size = wb["시가총액"]
    ws_forn = wb["외국인매수데이터"]
    ws_inst = wb["기관매수데이터"]

    # 백분위 기준값 (수급오실레이터 시트)
    p10 = p25 = p75 = p90 = None
    if "수급오실레이터" in wb.sheetnames:
        ws_osc = wb["수급오실레이터"]
        try:
            p90 = float(ws_osc.cell(7,  STAT_COL).value)
            p75 = float(ws_osc.cell(8,  STAT_COL).value)
            p25 = float(ws_osc.cell(10, STAT_COL).value)
            p10 = float(ws_osc.cell(11, STAT_COL).value)
        except (TypeError, ValueError):
            pass

    # 마지막 날짜
    last_date_val = ws_forn.cell(DATA_END, 1).value
    if isinstance(last_date_val, datetime):
        last_date_str = last_date_val.strftime("%Y-%m-%d")
    else:
        last_date_str = TODAY

    # 종목 헤더 (row 9, col 2~)
    max_col = 0
    for cell in ws_forn[9]:
        if cell.value is not None:
            max_col = cell.column
        else:
            break
    if max_col < 2:
        max_col = 800

    xl_stocks = {}
    for c in range(2, max_col + 1):
        name = ws_forn.cell(9, c).value
        code = ws_forn.cell(8, c).value
        if not name:
            break
        xl_stocks[str(name)] = {"col": c, "code": str(code or "").replace("A", "")}

    print(f"    [openpyxl] {len(xl_stocks)}종목 계산 중...")

    # 데이터 행렬 읽기 (iter_rows 방식 — cell() 반복보다 10~20배 빠름)
    def read_matrix(ws):
        mat = []
        for i, row in enumerate(ws.iter_rows(min_row=DATA_START, max_row=DATA_END,
                                              min_col=1, max_col=max_col, values_only=True)):
            mat.append(list(row))
        return mat

    mat_size = read_matrix(ws_size)
    mat_forn = read_matrix(ws_forn)
    mat_inst = read_matrix(ws_inst)
    wb.close()

    results = []
    for name, info in xl_stocks.items():
        c = info["col"] - 1
        size_v = [float(mat_size[r][c]) if mat_size[r][c] else None for r in range(len(mat_size))]
        forn_v = [float(mat_forn[r][c]) if mat_forn[r][c] else None for r in range(len(mat_forn))]
        inst_v = [float(mat_inst[r][c]) if mat_inst[r][c] else None for r in range(len(mat_inst))]
        pairs  = [((fv or 0) + (iv or 0), sv)
                  for fv, iv, sv in zip(forn_v, inst_v, size_v) if sv and sv > 0]
        if len(pairs) < 27:
            continue
        series = calc_osc([net / sz for net, sz in pairs])
        osc = series[-1]
        trend = "↑ 재진입" if len(series) >= 5 and osc > sum(series[-5:-1])/4 * 1.05 else \
                "↓ 빈집심화" if len(series) >= 5 and osc < sum(series[-5:-1])/4 * 0.95 else "→ 횡보"
        results.append({"name": name, "code": info["code"], "osc": osc, "trend": trend})

    if not results:
        return {"file": path.name, "error": "계산 가능 종목 없음"}

    all_sorted = sorted(r["osc"] for r in results)
    n = len(all_sorted)
    if p10 is None: p10 = all_sorted[int(n * 0.10)]
    if p25 is None: p25 = all_sorted[int(n * 0.25)]
    if p75 is None: p75 = all_sorted[int(n * 0.75)]
    if p90 is None: p90 = all_sorted[int(n * 0.90)]

    def pct(v):
        if v <= p10: return 10 * (v / p10) if p10 else 5
        if v <= p25: return 10 + (v - p10) / (p25 - p10) * 15
        if v <= p75: return 25 + (v - p25) / (p75 - p25) * 50
        if v <= p90: return 75 + (v - p75) / (p90 - p75) * 15
        return min(90 + (v - p90) / abs(p90) * 5 if p90 else 95, 100)

    빈집_A    = sorted([r for r in results if r["osc"] <= p10], key=lambda x: x["osc"])
    빈집_B    = sorted([r for r in results if p10 < r["osc"] <= p25], key=lambda x: x["osc"])
    과매수_상승 = sorted(
        [r for r in results if r["osc"] >= p75 and r.get("trend", "") == "↑ 재진입"],
        key=lambda x: x["osc"], reverse=True
    )[:30]
    과매수_하락 = sorted(
        [r for r in results if r["osc"] >= p75 and r.get("trend", "") == "↓ 빈집심화"],
        key=lambda x: x["osc"], reverse=True
    )[:30]
    for r in results:
        r["pct"] = pct(r["osc"])

    def _row(r): return {"name": r["name"], "code": r["code"], "osc": r["osc"], "pct": r["pct"], "trend": r["trend"]}
    return {
        "file":       path.name,
        "date":       last_date_str,
        "total":      len(results),
        "빈집_A":     [_row(r) for r in 빈집_A],
        "빈집_B":     [_row(r) for r in 빈집_B],
        "과매수_상승": [_row(r) for r in 과매수_상승],
        "과매수_하락": [_row(r) for r in 과매수_하락],
        "note": f"전체 {len(results)}종목 | 빈집A: {len(빈집_A)}개 | 빈집B: {len(빈집_B)}개 | 기준A≤{p10:.6f} B≤{p25:.6f}",
    }


def _run_osc(path: Path) -> dict:
    """COM 우선, 실패 시 openpyxl 폴백"""
    try:
        import win32com.client as win32
        return _run_osc_com(path)
    except Exception:
        print("    [COM 실패] openpyxl 폴백 사용")
        return _run_osc_openpyxl(path)


def parse_수급오실레이터(dry_run: bool) -> dict:
    path = find_excel("(700)")
    if not path:
        path = find_excel("수급오실레이터")
    if not path:
        return {"error": "수급오실레이터(700) 파일 없음"}
    return _run_osc(path)


# ─── 파서 6: 중소형주 오실레이터 (700-1400) ───────────────────────

def parse_중소형주오실레이터(dry_run: bool) -> dict:
    path = find_excel("(700-1400)")
    if not path:
        return {"error": "수급오실레이터(700-1400) 파일 없음"}
    return _run_osc(path)


# ─── 파서 7: 가속화모멘텀 ────────────────────────────────────────

def parse_가속화모멘텀(dry_run: bool) -> dict:
    # 태린이아빠 개편(2026-07): 가속화모멘텀 시트가 '카페라떼 회원용 종목피킹'에 포함
    path = find_excel("카페라떼 회원용 종목피킹")
    if not path:
        path = find_excel("유동성 체크")   # 구버전 호환
    if not path:
        return {"error": "가속화모멘텀(카페라떼 종목피킹) 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}
    if "가속화모멘텀" not in wb.sheetnames:
        wb.close(); return {"error": "가속화모멘텀 시트 없음"}

    ws = wb["가속화모멘텀"]

    # 4개 그룹: 각 그룹 = (순위col, 종목col, 스코어col, 변화24col, 변화25col)
    groups = {
        "추정이익3개+": (1, 2, 3, 4, 5),
        "추정이익1개+": (6, 7, 8, 9, 10),
        "시총상위300":  (11, 12, 13, 14, 15),
        "주당순이익1개+": (16, 17, 18, 19, 20),  # 태린이아빠 선호
    }

    result = {}
    for g_name, (c_rank, c_name, c_score, c_24, c_25) in groups.items():
        items = []
        for r in range(6, ws.max_row + 1):
            name = ws.cell(r, c_name).value
            score = ws.cell(r, c_score).value
            if not name or not isinstance(name, str): continue
            try: score = float(score)
            except (TypeError, ValueError): continue
            chg24 = ws.cell(r, c_24).value
            chg25 = ws.cell(r, c_25).value
            items.append({
                "name":   name.strip(),
                "score":  round(score, 4),
                "chg24":  float(chg24) if isinstance(chg24, (int, float)) else None,
                "chg25":  float(chg25) if isinstance(chg25, (int, float)) else None,
            })
        result[g_name] = items

    wb.close()
    return {"file": path.name, "results": result}


# ─── 파서 8: 상대강도 RS ─────────────────────────────────────────

def _rs_sorted_from_rows(all_rows) -> tuple[list, str]:
    """종가류 시트 rows(첫줄 헤더: DATE,코스피,종목...) → (norm_RS_avg 내림차순 리스트, 마지막날짜)"""
    import math
    if not all_rows:
        return [], TODAY
    headers = [str(v).strip() if v else None for v in all_rows[0]]
    prices = {h: [] for h in headers[1:] if h}
    dates = []
    for row in all_rows[1:]:
        if row[0] is None: continue
        dates.append(row[0])
        for c, h in enumerate(headers[1:], 1):
            if not h: continue
            v = row[c] if c < len(row) else None
            prices[h].append(float(v) if isinstance(v, (int, float)) else None)
    if not dates:
        return [], TODAY

    def mansfield_rs(price_list, kospi_list, ma_period):
        """Mansfield RS: (log(stock/kospi) - rolling_mean) * 100"""
        if len(price_list) < ma_period + 1:
            return None
        log_rel = []
        for p, k in zip(price_list, kospi_list):
            if p and k and p > 0 and k > 0:
                log_rel.append(math.log(p / k))
            else:
                log_rel.append(None)
        rs_series = []
        for i in range(len(log_rel)):
            window = [v for v in log_rel[max(0, i - ma_period + 1):i + 1] if v is not None]
            if len(window) >= ma_period and log_rel[i] is not None:
                ma = sum(window) / len(window)
                rs_series.append((log_rel[i] - ma) * 100)
            else:
                rs_series.append(None)
        return next((v for v in reversed(rs_series) if v is not None), None)

    def norm_sigmoid(x, scale=12):
        return round(100 / (1 + math.exp(-x / scale)), 2)

    kospi_list = prices.get("코스피", [])
    results = {}
    for stock, plist in prices.items():
        if stock == "코스피": continue
        if len(plist) < 63: continue
        row = {"name": stock}
        for period in [60, 120, 250]:
            raw = mansfield_rs(plist, kospi_list, period)
            if raw is not None:
                row[f"RS_{period}d"]    = round(raw, 2)
                row[f"norm_RS_{period}"] = norm_sigmoid(raw)
        results[stock] = row

    for v in results.values():
        rs_vals = [v[f"RS_{p}d"]      for p in [60, 120, 250] if f"RS_{p}d"      in v]
        nr_vals = [v[f"norm_RS_{p}"]  for p in [60, 120, 250] if f"norm_RS_{p}"  in v]
        if rs_vals:  v["RS_avg"]      = round(sum(rs_vals) / len(rs_vals), 2)
        if nr_vals:  v["norm_RS_avg"] = round(sum(nr_vals) / len(nr_vals), 2)

    sorted_r = sorted(
        [v for v in results.values() if "norm_RS_avg" in v],
        key=lambda x: x["norm_RS_avg"], reverse=True
    )
    last_date = str(dates[-1])[:10] if dates else TODAY
    return sorted_r, last_date


def _find_excel_ex(pattern: str, exclude=()) -> Path | None:
    """pattern 포함 + exclude 미포함 파일 중 이름이 가장 짧은(=기본) 것"""
    cands = [f for f in EXCEL_DIR.iterdir()
             if pattern in f.name and f.suffix in (".xlsx", ".xlsm")
             and not any(x in f.name for x in exclude)]
    return sorted(cands, key=lambda p: len(p.name))[0] if cands else None


def _load_rs_rows(pattern: str, sheet_candidates, exclude=()):
    path = _find_excel_ex(pattern, exclude)
    if not path:
        return None, None
    wb = load_wb(path)
    if not wb:
        return None, path
    sheet = next((s for s in sheet_candidates if s in wb.sheetnames), None)
    if not sheet:
        wb.close(); return None, path
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    return rows, path


def parse_rs(dry_run: bool) -> dict:
    # 태린이아빠 개편(2026-07): 개별종목 RS + ETF RS(소라티노) 둘 다 활용
    #   개별종목 → 종목상대강도데이터/종가 (per-stock 패널용)
    #   ETF(소라티노) → etf상대강도데이터/데이터 (섹터·ETF 강도용; irp/해외포함본 제외)
    stock_rows, stock_path = _load_rs_rows("종목상대강도데이터", ["종가"])
    if stock_rows is None:
        stock_rows, stock_path = _load_rs_rows("한국상대강도", ["종가"])   # 구버전 호환
    etf_rows, etf_path = _load_rs_rows("etf상대강도데이터", ["데이터", "종가"],
                                       exclude=["irp", "해외"])

    if stock_rows is None and etf_rows is None:
        return {"error": "상대강도(종목/ETF) 파일 없음"}

    stock_sorted, stock_date = _rs_sorted_from_rows(stock_rows or [])
    etf_sorted, etf_date = _rs_sorted_from_rows(etf_rows or [])

    src = stock_path or etf_path
    return {
        "file":     src.name if src else "",
        "date":     stock_date if stock_rows else etf_date,
        "total":    len(stock_sorted),
        "top30":    stock_sorted[:30],
        "bottom10": stock_sorted[-10:],
        # ETF RS(소라티노) — 섹터/ETF 강도용. 종목코드 매칭 안돼 per-stock 패널엔 미반영.
        "etf_file":     etf_path.name if etf_path else None,
        "etf_total":    len(etf_sorted),
        "etf_top30":    etf_sorted[:30],
        "etf_bottom10": etf_sorted[-10:],
    }


# ─── 파서 9: 업종 쏠림지수 ─────────────────────────────────────

def parse_쏠림지수(dry_run: bool) -> dict:
    path = find_excel("특정업종쏠림지수국내")
    if not path:
        return {"error": "특정업종쏠림지수국내 파일 없음"}
    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    # 상대강도 시트 — 업종별 RS 데이터
    if "상대강도" not in wb.sheetnames:
        wb.close(); return {"error": "상대강도 시트 없음"}

    ws = wb["상대강도"]
    all_rows = list(ws.iter_rows(min_row=8, max_row=2000, max_col=60, values_only=True))

    # 헤더: Code행(8), Name행(9)
    codes = list(all_rows[0])
    names = list(all_rows[1])
    name_map = {i: str(names[i]) for i in range(len(names)) if names[i] and names[i] != "Name"}

    # 날짜 데이터
    data_rows = [r for r in all_rows[4:] if r[0] is not None]
    if not data_rows:
        wb.close(); return {"error": "데이터 없음"}

    # 최신 30일 수익률 계산
    last_30 = data_rows[-30:]
    first_prices = {i: None for i in name_map}
    last_prices  = {i: None for i in name_map}

    for row in last_30:
        for i in name_map:
            if i < len(row) and isinstance(row[i], (int, float)):
                if first_prices[i] is None:
                    first_prices[i] = row[i]
                last_prices[i] = row[i]

    # 최종 시트 — 쏠림지수 (붉은선) 분석
    # 컬럼: 0=DATE, 1=코스피, 2=코스닥, 4=롱숏, 5=지수화, 6=12EMA, 7=26EMA, 8=MACD, 9=시그널, 10=오실레이터
    idx_val = macd_val = osc_val = None
    direction = None  # "up" or "down"

    if "최종" in wb.sheetnames:
        ws2 = wb["최종"]
        macd_rows = list(ws2.iter_rows(min_row=15, max_row=3000, max_col=11, values_only=True))
        valid = [r for r in macd_rows if r[0] is not None and isinstance(r[0], datetime)]
        if len(valid) >= 6:
            last_r  = valid[-1]
            prev5   = valid[-6:-1]
            try:
                idx_val  = round(float(last_r[5]),  2) if len(last_r) > 5  and isinstance(last_r[5],  (int,float)) else None
                macd_val = round(float(last_r[8]),  4) if len(last_r) > 8  and isinstance(last_r[8],  (int,float)) else None
                osc_val  = round(float(last_r[10]), 4) if len(last_r) > 10 and isinstance(last_r[10], (int,float)) else None
                # 5일 전 대비 지수화 방향
                prev_idx = next((r[5] for r in reversed(prev5) if r[5] and isinstance(r[5], (int,float))), None)
                if idx_val and prev_idx:
                    direction = "up" if idx_val > prev_idx else "down"
            except:
                pass

    wb.close()
    return {
        "file":      path.name,
        "idx":       idx_val,    # 지수화 현재값 (붉은선 레벨)
        "macd":      macd_val,   # MACD
        "osc":       osc_val,    # 오실레이터 (MACD 모멘텀)
        "direction": direction,  # 최근 5일 방향
    }


# ─── 파서 10: 액티브ETF 비중 변화 ────────────────────────────────

SECTOR_SHEETS_ETF = [
    "코스닥액티브", "반도체", "수급, 배당성장", "코스닥액티브2, 배당성장",
    "신재생, 2차전지", "이노, 소비, AI인프라", "코스피, 조선, 테크",
    "수출, 로봇, 컬쳐", "밸류업, 제조업, 바이오", "바이오"
]

def parse_액티브ETF(dry_run: bool) -> dict:
    # 태린이아빠 개편(2026-07): '액티브ETF관리' → '액티브ETF를 관찰하자'
    path = find_excel("액티브ETF를 관찰하자")
    if not path:
        path = find_excel("액티브ETF관리")   # 구버전 호환
    if not path:
        return {"error": "액티브ETF(관찰하자) 파일 없음"}
    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    increase, decrease = [], []

    for sname in SECTOR_SHEETS_ETF:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, max_row=100, max_col=60, values_only=True))

        # ETF 이름: 행3(idx=2)에서 col 1, 7, 13... (6 간격)
        etf_name_row = rows[2] if len(rows) > 2 else []
        etf_blocks = []  # (블록시작col, etf이름)
        for j, v in enumerate(etf_name_row):
            if v and isinstance(v, str) and len(str(v)) > 3:
                # 데이터 블록 시작 = j-1 (이름 컬럼 -1 = mkt 컬럼)
                etf_blocks.append((j - 1, str(v)))

        # 데이터 행 시작
        data_start = None
        for i, row in enumerate(rows):
            if row and row[0] in ("KS", "KQ"):
                data_start = i; break
        if data_start is None: continue

        # 각 ETF 블록: start=mkt, start+1=name, start+4=diff, start+5=rate
        for blk_start, etf_name in etf_blocks:
            if blk_start < 0: continue
            for row in rows[data_start:]:
                if blk_start >= len(row): continue
                mkt = row[blk_start]
                if mkt not in ("KS", "KQ"): continue
                name = row[blk_start + 1] if blk_start + 1 < len(row) else None
                diff = row[blk_start + 4] if blk_start + 4 < len(row) else None
                rate = row[blk_start + 5] if blk_start + 5 < len(row) else None
                if not name or not isinstance(diff, (int, float)): continue
                if diff > 0.3:
                    increase.append({"etf": etf_name, "name": str(name), "diff": round(diff, 2), "rate": round(float(rate)*100, 1) if isinstance(rate, float) else 0})
                elif diff < -0.3:
                    decrease.append({"etf": etf_name, "name": str(name), "diff": round(diff, 2), "rate": round(float(rate)*100, 1) if isinstance(rate, float) else 0})

    wb.close()

    # 중복 제거 (종목명 기준 최대 비중차이)
    inc_map, dec_map = {}, {}
    for it in increase:
        if it["name"] not in inc_map or it["diff"] > inc_map[it["name"]]["diff"]:
            inc_map[it["name"]] = it
    for it in decrease:
        if it["name"] not in dec_map or it["diff"] < dec_map[it["name"]]["diff"]:
            dec_map[it["name"]] = it

    return {
        "file": path.name,
        "increase": sorted(inc_map.values(), key=lambda x: x["diff"], reverse=True)[:20],
        "decrease": sorted(dec_map.values(), key=lambda x: x["diff"])[:10],
    }


# ─── 파서 11: 카페라떼 일정 D-30 ───────────────────────────────

def parse_일정(dry_run: bool) -> dict:
    path = find_excel("일정 및 수주잔고")
    if not path:
        return {"error": "일정 및 수주잔고 파일 없음"}
    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    from datetime import timedelta
    import re as _re

    today = datetime.today()
    d7  = today + timedelta(days=7)
    d30 = today + timedelta(days=30)
    events_7, events_30 = [], []

    if "25년 말~26년" not in wb.sheetnames:
        wb.close(); return {"error": "일정 시트 없음"}

    ws = wb["25년 말~26년"]
    for row in ws.iter_rows(min_row=2, max_row=500, max_col=6, values_only=True):
        # 컬럼: [0]=None, [1]=년도, [2]=날짜, [3]=관련종목, [4]=내용
        raw_d = row[2] if len(row) > 2 else None
        if not raw_d: continue
        dt = None
        try:
            if isinstance(raw_d, datetime):
                dt = raw_d
            elif isinstance(raw_d, str):
                # '6월 15~17일', '6월 8일', '6월 말' 등
                m = _re.search(r'(\d+)월\s*(\d+)', raw_d)
                if m:
                    yr = int(str(row[1] or "2026").replace("년","").strip()) if row[1] else 2026
                    dt = datetime(yr, int(m.group(1)), int(m.group(2)))
        except:
            continue
        if not dt: continue
        if not (today <= dt <= d30): continue

        related = str(row[3] or "").strip()
        content  = str(row[4] or row[3] or "").strip()
        label    = related if related else content
        if not label: continue

        ev = {"date": dt.strftime("%Y-%m-%d"), "related": related, "content": content or related}
        if dt <= d7:
            events_7.append(ev)
        else:
            events_30.append(ev)

    wb.close()
    return {
        "file": path.name,
        "d7":   events_7,
        "d30":  events_30,
    }


# ─── 파서 12: 한국ETF 상대강도 (Mansfield RS) ──────────────────────

def _etf_file_fresh(path) -> bool:
    """파일 내 데이터 마지막 날짜가 10일 이내인지 확인"""
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        dates = [r[0] for r in ws.iter_rows(min_row=2, max_row=5000, max_col=1, values_only=True) if r[0]]
        wb.close()
        if not dates: return False
        last = dates[-1] if isinstance(dates[-1], datetime) else datetime.strptime(str(dates[-1])[:10], "%Y-%m-%d")
        return (datetime.today() - last).days <= 10
    except Exception:
        return False


def parse_한국ETF상대강도(dry_run: bool) -> dict:
    # 한국ETF상대강도 신선하면 사용, 오래됐으면 소라티노ETF 파일로 폴백
    path = find_excel("한국ETF상대강도")
    fallback_used = False
    if not path or not _etf_file_fresh(path):
        path = find_excel("소라티노ETF상대강도")
        fallback_used = True
    if not path:
        return {"error": "한국ETF상대강도 / 소라티노ETF상대강도 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    sheet_name = "데이터" if "데이터" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {"error": "데이터 없음"}

    headers = [str(v).strip() if v else None for v in all_rows[0]]
    prices = {h: [] for h in headers[1:] if h}
    dates = []

    for row in all_rows[1:]:
        if row[0] is None: continue
        dates.append(row[0])
        for c, h in enumerate(headers[1:], 1):
            if not h: continue
            v = row[c] if c < len(row) else None
            prices[h].append(float(v) if isinstance(v, (int, float)) else None)

    if not dates:
        return {"error": "데이터 없음"}

    import math

    def _mansfield(price_list, kospi_list, ma_period):
        log_rel = []
        for p, k in zip(price_list, kospi_list):
            if p and k and p > 0 and k > 0:
                log_rel.append(math.log(p / k))
            else:
                log_rel.append(None)
        rs_series = []
        for i in range(len(log_rel)):
            window = [v for v in log_rel[max(0, i - ma_period + 1):i + 1] if v is not None]
            if len(window) >= ma_period and log_rel[i] is not None:
                ma = sum(window) / len(window)
                rs_series.append((log_rel[i] - ma) * 100)
            else:
                rs_series.append(None)
        return next((v for v in reversed(rs_series) if v is not None), None)

    def _norm(x, scale=12):
        return round(100 / (1 + math.exp(-x / scale)), 2)

    kospi_key = next((k for k in prices if "코스피" in k), None)
    kospi_list = prices.get(kospi_key, []) if kospi_key else []

    results = []
    for etf, plist in prices.items():
        if etf == kospi_key: continue
        if len(plist) < 63: continue
        row = {"name": etf}
        rs_vals, nr_vals = [], []
        for period in [60, 120, 250]:
            raw = _mansfield(plist, kospi_list, period)
            if raw is not None:
                row[f"RS_{period}d"] = round(raw, 2)
                row[f"norm_{period}"] = _norm(raw)
                rs_vals.append(raw); nr_vals.append(_norm(raw))
        if rs_vals:
            row["RS_avg"]       = round(sum(rs_vals) / len(rs_vals), 2)
            row["norm_RS_avg"]  = round(sum(nr_vals) / len(nr_vals), 2)
            results.append(row)

    sorted_r = sorted(results, key=lambda x: x.get("norm_RS_avg", 0), reverse=True)
    last_date = str(dates[-1])[:10] if dates else TODAY
    return {
        "file":   path.name,
        "date":   last_date,
        "total":  len(sorted_r),
        "top20":  sorted_r[:20],
        "rs70+":  [v for v in sorted_r if v.get("norm_RS_avg", 0) >= 70],
    }


# ─── 파서 13: 투자아이디어정리 ──────────────────────────────────────

def parse_투자아이디어(dry_run: bool) -> dict:
    path = find_excel("투자아이디어정리")
    if not path:
        return {"error": "투자아이디어정리 파일 없음"}

    wb = load_wb(path)
    if not wb: return {"error": "열기 실패"}

    today_dt = datetime.today()
    month = today_dt.month
    year  = today_dt.year
    q     = (month - 1) // 3 + 1
    next_q = q + 1 if q < 4 else 1
    next_y = year if q < 4 else year + 1

    # 분기 문자열 매칭 (다양한 표기 대응: "2026 2Q", "26년 2Q", "2Q26" 등)
    def match_q(v) -> bool:
        if not isinstance(v, str): return False
        v = v.strip()
        targets = [
            f"{year} {q}Q", f"{year % 100}년 {q}Q", f"{q}Q{year % 100}",
            f"{next_y} {next_q}Q", f"{next_y % 100}년 {next_q}Q", f"{next_q}Q{next_y % 100}",
        ]
        return any(t in v for t in targets)

    SKIP_SHEETS = {"정기변경", "Sheet1", "해외주식"}
    ideas = []

    def _clean(v) -> str:
        if v is None: return ""
        s = str(v).strip().lstrip("'")  # Excel 텍스트 prefix ' 제거
        return s

    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS: continue
        # IDEA_SHEET_SECTOR_MAP에서 섹터 확인 (None이면 제외)
        sector_key = IDEA_SHEET_SECTOR_MAP.get(sname, sname)  # 매핑 없으면 시트명=종목명으로 간주
        if sector_key is None: continue

        ws = wb[sname]
        rows = list(ws.iter_rows(min_row=1, max_row=200, max_col=6, values_only=True))
        if not rows: continue

        # 데이터가 B열(idx=1)에 있음 — 종목명: 헤더('분기') 이전 첫 비어있지 않은 B열 값
        company = None
        for r in rows[:6]:
            v = _clean(r[1] if len(r) > 1 else None)
            if v and "분기" not in v and len(v) > 1:
                company = v
                break
        if not company:
            company = sname

        # 섹터 폴더 (IDEA_SHEET_SECTOR_MAP 사용, 없으면 종목명으로 wiki 검색)
        if sname in IDEA_SHEET_SECTOR_MAP and IDEA_SHEET_SECTOR_MAP[sname]:
            sector_folder = get_sector_folder(IDEA_SHEET_SECTOR_MAP[sname])
        else:
            sector_folder = get_sector_folder(get_stock_sector(company) or "")

        # 헤더 행: B열에 '분기' 있는 행
        data_start = None
        for i, r in enumerate(rows):
            v = _clean(r[1] if len(r) > 1 else None)
            if "분기" in v:
                data_start = i + 1
                break
        if data_start is None: continue

        for r in rows[data_start:]:
            quarter_v  = _clean(r[1] if len(r) > 1 else None)
            momentum_v = _clean(r[2] if len(r) > 2 else None)
            detail_v   = _clean(r[3] if len(r) > 3 else None)
            if not match_q(quarter_v): continue
            if not momentum_v: continue
            ideas.append({
                "sheet":         sname,
                "company":       company,
                "sector":        IDEA_SHEET_SECTOR_MAP.get(sname, get_stock_sector(company) or "—"),
                "sector_folder": str(sector_folder) if sector_folder else None,
                "quarter":       quarter_v,
                "momentum":      momentum_v[:80],
                "detail":        detail_v[:100],
            })

    wb.close()
    quarters_label = f"{year} {q}Q / {next_y} {next_q}Q"
    return {
        "file":    path.name,
        "quarters": quarters_label,
        "ideas":   ideas,
    }


# ─── 텔레그램 빌더 — 쏠림/ETF/일정 ───────────────────────────────

def _build_컨센움직임_tg(r: dict) -> str:
    last_upd = r.get("last_update", TODAY)
    res = r.get("results", {})
    lines = [f"📊 <b>컨센움직임·서프·쇼크</b> — {last_upd}", ""]

    # 컨센상향 TOP10
    up = sorted(
        [it for it in res.get("컨센상향", []) if isinstance(it.get("csen_chg"), (int, float))],
        key=lambda x: x["csen_chg"], reverse=True
    )[:10]
    if up:
        lines.append(f"<b>📈 컨센상향 TOP10 (컨센변화율 기준)</b>")
        for it in up:
            chg = f"{it['csen_chg']:+.2f}" if isinstance(it['csen_chg'], float) else "—"
            r1m = f"{it['ret_1m']:+.1f}%" if isinstance(it['ret_1m'], (int, float)) else "—"
            lines.append(f"  {it['name']}({it['sector']})  컨센{chg}  1M{r1m}")
        lines.append("")

    # 서프라이즈 TOP10
    surf = sorted(
        [it for it in res.get("서프라이즈", []) if isinstance(it.get("surprise_rate"), (int, float))],
        key=lambda x: x["surprise_rate"], reverse=True
    )[:10]
    if surf:
        lines.append(f"<b>🎉 어닝서프라이즈 TOP10</b>")
        for it in surf:
            sr = f"{it['surprise_rate']:+.2f}" if isinstance(it['surprise_rate'], float) else "—"
            r1m = f"{it['ret_1m']:+.1f}%" if isinstance(it['ret_1m'], (int, float)) else "—"
            lines.append(f"  {it['name']}({it['sector']})  서프{sr}  1M{r1m}")
        lines.append("")

    # 쇼크 (피해야 할 종목, 간략히)
    shock = sorted(
        [it for it in res.get("쇼크", []) if isinstance(it.get("surprise_rate"), (int, float))],
        key=lambda x: x["surprise_rate"]
    )[:5]
    if shock:
        lines.append(f"<b>💥 어닝쇼크 TOP5 (주의)</b>")
        for it in shock:
            sr = f"{it['surprise_rate']:+.2f}" if isinstance(it['surprise_rate'], float) else "—"
            lines.append(f"  {it['name']}({it['sector']})  쇼크{sr}")

    return "\n".join(lines)


def _build_쏠림_tg(r: dict) -> str:
    idx  = r.get("idx")
    macd = r.get("macd")
    osc  = r.get("osc")
    dirn = r.get("direction")

    # 시장 판단
    if macd is not None and macd > 0:
        if dirn == "up":
            judge = "🔥 쏠림 장세 — 주도업종 집중 상승 중\n→ 주도섹터 안에서 종목 선별 필수"
        else:
            judge = "🟡 쏠림 약화 중 — 순환매 전환 가능성\n→ 주도업종 유지 여부 체크"
    elif macd is not None and macd <= 0:
        judge = "🔄 순환매 장세 — 여러 종목 분산 상승\n→ 종목 선택 폭 넓음, 수익내기 상대적으로 쉬운 장"
    else:
        judge = "—"

    lines = [
        f"🌊 <b>업종 쏠림지수 (시장 상단 판단)</b> — {TODAY}",
        "",
        judge,
        "",
        f"붉은선(지수화): <code>{idx:,.2f}</code>" if idx else "",
        f"MACD: <code>{macd:+.2f}</code>  오실레이터: <code>{osc:+.4f}</code>" if macd else "",
        f"방향: {'↑ 상승 (쏠림 강화)' if dirn=='up' else '↓ 하락 (순환매 진행)' if dirn=='down' else '—'}",
    ]
    return "\n".join(l for l in lines if l is not None)


def _build_액티브ETF_tg(r: dict) -> str:
    inc = r.get("increase", [])
    dec = r.get("decrease", [])
    lines = [f"📊 <b>액티브ETF 비중 변화</b> — {TODAY}", ""]
    if inc:
        lines.append(f"<b>📈 편입 증가 ({len(inc)}종목)</b>")
        for it in inc[:15]:
            lines.append(f"  {it['name']}  <code>+{it['diff']:.2f}%</code>  ({it['etf'][:8]})")
        lines.append("")
    if dec:
        lines.append(f"<b>📉 편출 감소 ({len(dec)}종목)</b>")
        for it in dec[:8]:
            lines.append(f"  {it['name']}  <code>{it['diff']:.2f}%</code>")
    return "\n".join(lines)


def _build_일정_tg(r: dict) -> str:
    d7  = r.get("d7",  [])
    d30 = r.get("d30", [])
    lines = [f"📅 <b>이벤트 일정</b> — {TODAY}", ""]
    if d7:
        lines.append(f"<b>🚨 D-7 이내 ({len(d7)}건)</b>")
        for ev in d7:
            lines.append(f"  {ev['date']}  {ev['related']}  {ev['content'][:30]}")
        lines.append("")
    if d30:
        lines.append(f"<b>📌 D-30 이내 ({len(d30)}건)</b>")
        for ev in d30[:10]:
            lines.append(f"  {ev['date']}  {ev['related']}  {ev['content'][:30]}")
    if not d7 and not d30:
        lines.append("D-30 이내 주요 이벤트 없음")
    return "\n".join(lines)


# ─── 리포트 생성 ──────────────────────────────────────────────

def build_report(results: dict) -> str:
    lines = [
        f"# Excel Ingest 리포트 — {TODAY}",
        f"> 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 소스: `raw/매일 엑셀넣을것/`",
        "",
    ]

    # 1. 추정이익변경
    r = results.get("추정이익변경", {})
    if "results" in r:
        lines += ["## 1. 추정이익변경 (Rating/TP)", ""]
        for sheet, label in [("Rating_Up","📈 레이팅 상향"), ("Rating_Down","📉 레이팅 하향"),
                              ("TP_Up","🎯 TP 상향"), ("TP_Down","⬇️ TP 하향")]:
            items = r["results"].get(sheet, [])
            if items:
                lines.append(f"### {label} ({len(items)}건)")
                lines.append("| 날짜 | 종목 | 증권사 | 의견변경 | TP변경 |")
                lines.append("|------|------|--------|---------|--------|")
                for it in items[:15]:
                    op = f"{it['op_old']}→{it['op_new']}" if it['op_old'] != it['op_new'] else it['op_new']
                    tp_old = fmt_num(it['tp_old'], "원") if it['tp_old'] else "—"
                    tp_new = fmt_num(it['tp_new'], "원") if it['tp_new'] else "—"
                    tp_str = f"{tp_old}→{tp_new}" if it['tp_old'] else tp_new
                    lines.append(f"| {it['date']} | {it['name']} | {it['brokerage']} | {op} | {tp_str} |")
                lines.append("")
        if r.get("updated"):
            lines += ["**wiki 업데이트:**"] + [f"- {u}" for u in r["updated"]] + [""]

    # 2. 컨센움직임
    r = results.get("컨센움직임", {})
    if "results" in r:
        lines += ["## 2. 컨센움직임 / 서프·쇼크", ""]
        for sheet, label in [("컨센상향","📈 컨센상향"), ("서프라이즈","🎉 어닝서프"),
                              ("컨센하향","📉 컨센하향"), ("쇼크","💥 어닝쇼크")]:
            items = r["results"].get(sheet, [])
            if items:
                lines.append(f"### {label} ({len(items)}종목)")
                lines.append("| 종목 | 섹터 | 1M수익률 | 2026E영업이익(억) | 1M변화 |")
                lines.append("|------|------|---------|----------------|--------|")
                for it in items[:10]:
                    op_val = it.get('op_2026')
                    op_str = fmt_num(op_val, "") if op_val else "—"
                    lines.append(f"| {it['name']} | {it.get('sector','—')} | {fmt_pct(it.get('ret_1m'))} | {op_str} | {fmt_pct(it.get('op_chg1m'))} |")
                lines.append("")

    # 3. 수출
    r = results.get("수출", {})
    if "results" in r:
        lines += ["## 3. 핵심 수출 신호", ""]
        lines.append("| 품목 | 섹터 | 최근월 | 물량YoY | 판가YoY | 신호 |")
        lines.append("|------|------|--------|--------|--------|------|")
        for sheet_name, data in r["results"].items():
            if not data.get("monthly"): continue
            latest = data["monthly"][-1]
            lines.append(
                f"| {sheet_name} | {data['sector']} | {latest['month'][:7]} "
                f"| {fmt_pct(latest['yoy_vol'])} | {fmt_pct(latest['price_yoy'])} "
                f"| {data['signal']} |"
            )
        lines.append("")

    # 4. 유동성체크
    r = results.get("유동성", {})
    if "results" in r:
        lines += ["## 4. 컨센 신고가 TOP10", ""]
        items = r["results"].get("컨센신고가", [])[:10]
        if items:
            lines.append("| 순위 | 종목 | 최근발표일 | 증권사 |")
            lines.append("|------|------|-----------|--------|")
            for it in items:
                lines.append(f"| {it['rank']} | {it['name']} | {it['latest_date']} | {it['brokerage']} |")
        lines.append("")

    def _빈집_섹션(lines, r, title_prefix, sec_num):
        if not (r.get("빈집_A") or r.get("빈집_B")):
            return
        lines += [f"## {sec_num}. {title_prefix} 수급 빈집 탐지", ""]
        if r.get("date"):
            lines.append(f"> 기준일: {r['date']} | {r.get('note', '')}")
            lines.append("")
        for grade, label in [("빈집_A", "🏚️ 빈집A — 완전빈집 (하위 10%)"),
                              ("빈집_B", "🏠 빈집B — 반빈집 (하위 10~25%)")]:
            items = r.get(grade, [])
            if not items: continue
            lines.append(f"### {label} {len(items)}종목")
            lines.append("| 종목 | 오실레이터 | 퍼센타일 | 방향 |")
            lines.append("|------|----------|---------|------|")
            for it in items[:20]:
                lines.append(f"| {it['name']} | {it['osc']:+.6f} | 하위{it['pct']:.1f}% | {it.get('trend','')} |")
            lines.append("")

    _빈집_섹션(lines, results.get("수급", {}),       "대형주(700)",       "5")
    _빈집_섹션(lines, results.get("중소형주수급", {}), "중소형주(700-1400)", "6")

    # 7. 가속화모멘텀 (Q열 — 주당순이익 1개+)
    r = results.get("가속화모멘텀", {})
    if "results" in r:
        q_items = r["results"].get("주당순이익1개+", [])
        if q_items:
            lines += ["## 7. 가속화모멘텀 TOP30 (주당순이익 1개+)", ""]
            lines.append("| 순위 | 종목 | 모멘텀스코어 | 이익변화24 | 이익변화25 |")
            lines.append("|------|------|------------|---------|---------|")
            for i, it in enumerate(q_items[:30], 1):
                c24 = f"{it['chg24']:+.0%}" if it['chg24'] is not None else "—"
                c25 = f"{it['chg25']:+.0%}" if it['chg25'] is not None else "—"
                lines.append(f"| {i} | {it['name']} | {it['score']:.2f} | {c24} | {c25} |")
            lines.append("")

    # 8. RS 상대강도 TOP30
    r = results.get("RS", {})
    if r.get("top30"):
        lines += [f"## 8. RS 상대강도 TOP30 — {r.get('date','')}", ""]
        lines.append("| 종목 | RS_60d | RS_120d | RS_250d | norm평균 |")
        lines.append("|------|--------|---------|---------|---------|")
        for it in r["top30"]:
            r60  = f"{it.get('RS_60d',0):+.1f}%"  if 'RS_60d'  in it else "—"
            r120 = f"{it.get('RS_120d',0):+.1f}%" if 'RS_120d' in it else "—"
            r250 = f"{it.get('RS_250d',0):+.1f}%" if 'RS_250d' in it else "—"
            nav  = f"{it.get('norm_RS_avg',0):.1f}"
            lines.append(f"| {it['name']} | {r60} | {r120} | {r250} | {nav} |")
        lines.append("")

    # 9. 교차분석 탑픽
    교차_msg = _build_교차분석_tg(results)
    if 교차_msg:
        lines += ["## 9. 교차분석 탑픽 (수급빈집 × RS × 가속화)", ""]
        lines.append("```")
        lines.append(교차_msg.replace("<b>", "").replace("</b>", "").replace("<code>", "").replace("</code>", ""))
        lines.append("```")
        lines.append("")

    return "\n".join(lines)


# ─── 텔레그램 발송 ────────────────────────────────────────────────

def _load_env() -> dict:
    env_path = ROOT / ".env"
    cfg = {}
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                cfg[k.strip()] = v.strip()
    return cfg


def send_telegram(text: str) -> bool:
    import json as _json, urllib.request as _req, urllib.parse as _parse
    cfg     = _load_env()
    token   = cfg.get("BOT_TOKEN", "")
    chat_id = cfg.get("CHAT_ID", "")
    if not token or not chat_id:
        print("  ⚠️ .env BOT_TOKEN/CHAT_ID 없음 — 텔레그램 스킵")
        return False
    chunks = [text[i:i+4000] for i in range(0, len(text), 4000)]
    for idx, chunk in enumerate(chunks, 1):
        url  = f"https://api.telegram.org/bot{token}/sendMessage"
        data = _parse.urlencode({
            "chat_id": chat_id, "text": chunk,
            "parse_mode": "HTML", "disable_web_page_preview": "true",
        }).encode()
        try:
            with _req.urlopen(_req.Request(url, data=data), timeout=10) as r:
                res = _json.loads(r.read())
                if not res.get("ok"):
                    print(f"  ❌ TG 오류: {res}")
                    return False
        except Exception as e:
            print(f"  ❌ TG 실패: {e}")
            return False
    print(f"  ✅ 텔레그램 발송 ({len(chunks)}건)")
    return True


def _빈집_tg_블록(a_list: list) -> list:
    """빈집A 리스트 → 심화중 / 유턴중 분리 라인 반환"""
    심화 = [it for it in a_list if '심화' in it.get('trend', '')]
    유턴 = [it for it in a_list if '재진입' in it.get('trend', '')]
    lines = []
    if 심화:
        lines.append(f"<b>📉 빈집 심화중 — {len(심화)}종목</b>")
        for it in 심화:
            lines.append(f"  {it['name']}  <code>{it['osc']:+.6f}</code>  하위{it.get('pct',0):.0f}%")
        lines.append("")
    if 유턴:
        lines.append(f"<b>↩️ 빈집 유턴중 — {len(유턴)}종목</b>")
        for it in 유턴:
            lines.append(f"  {it['name']}  <code>{it['osc']:+.6f}</code>  하위{it.get('pct',0):.0f}%")
        lines.append("")
    return lines


def _build_중소형주_tg(r: dict) -> str:
    date_str = r.get("date", TODAY)
    total    = r.get("total", 0)
    a_list   = r.get("빈집_A", [])
    심화 = [it for it in a_list if '심화' in it.get('trend', '')]
    유턴 = [it for it in a_list if '재진입' in it.get('trend', '')]
    lines = [
        f"🏚️ <b>중소형주(700-1400) 수급 빈집</b> — {date_str}",
        f"전체 {total}종목 | 심화중: {len(심화)}개  유턴중: {len(유턴)}개",
        "",
    ]
    lines += _빈집_tg_블록(a_list)
    return "\n".join(lines)


def build_빈집_tg(r: dict, results: dict) -> str:
    date_str = r.get("date", TODAY)
    total    = r.get("total", 0)
    a_list   = r.get("빈집_A", [])
    심화 = [it for it in a_list if '심화' in it.get('trend', '')]
    유턴 = [it for it in a_list if '재진입' in it.get('trend', '')]
    lines = [
        f"🏚️ <b>대형주(700) 수급 빈집</b> — {date_str}",
        f"전체 {total}종목 | 심화중: {len(심화)}개  유턴중: {len(유턴)}개",
        "",
    ]
    lines += _빈집_tg_블록(a_list)

    # ── 레이팅 상향 TOP5
    rating_up = results.get("추정이익변경", {}).get("results", {}).get("Rating_Up", [])
    if rating_up:
        lines.append("<b>📈 레이팅 상향 TOP5</b>")
        for it in rating_up[:5]:
            op = f"{it['op_old']}→{it['op_new']}"
            tp = fmt_num(it['tp_new'], "원") if it['tp_new'] else "—"
            lines.append(f"  · {it['name']} ({it['brokerage']}) {op} TP {tp}")
        lines.append("")

    # ── 컨센 신고가 TOP5
    high_list = results.get("유동성", {}).get("results", {}).get("컨센신고가", [])
    if high_list:
        lines.append("<b>🎯 컨센 신고가 TOP5</b>")
        for it in high_list[:5]:
            lines.append(f"  {it['rank']}. {it['name']} ({it['brokerage']})")
        lines.append("")

    return "\n".join(lines)


def _build_가속화_tg(r: dict) -> str:
    items = r.get("results", {}).get("주당순이익1개+", [])
    lines = [
        f"⚡ <b>가속화모멘텀 TOP30</b> (주당순이익 1개+)",
        f"총 {len(items)}종목 중 상위 30",
        "",
    ]
    for i, it in enumerate(items[:30], 1):
        star = "⭐" if it["score"] >= 1.0 else "  "
        lines.append(f"{star}{i:2d}. {it['name']}  <code>{it['score']:.2f}</code>")
    return "\n".join(lines)


def _build_rs_tg(r: dict) -> str:
    date_str = r.get("date", TODAY)
    items    = r.get("top30", [])
    lines = [
        f"📈 <b>RS 상대강도 TOP30</b> — {date_str}",
        f"전체 {r.get('total',0)}종목 | norm 평균 퍼센타일 기준",
        "",
        "<b>종목  RS60  RS120  RS250  norm평균</b>",
    ]
    for it in items[:30]:
        r60  = f"{it.get('RS_60d', 0):+.0f}"  if 'RS_60d'  in it else "—"
        r120 = f"{it.get('RS_120d',0):+.0f}" if 'RS_120d' in it else "—"
        r250 = f"{it.get('RS_250d',0):+.0f}" if 'RS_250d' in it else "—"
        nav  = f"{it.get('norm_RS_avg',0):.0f}"
        lines.append(f"{it['name']}  {r60}  {r120}  {r250}  <code>{nav}</code>")
    return "\n".join(lines)


def _build_etf_rs_tg(r: dict) -> str:
    date_str = r.get("date", TODAY)
    rs70     = r.get("rs70+", [])
    top20    = r.get("top20", [])
    lines = [
        f"📈 <b>한국ETF 상대강도 (Mansfield RS)</b> — {date_str}",
        f"전체 {r.get('total', 0)}개 ETF | RS≥70: {len(rs70)}개",
        "",
    ]
    if rs70:
        lines.append(f"<b>🔴 RS ≥ 70 ETF ({len(rs70)}개)</b>")
        for it in rs70[:15]:
            lines.append(f"  {it['name']}  <code>{it.get('norm_RS_avg', 0):.0f}</code>")
        lines.append("")
    if top20:
        lines.append("<b>▶ Top 10</b>")
        for i, it in enumerate(top20[:10], 1):
            lines.append(f"  {i:2d}. {it['name']}  <code>{it.get('norm_RS_avg', 0):.1f}</code>")
    return "\n".join(lines)


def _build_투자아이디어_tg(r: dict) -> str:
    ideas    = r.get("ideas", [])
    quarters = r.get("quarters", "")
    lines = [
        f"💡 <b>투자아이디어 트리거 ({quarters})</b> — {TODAY}",
        f"총 {len(ideas)}개 트리거",
        "",
    ]
    if not ideas:
        lines.append("해당 분기 트리거 없음")
        return "\n".join(lines)
    by_company: dict = {}
    for it in ideas:
        by_company.setdefault(it["company"], []).append(it)
    for company, items in list(by_company.items())[:15]:
        lines.append(f"<b>▶ {company}</b>")
        for it in items[:3]:
            lines.append(f"  [{it['quarter']}] {it['momentum']}")
        lines.append("")
    return "\n".join(lines)


def _build_컨센_탑픽_tg(results: dict) -> str:
    """수급빈집 × 컨센(리레이팅·신고가·상향·TP·가속화) 종합 스코어링 → 탑픽"""

    # ── 수급빈집 ──────────────────────────────────────────────────
    빈집_A: set[str] = set()
    빈집_B: set[str] = set()
    빈집_detail: dict[str, dict] = {}
    for key, label in [("수급", "대형주"), ("중소형주수급", "중소형주")]:
        r = results.get(key, {})
        for it in r.get("빈집_A", []):
            빈집_A.add(it["name"])
            빈집_detail[it["name"]] = {"pct": it.get("pct", 0), "trend": it.get("trend", "")}
        for it in r.get("빈집_B", []):
            if it["name"] not in 빈집_detail:
                빈집_B.add(it["name"])
                빈집_detail[it["name"]] = {"pct": it.get("pct", 0), "trend": it.get("trend", "")}

    # ── 리레이팅 (Rating_Up) ──────────────────────────────────────
    r_추정 = results.get("추정이익변경", {})
    리레이팅 = {it["name"] for it in r_추정.get("results", {}).get("Rating_Up", [])}
    리레이팅_detail = {it["name"]: f"{it['op_old']}→{it['op_new']} ({it['brokerage']})"
                      for it in r_추정.get("results", {}).get("Rating_Up", [])}

    # ── TP 상향 ───────────────────────────────────────────────────
    tp_up = {it["name"] for it in r_추정.get("results", {}).get("TP_Up", [])}

    # ── 컨센신고가 TOP10 ──────────────────────────────────────────
    r_유동 = results.get("유동성", {})
    신고가 = {it["name"] for it in r_유동.get("results", {}).get("컨센신고가", [])[:10]}

    # ── 컨센상향 (방향 양수) ──────────────────────────────────────
    r_컨센 = results.get("컨센움직임", {})
    컨센상향 = {
        it["name"] for it in r_컨센.get("results", {}).get("컨센상향", [])
        if isinstance(it.get("csen_chg"), (int, float)) and it["csen_chg"] > 0
    }
    컨센_detail = {
        it["name"]: it.get("csen_chg", 0)
        for it in r_컨센.get("results", {}).get("컨센상향", [])
    }

    # ── 가속화모멘텀 TOP30 ────────────────────────────────────────
    r_가속 = results.get("가속화모멘텀", {})
    가속화 = {it["name"] for it in r_가속.get("results", {}).get("주당순이익1개+", [])[:30]}

    # ── 전체 종목 스코어링 ────────────────────────────────────────
    # 수급빈집 없으면 후순위 처리 (0.5점 감점 아닌 별도 섹션)
    all_names = (빈집_A | 빈집_B | 리레이팅 | 신고가 | 컨센상향 | tp_up | 가속화)

    scored = []
    for name in all_names:
        pts = []
        score = 0
        if name in 빈집_A:   score += 2; pts.append("빈집A(+2)")
        if name in 빈집_B:   score += 1; pts.append("빈집B(+1)")
        if name in 리레이팅: score += 2; pts.append("리레이팅(+2)")
        if name in 신고가:   score += 2; pts.append("컨센신고가(+2)")
        if name in 가속화:   score += 1; pts.append("가속화(+1)")
        if name in 컨센상향: score += 1; pts.append("컨센상향(+1)")
        if name in tp_up:    score += 1; pts.append("TP상향(+1)")
        has_빈집 = name in 빈집_A or name in 빈집_B
        scored.append({
            "name":     name,
            "score":    score,
            "has_빈집": has_빈집,
            "pts":      pts,
            "빈집_info": 빈집_detail.get(name, {}),
            "컨센_chg":  컨센_detail.get(name),
            "rating":   리레이팅_detail.get(name),
        })

    # 수급 상태별 분류
    재진입 = sorted(
        [s for s in scored if s["has_빈집"] and "재진입" in s["빈집_info"].get("trend", "")],
        key=lambda x: x["score"], reverse=True
    )
    빠지는중 = sorted(
        [s for s in scored if s["has_빈집"] and "재진입" not in s["빈집_info"].get("trend", "")],
        key=lambda x: x["score"], reverse=True
    )

    # 과매수 상승/하락 — 오실레이터 결과에서 직접
    꽉참_상승_names: set[str] = set()
    꽉참_턴_names: set[str] = set()
    for key in ["수급", "중소형주수급"]:
        for it in results.get(key, {}).get("과매수_상승", []):
            꽉참_상승_names.add(it["name"])
        for it in results.get(key, {}).get("과매수_하락", []):
            꽉참_턴_names.add(it["name"])

    꽉참_상승중 = sorted(
        [s for s in scored if not s["has_빈집"] and s["name"] in 꽉참_상승_names],
        key=lambda x: x["score"], reverse=True
    )
    꽉참_턴중 = sorted(
        [s for s in scored if not s["has_빈집"] and s["name"] in 꽉참_턴_names],
        key=lambda x: x["score"], reverse=True
    )

    def _fmt(s):
        pts_clean = [p.split("(")[0] for p in s["pts"] if "빈집" not in p]
        return f"  <b>{s['name']}</b>  {s['score']}점  {' · '.join(pts_clean)}"

    lines = [
        f"🎯 <b>컨센 종합 탑픽</b> — {TODAY}",
        f"수급 × 리레이팅 × 컨센신고가 × 상향 × TP × 가속화 (10점 만점)",
        "",
    ]

    if 재진입:
        lines.append(f"<b>🔴 수급 바닥 턴중 ({len(재진입)}종목)</b>")
        for s in 재진입[:10]:
            lines.append(_fmt(s))
            if s["rating"]:
                lines.append(f"    └ 리레이팅: {s['rating']}")
        lines.append("")

    if 빠지는중:
        lines.append(f"<b>🟡 수급 바닥 하락중 ({len(빠지는중)}종목, 상위 5개)</b>")
        for s in 빠지는중[:5]:
            lines.append(_fmt(s))
        lines.append("")

    if 꽉참_상승중:
        lines.append(f"<b>⚪ 수급 꽉참 상승중 ({len(꽉참_상승중)}종목, 상위 5개)</b>")
        for s in 꽉참_상승중[:5]:
            pts_clean = [p.split("(")[0] for p in s["pts"]]
            lines.append(f"  {s['name']}  {s['score']}점  {' · '.join(pts_clean)}")
        lines.append("")

    if 꽉참_턴중:
        lines.append(f"<b>🔵 수급 꽉참 턴중 ({len(꽉참_턴중)}종목, 상위 5개)</b>")
        for s in 꽉참_턴중[:5]:
            pts_clean = [p.split("(")[0] for p in s["pts"]]
            lines.append(f"  {s['name']}  {s['score']}점  {' · '.join(pts_clean)}")
        lines.append("")

    # 빈집/과매수 없지만 컨센 신호 2개+ 종목
    컨센단독 = sorted(
        [s for s in scored
         if not s["has_빈집"]
         and s["name"] not in 꽉참_상승_names
         and s["name"] not in 꽉참_턴_names
         and s["score"] >= 2],
        key=lambda x: x["score"], reverse=True
    )
    if 컨센단독:
        lines.append(f"<b>⚡ 컨센 단독 후보 ({len(컨센단독)}종목, 상위 10개)</b>")
        lines.append("<i>수급 중립 — 컨센·리레이팅·TP 신호만</i>")
        for s in 컨센단독[:10]:
            pts_clean = [p.split("(")[0] for p in s["pts"]]
            line = f"  <b>{s['name']}</b>  {s['score']}점  {' · '.join(pts_clean)}"
            if s["rating"]:
                line += f"\n    └ {s['rating']}"
            lines.append(line)

    if not any([재진입, 빠지는중, 꽉참_상승중, 꽉참_턴중, 컨센단독]):
        lines.append("후보 없음")

    return "\n".join(lines)


def _build_교차분석_tg(results: dict) -> str:
    """수급빈집 × RS × 가속화모멘텀 교집합 → 탑픽 후보 텔레그램 메시지"""
    # 1. 수급빈집 집합 (대형주 A, 중소형주 A/B)
    빈집_names: set[str] = set()
    빈집_detail: dict[str, dict] = {}

    for key, label in [("수급", "대형주"), ("중소형주수급", "중소형주")]:
        r = results.get(key, {})
        for it in r.get("빈집_A", []):
            빈집_names.add(it["name"])
            빈집_detail[it["name"]] = {
                "grade": f"{label}A",
                "pct":   it.get("pct", 0),
                "trend": it.get("trend", ""),
            }
        for it in r.get("빈집_B", []):
            if it["name"] not in 빈집_detail:
                빈집_names.add(it["name"])
                빈집_detail[it["name"]] = {
                    "grade": f"{label}B",
                    "pct":   it.get("pct", 0),
                    "trend": it.get("trend", ""),
                }

    # 2. RS top30 집합
    r_rs = results.get("RS", {})
    rs_names  = {it["name"] for it in r_rs.get("top30", [])}
    rs_score  = {it["name"]: it.get("norm_RS_avg", 0) for it in r_rs.get("top30", [])}

    # 3. 가속화모멘텀 top30 집합 (주당순이익1개+)
    r_가속  = results.get("가속화모멘텀", {})
    가속_items = r_가속.get("results", {}).get("주당순이익1개+", [])[:30]
    가속_names = {it["name"] for it in 가속_items}
    가속_score = {it["name"]: it.get("score", 0) for it in 가속_items}

    if not 빈집_names:
        return ""

    triple     = 빈집_names & rs_names & 가속_names
    double_rs  = (빈집_names & rs_names)  - triple
    double_가속 = (빈집_names & 가속_names) - triple

    lines = [
        f"🎯 <b>교차분석 탑픽 후보</b> — {TODAY}",
        f"빈집({len(빈집_names)}) × RS({len(rs_names)}) × 가속화({len(가속_names)})",
        "",
    ]

    def _row(name: str) -> str:
        d = 빈집_detail.get(name, {})
        rs = f"  RS:{rs_score[name]:.0f}" if name in rs_score else ""
        가속 = f"  가속:{가속_score[name]:.2f}" if name in 가속_score else ""
        trend = f"  {d.get('trend','')}" if d.get("trend") else ""
        return f"  {name}  빈집{d.get('grade','')}(하위{d.get('pct',0):.0f}%){rs}{가속}{trend}"

    if triple:
        lines.append(f"<b>🔴 3중 교집합 ({len(triple)}종목)</b>")
        for name in sorted(triple):
            lines.append(_row(name))
        lines.append("")

    if double_rs:
        lines.append(f"<b>🟠 빈집×RS ({len(double_rs)}종목)</b>")
        for name in sorted(double_rs):
            lines.append(_row(name))
        lines.append("")

    if double_가속:
        lines.append(f"<b>🟡 빈집×가속화 ({len(double_가속)}종목)</b>")
        for name in sorted(double_가속):
            lines.append(_row(name))
        lines.append("")

    if not triple and not double_rs and not double_가속:
        lines.append("교집합 없음 — 빈집 단독 후보")
        빈집_only = sorted(빈집_names, key=lambda n: 빈집_detail[n].get("pct", 100))[:10]
        for name in 빈집_only:
            lines.append(_row(name))

    return "\n".join(lines)


# ─── log.md 기록 ───────────────────────────────────────────────

def append_log(summary: str):
    if not LOG_FILE.exists(): return
    existing = LOG_FILE.read_text(encoding="utf-8")
    entry = f"- {TODAY} — Excel ingest 완료: {summary}\n"
    LOG_FILE.write_text(entry + existing, encoding="utf-8")


# ─── 메인 ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="wiki 수정 없이 리포트만 생성")
    parser.add_argument("--only", help="특정 파서만 실행 (추정이익|컨센|수출|유동성|수급)")
    args = parser.parse_args()

    dry_run = args.dry_run
    only = args.only

    print(f"{'[DRY-RUN] ' if dry_run else ''}Excel Ingest 시작 — {TODAY}")
    print(f"소스 폴더: {EXCEL_DIR}")
    print()

    results = {}

    parsers = {
        "추정이익": ("추정이익변경", lambda: parse_추정이익변경(dry_run)),
        "컨센":     ("컨센움직임",   lambda: parse_컨센움직임(dry_run)),
        "수출":     ("수출",         lambda: parse_수출정리(dry_run)),
        "유동성":   ("유동성",       lambda: parse_유동성체크(dry_run)),
        "수급":     ("수급",         lambda: parse_수급오실레이터(dry_run)),
        "중소형주":  ("중소형주수급",  lambda: parse_중소형주오실레이터(dry_run)),
        "가속화":    ("가속화모멘텀",  lambda: parse_가속화모멘텀(dry_run)),
        "RS":        ("RS",           lambda: parse_rs(dry_run)),
        "쏠림":      ("쏠림지수",      lambda: parse_쏠림지수(dry_run)),
        "ETF":       ("액티브ETF",     lambda: parse_액티브ETF(dry_run)),
        "일정":      ("일정",          lambda: parse_일정(dry_run)),
        "ETFRS":     ("한국ETF_RS",    lambda: parse_한국ETF상대강도(dry_run)),
        "아이디어":  ("투자아이디어",  lambda: parse_투자아이디어(dry_run)),
    }

    for key, (label, fn) in parsers.items():
        if only and only not in key: continue
        print(f"  처리 중: {label}...")
        try:
            results[label] = fn()
            if "error" in results[label]:
                print(f"    ⚠️ {results[label]['error']}")
            else:
                print(f"    ✅ 완료")
        except Exception as e:
            print(f"    ❌ 오류: {e}")
            results[label] = {"error": str(e)}

    # 종목별 태린이 지표 스냅샷 (대시보드 종목 카드용)
    try:
        idx = build_stock_index(results)
        msg = f"  📦 taerini_stock.json — {idx['meta']['stock_count']}종목"
        if idx["meta"]["unmatched"]:
            msg += f" (미매칭 {len(idx['meta']['unmatched'])})"
        print(msg)
    except Exception as e:
        print(f"  ⚠️ build_stock_index 실패: {e}")

    # 리포트 저장
    report = build_report(results)
    report_path = ROOT / "raw" / f"ingest_report_{TODAY}.md"
    report_path.write_text(report, encoding="utf-8")
    print(f"\n📄 리포트: {report_path}")

    # log.md 기록 + 텔레그램 발송
    if not dry_run:
        summary_parts = []
        for label, data in results.items():
            if "error" not in data:
                summary_parts.append(label)
        append_log("·".join(summary_parts) + f" → {report_path.name}")

        # 컨센움직임 텔레그램 발송
        r_컨센 = results.get("컨센움직임", {})
        if r_컨센.get("results"):
            has_data = any(r_컨센["results"].get(k) for k in ["컨센상향", "서프라이즈", "쇼크"])
            if has_data:
                print("\n📲 [컨센움직임] 텔레그램 발송 중...")
                send_telegram(_build_컨센움직임_tg(r_컨센))

        # 대형주 수급빈집 텔레그램 발송
        r_수급 = results.get("수급", {})
        if r_수급.get("빈집_A") or r_수급.get("빈집_B"):
            print("\n📲 [대형주] 텔레그램 발송 중...")
            msg = build_빈집_tg(r_수급, results)
            send_telegram(msg)

        # 중소형주 빈집A 텔레그램 발송
        r_중소 = results.get("중소형주수급", {})
        if r_중소.get("빈집_A"):
            print("\n📲 [중소형주] 텔레그램 발송 중...")
            send_telegram(_build_중소형주_tg(r_중소))

        # 가속화모멘텀 텔레그램 발송
        r_가속 = results.get("가속화모멘텀", {})
        if r_가속.get("results"):
            print("\n📲 [가속화모멘텀] 텔레그램 발송 중...")
            send_telegram(_build_가속화_tg(r_가속))

        # RS 상대강도 텔레그램 발송
        r_rs = results.get("RS", {})
        if r_rs.get("top30"):
            print("\n📲 [RS] 텔레그램 발송 중...")
            send_telegram(_build_rs_tg(r_rs))

        # 업종 쏠림지수 텔레그램 발송
        r_쏠림 = results.get("쏠림지수", {})
        if r_쏠림.get("idx") or r_쏠림.get("macd"):
            print("\n📲 [쏠림지수] 텔레그램 발송 중...")
            send_telegram(_build_쏠림_tg(r_쏠림))

        # 액티브ETF 비중 변화 텔레그램 발송
        r_etf = results.get("액티브ETF", {})
        if r_etf.get("increase") or r_etf.get("decrease"):
            print("\n📲 [액티브ETF] 텔레그램 발송 중...")
            send_telegram(_build_액티브ETF_tg(r_etf))

        # 이벤트 일정 텔레그램 발송 (D-7 이내 있을 때만)
        r_일정 = results.get("일정", {})
        if r_일정.get("d7") or r_일정.get("d30"):
            print("\n📲 [일정] 텔레그램 발송 중...")
            send_telegram(_build_일정_tg(r_일정))

        # 교차분석 텔레그램 발송 (수급빈집 있을 때)
        교차_msg = _build_교차분석_tg(results)
        if 교차_msg:
            print("\n📲 [교차분석] 텔레그램 발송 중...")
            send_telegram(교차_msg)

        # 컨센 종합 탑픽 발송
        컨센_탑픽_msg = _build_컨센_탑픽_tg(results)
        if 컨센_탑픽_msg:
            print("\n📲 [컨센 종합 탑픽] 텔레그램 발송 중...")
            send_telegram(컨센_탑픽_msg)

        # 한국ETF RS 텔레그램 발송
        r_etf_rs = results.get("한국ETF_RS", {})
        if r_etf_rs.get("top20"):
            print("\n📲 [한국ETF RS] 텔레그램 발송 중...")
            send_telegram(_build_etf_rs_tg(r_etf_rs))

        # 투자아이디어 텔레그램 발송 (트리거 있을 때만)
        r_아이디어 = results.get("투자아이디어", {})
        if r_아이디어.get("ideas"):
            print("\n📲 [투자아이디어] 텔레그램 발송 중...")
            send_telegram(_build_투자아이디어_tg(r_아이디어))

        # 업종 오실레이터 실행 (별도 스크립트)
        업종_script = ROOT / "scripts" / "scan_업종오실레이터.py"
        if 업종_script.exists():
            print("\n⚙️ 업종 오실레이터 실행 중...")
            subprocess.run([sys.executable, str(업종_script), "--tg"], check=False)

        # 소라티노 ETF 랭킹 실행 (별도 스크립트)
        sortino_script = ROOT / "scripts" / "scan_sortino.py"
        if sortino_script.exists():
            print("\n⚙️ 소라티노 ETF 실행 중...")
            subprocess.run([sys.executable, str(sortino_script), "--tg"], check=False)

        # 추정이익 카드뉴스 자동 생성 (오늘 TP 상향 TOP3)
        r_추정 = results.get("추정이익변경", {})
        tp_up = r_추정.get("results", {}).get("TP_Up", [])
        if tp_up:
            seen: dict[str, dict] = {}
            for it in tp_up:
                if it["date"] == TODAY and it["name"] not in seen:
                    seen[it["name"]] = it
            top_names = list(seen.keys())[:3]
            if top_names:
                card_script = ROOT / "scripts" / "viz_card.py"
                if card_script.exists():
                    print(f"\n카드뉴스 자동 생성: {', '.join(top_names)}")
                    subprocess.run(
                        [sys.executable, str(card_script)] + top_names + ["--tg"],
                        check=False
                    )

    print("\n" + "="*50)
    print(report[:2000])

if __name__ == "__main__":
    main()
